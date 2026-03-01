import os
import csv
import json
import time
import math
import queue
import serial
import socket
import struct
import threading
import pandas as pd
import numpy as np
import tkinter as tk
import minimalmodbus
from tkinter import ttk, TclError
from threading import Lock
from datetime import datetime
from collections import deque
import serial.tools.list_ports
from tkinter import filedialog, messagebox
import matplotlib.pyplot as plt
from tkinter import scrolledtext
from matplotlib.image import imread
from dataclasses import dataclass
from matplotlib.animation import FuncAnimation
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk



# === Globals ===
HOST, PORT = '0.0.0.0', 52740
client_socket = None
shutdown_flag = False
data_lock = Lock()
acc_fsr = 16.0  # ±16g full scale range
gyro_fsr = 2000.0  # ±2000dps full scale range

last_position = None
last_time = None
total_distance = 0.0
total_time = 0.0

ser_uwb = None
uwb_lock = Lock()
anchor_instrument = None # MODBUS instrument for anchor communication
anchor_lock = Lock()

ANCHOR_MAX_COUNT = 16
IMU_DATA_ACC_EN = 0x0001
IMU_DATA_GYRO_EN = 0x0002
IMU_DATA_EULER_EN = 0x0004
IMU_DATA_TEMP_EN = 0x0008
IMU_DATA_Q_EN = 0x0010
IMU_DATA_MAGN_EN = 0x0020

log_queue = queue.Queue()
data_queue = queue.Queue()

map_multiple = 1.0  # Map scaling factor
map_origin_x = 0    # Map origin X coordinate
map_origin_y = 0    # Map origin Y coordinate
max_tags = 100      # Maximum number of tags to track

MODBUS_START_POSITIONING = bytes([0x01, 0x10, 0x00, 0x3B, 
                                  0x00, 0x01, 0x02, 0x00, 
                                  0x04, 0xA3, 0x18])
MODBUS_STOP_POSITIONING = bytes([0x01, 0x10, 0x00, 0x3B, 
                                 0x00, 0x01, 0x02, 0x00, 
                                 0x00, 0xA2, 0xDB])
BAUD_RATE_DELAYS = {
    9600: 0.013,
    19200: 0.010,
    38400: 0.004,
    57600: 0.003,
    115200: 0.002,
    230400: 0.001,
    460800: 0.001,
    921600: 0.001
}

MODBUS_ADDR_POSITIONING_CONTROL = 59 # Register address for starting/stopping
MODBUS_VAL_START_POSITIONING = 4      # Value to start (matches 0x04 in your bytes)
MODBUS_VAL_STOP_POSITIONING = 0       # Value to stop (matches 0x00 in your bytes)

MODBUS_ADDR_MAP = {
    'MODBUS-ID': 0x01,        # Typically a device parameter, not a register
    'Kalman-Q': 0x10,
    'Kalman-R': 0x11,
    'Antenna Delay': 0x12,
    'Device Type': 0x14,
    'Device ID': 0x15,
    'Positioning Mode': 0x16,
    'Ranging Mode': 0x17,
    'UWB Channel': 0x18,
    'Data Rate': 0x19,
    'Network ID': 0x20
}

ANC_PROTOCAL_RTLS = 0      # Base station positioning data flag
ANC_PROTOCAL_DIST = 1      # Base station ranging data flag
ANC_PROTOCAL_RXDIAG = 2    # Base station reception info flag
ANC_PROTOCAL_TIMESTAMP = 3 # Base station timestamp info flag

TAG_OUTPUT_DIST = 0
TAG_OUTPUT_RTLS = 1

ANCHOR_COMMANDS = {
    'start_pos': {'cmd': 'pc', 'desc': 'Start positioning cycle'},
    'stop_pos': {'cmd': ' \r\n', 'desc': 'Stop positioning'},
    'reset': {'cmd': 'reset', 'desc': 'Reset anchor'},
    'set_major': {'cmd': 'nvm 0x80 1', 'desc': 'Set as major anchor'},
    'get_info': {'cmd': 'si', 'desc': 'Get system information'},
    'set_rate': {'cmd': 'nvm 0x82 {rate}', 'desc': 'Set update rate (ms)'},
    'set_network': {'cmd': 'nvm 0x81 {id}', 'desc': 'Set network ID (1-65535)'}
}

anchor_state = {
    'connected': False,
    'positioning': False,
    'network_id': None,
    'update_rate': None,
    'last_response': None
}
anchor_state_lock = Lock()

# Data buffers for plotting
max_len = 100
x_data = deque(maxlen=max_len)
y_data = deque(maxlen=max_len)
angle_data = deque(maxlen=max_len)
time_stamps = deque(maxlen=max_len)
z_data = deque(maxlen=max_len)

@dataclass
class UWBTag:
    def __init__(self, id):
        self.id = str(id)
        self.x = 0.0
        self.y = 0.0
        self.z = 0.0
        self.angle = 0.0
        self.velocity = 0.0
        self.distance_traveled = 0.0
        self.last_update = 0.0
        self.status = "Inactive"
        self.channel = 2  # Default channel
        self.movement_history = deque(maxlen=100)  # Stores (x, y, z, angle, timestamp)
        self.lock = threading.Lock()
        
        # IMU data attributes
        self.acc_x = 0.0
        self.acc_y = 0.0
        self.acc_z = 0.0
        self.gyro_x = 0.0
        self.gyro_y = 0.0
        self.gyro_z = 0.0
        self.rotation_x = 0.0
        self.rotation_y = 0.0
        self.rotation_z = 0.0
        self.temperature = 0.0
        self.q0 = 0.0
        self.q1 = 0.0
        self.q2 = 0.0
        self.q3 = 0.0
        self.magn_x = 0.0
        self.magn_y = 0.0
        self.magn_z = 0.0
        self.is_get_newdata = False
    
    def update_imu_data(self, imu_data):
        """Update IMU data fields"""
        with self.lock:
            if 'acc_x' in imu_data: self.acc_x = imu_data['acc_x']
            if 'acc_y' in imu_data: self.acc_y = imu_data['acc_y']
            if 'acc_z' in imu_data: self.acc_z = imu_data['acc_z']
            if 'gyro_x' in imu_data: self.gyro_x = imu_data['gyro_x']
            if 'gyro_y' in imu_data: self.gyro_y = imu_data['gyro_y']
            if 'gyro_z' in imu_data: self.gyro_z = imu_data['gyro_z']
            if 'rotation_x' in imu_data: self.rotation_x = imu_data['rotation_x']
            if 'rotation_y' in imu_data: self.rotation_y = imu_data['rotation_y']
            if 'rotation_z' in imu_data: self.rotation_z = imu_data['rotation_z']
            if 'temperature' in imu_data: self.temperature = imu_data['temperature']
            if 'q0' in imu_data: self.q0 = imu_data['q0']
            if 'q1' in imu_data: self.q1 = imu_data['q1']
            if 'q2' in imu_data: self.q2 = imu_data['q2']
            if 'q3' in imu_data: self.q3 = imu_data['q3']
            if 'magn_x' in imu_data: self.magn_x = imu_data['magn_x']
            if 'magn_y' in imu_data: self.magn_y = imu_data['magn_y']
            if 'magn_z' in imu_data: self.magn_z = imu_data['magn_z']
            if 'is_get_newdata' in imu_data: self.is_get_newdata = imu_data['is_get_newdata']
            
    def update_position(self, x, y, z, angle=None, channel=None):
        """Thread-safe position update"""
        with self.lock:
            self.x = float(x)
            self.y = float(y)
            self.z = float(z)
            if angle is not None:
                self.angle = float(angle)
            if channel is not None:
                self.channel = int(channel)
            self.last_update = time.time()
            
    def get_position(self):
        """Thread-safe position access"""
        with self.lock:
            return (self.x, self.y, self.z, self.angle)
            
    def calculate_velocity(self, new_x, new_y, new_z):
        """Calculate velocity from new position"""
        with self.lock:
            if len(self.movement_history) > 0:
                prev_x, prev_y, prev_z, _, prev_time = self.movement_history[-1]
                dt = time.time() - prev_time
                
                if dt > 0:
                    dx = new_x - prev_x
                    dy = new_y - prev_y
                    dz = new_z - prev_z
                    distance = (dx**2 + dy**2 + dz**2) ** 0.5
                    
                    self.velocity = distance / dt
                    self.distance_traveled += distance
                    # Update status based on movement
                    self.status = "Moving" if self.velocity > 0.1 else "Stationary"
    
class TagManager:
    def __init__(self):
        self.tags = {}
        self.active_tag = None
        self.active_tag_id = "0"
                
        self.tag_colors = {
            'default': (255, 255, 255),  # White
            '0': (255, 0, 0),     # Red
            '1': (0, 255, 0),     # Green
            '2': (0, 0, 255),     # Blue
            '3': (255, 255, 0),   # Yellow
            '4': (255, 0, 255),   # Magenta
            '5': (0, 255, 255),   # Cyan
            '6': (255, 128, 0),   # Orange
            '7': (128, 0, 128)    # Purple
        }
        
        # PG-specific data structures
        self.last_cal_data = {
            'Tag_ID': 0,
            'x': 0,
            'y': 0,
            'z': 0,
            'Cal_Flag': 0,
            'Dist': [0] * 16,
            'Time_ts': [0] * 6
        }
        
        self.rx_diag = {
            'Max_noise': 0,
            'Std_noise': 0,
            'Fp_amp1': 0,
            'Fp_amp2': 0,
            'Fp_amp3': 0,
            'Max_growthCIR': 0,
            'Rx_preambleCount': 0,
            'Fp': 0,
            'Fp_power': 0.0,
            'Rx_power': 0.0
        }
        
        self.help_texts = {
            'Ranging Mode': (
                "Ranging Mode Description\n"
                "<Usage> If using HDS-TWR, the rate is fixed at 6.8Mbps and all anchors/tags must use the same ranging mode!\n"
                "<Default Setting> HDS-TWR\n"
                "<Activation> Takes effect immediately after configuration is loaded"
                ),
            'UWB Channel': (
                "UWB Channel\n"
                "<Usage> The wireless channel for DW1000 data transmission\n"
                "Different channels in the same area will cause interference\n"
                "All devices must use the same channel. Channel 2 is most commonly used with the longest range\n"
                "<Default Setting> 2\n"
                "<Activation> Takes effect immediately after configuration is loaded"
                ),
            '3D View': (
                "Dynamic 3D View Configuration Guide\n"
                "<Usage>\n"
                "<X-axis Range> Sets the real-world range represented by the x-axis (blue)\n"
                "<Y-axis Range> Sets the real-world range represented by the y-axis (green)\n"
                "<Z-axis Range> Sets the real-world range represented by the z-axis (red)\n"
                "<X-axis Step> Gridline intervals for x-axis (must be smaller than x-axis range)\n"
                "<Y-axis Step> Gridline intervals for y-axis (must be smaller than y-axis range)\n"
                "<Z-axis Step> Gridline intervals for z-axis (must be smaller than z-axis range)\n"
                "<Trail Length> Maximum length of tag movement trails\n"
                "<Apply Changes> Click 'Update Configuration' to apply settings\n"
                "<View Navigation>\n"
                "Left-click+drag to rotate view | Mouse wheel to zoom\n"
                "W/S/A/D keys move camera forward/back/left/right\n"
                "Z/C keys move camera up/down\n"
                "View adjustments only work during 3D positioning operation"
                ),
            'Hardware Calculation': (
                "Calculation Mode Selection\n"
                "<Usage> Choose between hardware/software calculation\n"
                "<Control>\n"
                "Checked = Hardware calculation | Unchecked = Software calculation\n"
                "Requires writing configuration to take effect\n"
                "In software mode, hardware will report raw ranging data without calculation\n"
                "For 3D positioning with >10 anchors, hardware calculation may fail - use software mode"
                ),
            'Waveform': (
                "Waveform Display\n"
                "<Usage> Waveform 1/2 corresponds to Channel Data Table 1/2\n"
                "<Controls>\n"
                "Edit tag ID to display specific tag data\n"
                "Left-click lines to view detailed values\n"
                "Right-click to pause scrolling and pan view\n"
                "Mouse wheel to zoom in/out\n"
                "After manual interaction, click 'Reset View' to resume auto-scrolling\n"
                "<Default> Channel 1 Tag ID: 0 | Channel 2 Tag ID: 1\n"
                "<Activation> Press Enter after changing ID"
                )
        }
        
        self.config_vars = None
        self.log_area = None  # Will be set later
        self.last_position = None
        self.total_distance = 0.0  # Add these tracking variables
        self.total_time = 0.0
        self.tag_table = None  # Will be set to reference the GUI table
        self.max_tags = max_tags
        self.target_position = None
        self.is_show_trajectory = True
        self.trajectory_button_text = tk.StringVar(value="Clear Trajectory Display")
        self.tag_lock = threading.Lock()
        self.data_transmission_enabled = False
        self.transmission_data = bytearray()
        self.transmission_interval = 1000  # ms
        self.transmission_thread = None
        self.transmission_running = False
        self.signal_strength_tag = "0"  # Default to tag 0
        self.signal_filter_alpha = 0.35  # Default filter coefficient
        
    def get_tag_color(self, tag_id):
        """Get color for a tag, assigning one if it doesn't exist"""
        # Use the tag's ID to get a consistent color
        with self.tag_lock:
            if tag_id not in self.tag_colors:
                # Generate a new color if we run out of predefined ones
                color_idx = len(self.tag_colors) - len(self.tag_colors) % 8  # Cycle through colors
                r = (color_idx * 37) % 256
                g = (color_idx * 73) % 256
                b = (color_idx * 109) % 256
                self.tag_colors[tag_id] = (r, g, b)
            
            return self.tag_colors.get(tag_id, self.tag_colors['default'])

    def get_anchor_color(self, anchor_id):
        """Get color for an anchor"""
        return self.anchor_positions.get(anchor_id, {}).get('color', (255, 255, 255))     
    
    def find_tag(self, tag_id):
        """Thread-safe tag lookup"""
        with self.tag_lock:
            return self.tags.get(tag_id, None)
        
    def check_bit_is_true(self, data, b):
        """Check if a specific bit is set"""
        return ((data >> b) & 0x01) == 0x01
    
    def set_log_area(self, log_area):
        """Set the log area after GUI initialization"""
        self.log_area = log_area    
    
    def safe_log(self, message):
        """Thread-safe logging with fallback"""
        if self.log_area:
            try:
                self.log_area.after(0, lambda: self.log_area.insert(tk.END, message))
            except:
                print(f"GUI log failed: {message}")
        else:
            print(message)  # Console fallback
            
    def process_modbus_data(self, data):
        """Process incoming MODBUS data from PG anchors"""
        try:
            if len(data) < 5:
                return False
                
            if data[0] == 0x01 and data[1] == 0x03:  # MODBUS response
                length = data[2]
                if len(data) < length + 5:
                    return False  # Incomplete message
                
                # Verify CRC
                crc = self.crc16(data[:-2])
                if crc != (data[-2] << 8 | data[-1]):
                    self.safe_log("CRC check failed")
                    return False
                
                # Check for calibration data (0xDA 0xDA)
                if data[3] == 0xDA and data[4] == 0xDA:
                    return self.handle_calibration_data(data)
                
                if data[3] == 0xCA and data[4] == 0xDA:
                    self.process_anchor_data(data[5:-2])
                elif data[3] == 0xAC and data[4] == 0xDA:
                    self.process_tag_data(data[5:-2])
                
                return True
            
        except Exception as e:
            self.safe_log(f"Error processing MODBUS data: {str(e)}")
            return False
    
    def process_anchor_data(self, data):
        """Process data from anchor station"""
        ptr = 0
        output_protocal = (data[ptr] << 8) | data[ptr+1]
        ptr += 2
        
        self.last_cal_data['Tag_ID'] = (data[ptr] << 8) | data[ptr+1]
        ptr += 2
        
        # Status flags
        cal_flag = (data[ptr] << 24) | (data[ptr+1] << 16) | (data[ptr+2] << 8) | data[ptr+3]
        ptr += 4
        self.last_cal_data['Cal_Flag'] = cal_flag
        
        # Positioning data
        if self.check_bit_is_true(output_protocal, ANC_PROTOCAL_RTLS):
            if not self.check_bit_is_true(cal_flag, 16):  # Positioning failed
                ptr += 6
                self.safe_log("Positioning calculation failed")
            else:
                self.last_cal_data['x'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.last_cal_data['y'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.last_cal_data['z'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.safe_log(f"Positioning success - X: {self.last_cal_data['x']} Y: {self.last_cal_data['y']} Z: {self.last_cal_data['z']}")
        
        # Distance data
        if self.check_bit_is_true(output_protocal, ANC_PROTOCAL_DIST):
            for i in range(16):
                if self.check_bit_is_true(cal_flag, i):
                    self.last_cal_data['Dist'][i] = (data[ptr] << 8) | data[ptr+1]
                    ptr += 2
                else:
                    self.last_cal_data['Dist'][i] = 0
                    ptr += 2
        
        # Reception diagnostics
        if self.check_bit_is_true(output_protocal, ANC_PROTOCAL_RXDIAG):
            if self.last_cal_data['Tag_ID'] == self.active_tag:
                self.rx_diag['Max_noise'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Std_noise'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Fp_amp1'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Fp_amp2'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Fp_amp3'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Max_growthCIR'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Rx_preambleCount'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Fp'] = ((data[ptr] << 8) | data[ptr+1]) / 64
                ptr += 2
                
                self.rx_diag['Fp_power'] = self.calculate_fp_power(self.rx_diag)
                self.rx_diag['Rx_power'] = self.calculate_rx_power(self.rx_diag)
            else:
                ptr += 16
        
        # Timestamps
        if self.check_bit_is_true(output_protocal, ANC_PROTOCAL_TIMESTAMP):
            if self.last_cal_data['Tag_ID'] == self.active_tag:
                for j in range(6):
                    self.last_cal_data['Time_ts'][j] = (data[ptr] << 24) | (data[ptr+1] << 16) | (data[ptr+2] << 8) | data[ptr+3]
                    ptr += 4
    
    def process_tag_data(self, data):
        """Process data from tag"""
        ptr = 0
        output_protocal = (data[ptr] << 8) | data[ptr+1]
        ptr += 2
        
        # Distance data
        if self.check_bit_is_true(output_protocal, TAG_OUTPUT_DIST):
            dist_flag = (data[ptr] << 8) | data[ptr+1]
            ptr += 2
            
            for i in range(16):
                if self.check_bit_is_true(dist_flag, i):
                    self.last_cal_data['Dist'][i] = (data[ptr] << 8) | data[ptr+1]
                    ptr += 2
                else:
                    ptr += 2
        
        # Positioning data
        if self.check_bit_is_true(output_protocal, TAG_OUTPUT_RTLS):
            if (data[ptr] << 8) | data[ptr+1] == 1:  # Positioning enabled
                ptr += 2
                self.last_cal_data['x'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.last_cal_data['y'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.last_cal_data['z'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
            else:
                ptr += 2
    
    def process_rtls_packet(self, data: bytes):
        """
        NEW: This is the primary function to parse the RTLS packet.
        It's based on the C# Rtls_DataRecv function.
        It expects the PAYLOAD part of the Modbus frame (after the 0xCA 0xDA).
        """
        try:
            # --- Unpack header fields ---
            # '>' for big-endian, 'H' for unsigned short, 'I' for unsigned int
            unpacker_header = struct.Struct('>HHI')
            output_protocol, tag_id, cal_flag = unpacker_header.unpack_from(data, 0)
            cursor = unpacker_header.size

            tag = self.get_or_create_tag(str(tag_id))

            # --- Extract Position (if available) ---
            if (output_protocol & (1 << ANC_PROTOCAL_RTLS)):
                position_valid = bool((cal_flag >> 16) & 0x01)
                if position_valid:
                    # '>' for big-endian, 'h' for signed short
                    pos_unpacker = struct.Struct('>hhh')
                    raw_x, raw_y, raw_z = pos_unpacker.unpack_from(data, cursor)
                    # The module sends coordinates in cm, convert to meters for consistency
                    tag.update_position(raw_x / 100.0, raw_y / 100.0, raw_z / 100.0)
                cursor += 6  # Always advance cursor by 6 bytes for position data block

            # --- Extract Distances (if available) ---
            if (output_protocol & (1 << ANC_PROTOCAL_DIST)):
                dist_unpacker = struct.Struct('>H') # unsigned short
                for i in range(ANCHOR_MAX_COUNT):
                    # The distance field is always present, but we check the flag to see if it's valid
                    distance_cm = dist_unpacker.unpack_from(data, cursor)[0]
                    is_valid = bool((cal_flag >> i) & 0x01)
                    if is_valid:
                        # You can store this distance data in your anchor/tag objects if needed
                        # For now, we just log it for debugging
                        if i == 0: # Log distance to anchor A for a quick check
                            # print(f"Tag {tag_id} Dist to A: {distance_cm} cm")
                            pass
                    cursor += 2

            # After parsing, we can update stats and GUI
            root.after(0, self.update_gui_for_tag, tag)
            return True

        except (struct.error, IndexError) as e:
            safe_log_insert(log_area, f"ERROR: Failed to parse RTLS packet: {e}\n")
            return False
        
    def decode_int16(self, data):
        """Decode a 16-bit signed integer from bytes"""
        return struct.unpack('>h', bytes(data))[0]
    
    def crc16(self, data):
        """Calculate MODBUS CRC-16 checksum"""
        crc = 0xFFFF
        for byte in data:
            crc ^= byte
            for _ in range(8):
                if crc & 0x0001:
                    crc >>= 1
                    crc ^= 0xA001
                else:
                    crc >>= 1
        return crc

    def update_or_create_tag(self, tag_data):
        """Update existing tag or create new one with thread safety"""
        tag_id = str(tag_data['id'])
        with self.tag_lock:
            if tag_id not in self.tags:
                self.tags[tag_id] = UWBTag(id=tag_id)
                
            tag = self.tags[tag_id]
            tag.x = tag_data.get('x', tag.x)
            tag.y = tag_data.get('y', tag.y)
            tag.z = tag_data.get('z', tag.z)
            tag.angle = tag_data.get('angle', tag.angle)
            tag.channel = tag_data.get('channel', tag.channel)
            tag.last_update = time.time()
            
            # Calculate velocity and distance
            self._update_tag_movement_stats(tag)
            
            # Update movement history
            tag.movement_history.append((tag.x, tag.y, tag.z, tag.angle, time.time()))
            
            return tag
        
    def _update_tag_movement_stats(self, tag):
        """Calculate velocity and distance traveled"""
        if len(tag.movement_history) > 0:
            prev_x, prev_y, prev_z, _, prev_time = tag.movement_history[-1]
            dt = time.time() - prev_time
            
            if dt > 0:
                dx = tag.x - prev_x
                dy = tag.y - prev_y
                dz = tag.z - prev_z
                distance_step = (dx**2 + dy**2 + dz**2) ** 0.5
                
                tag.velocity = distance_step / dt
                tag.distance_traveled += distance_step
                
                # Update status based on movement
                tag.status = "Moving" if tag.velocity > 0.1 else "Stationary"
                
                # Check for timeout (no updates in 5 seconds)
                if time.time() - tag.last_update > 5:
                    tag.status = "Inactive"

    def filter_multipath(self, x, y, z):
        """Filter positions based on installation guidelines"""
        # Height check (minimum 1m per docs)
        if z < 1.0:
            return False
            
        # Wall proximity check (30cm clearance)
        if x < 0.3 or y < 0.3:
            return False
            
        # Velocity sanity check (5m/s threshold)
        if self.last_position:
            dx = x - self.last_position['x']
            dy = y - self.last_position['y']
            dt = time.time() - self.last_position['time']
            if dt > 0 and (dx**2 + dy**2)**0.5/dt > 5.0:
                return False
                
        self.last_position = {'x': x, 'y': y, 'time': time.time()}
        return True

    def save_config(self, config_vars=None, log_area=None):
        """Unified configuration save handler"""
        config_vars = config_vars or self.config_vars
        if not config_vars:
            raise ValueError("No configuration variables set")
        log_area = log_area or self.log_area
        
        try:
            device_type = config_vars['Device Type'].get().lower()
            type_map = {
                'major anchor': '1',
                'sub anchor': '2', 
                'tag': '3'
            }
            device_type_value = type_map.get(device_type, '1')

            commands = [
                f"nvm 0x80 {config_vars['MODBUS-ID'].get()}",
                f"kalman_q {config_vars['Kalman-Q'].get()}",
                f"kalman_r {config_vars['Kalman-R'].get()}",
                f"antdly {config_vars['Antenna Delay'].get()}",
                f"mode {config_vars['Positioning Mode'].get().lower()}",
                f"chan {config_vars['UWB Channel'].get()}",
                f"rate {config_vars['Data Rate'].get()}",
                f"nvm 0x84 {device_type_value}",
                f"nvm 0x85 {config_vars['Device ID'].get()}",
                f"ranging {config_vars['Ranging Mode'].get().lower()}",
                f"nvm 0x86 {config_vars['Network ID'].get()}"
            ]

            for cmd in commands:
                if not send_to_major_anchor(cmd):
                    raise Exception(f"Command failed: {cmd}")
                    
            safe_log_insert(log_area, "Configuration saved successfully\n")
            return True
            
        except Exception as e:
            safe_log_insert(log_area, f"Configuration save failed: {str(e)}\n")
            return False

    def start_positioning(self):
        """Start HDS-TWR positioning cycle"""
        try:
            # Set network ID first
            if not send_to_major_anchor('set_network', 
                                      {'id': self.config_vars['Network ID'].get()}):
                raise Exception("Network ID config failed")

            # Set data rate based on coverage area
            coverage = float(self.config_vars['Coverage Area'].get())
            rate = 110 if coverage > 15 else 6800  # 110kbps or 6.8Mbps
            send_to_major_anchor('set_rate', {'rate': rate})

            # Execute anchor sequence
            seq = [
                ('start_pos', None),    # Anchor A
                ('trigger_B', 0.003),   # 3ms delay
                ('trigger_C', 0.003),   # 3ms delay
                ('trigger_D', 0.003)    # 3ms delay
            ]
            
            for cmd, delay in seq:
                if not send_to_major_anchor(cmd):
                    raise Exception(f"{cmd} failed")
                if delay:
                    time.sleep(delay)
                    
            safe_log_insert(self.log_area, "HDS-TWR positioning started\n")
            return True
            
        except Exception as e:
            safe_log_insert(self.log_area, f"Start failed: {str(e)}\n")
            return False

    def get_all_tags(self):
        """Get a thread-safe copy of all tags"""
        with self.tag_lock:
            return list(self.tags.values())

    def get_active_tag_info(self):
        """Get info for the currently active tag"""
        if not self.active_tag:
            return None
            
        tag = self.find_tag(self.active_tag)
        if not tag:
            return None
            
        return {
            'id': tag.id,
            'x': tag.x,
            'y': tag.y,
            'z': tag.z,
            'angle': tag.angle,
            'velocity': tag.velocity,
            'distance': tag.distance_traveled,
            'status': tag.status,
            'channel': tag.channel,
            'last_update': tag.last_update
        }

    def reset_tag_stats(self, tag_id):
        """Reset distance and velocity for a tag"""
        with self.tag_lock:
            if tag_id in self.tags:
                tag = self.tags[tag_id]
                tag.distance_traveled = 0.0
                tag.velocity = 0.0
                tag.movement_history.clear()
                
    def validate_position(self, x, y, z):
        """Enhanced position validation from C# code"""
        # Ground/obstruction check
        if z < 1.0:  # Minimum height per docs
            self.safe_log(f"WARNING: Invalid height {z}m - must be ≥1m")
            return False
        
        # Origin rejection (common error state)
        if abs(x) < 0.1 and abs(y) < 0.1:
            self.safe_log("WARNING: Position at origin - likely error")
            return False
        
        # Velocity sanity check
        if self.last_position and self.last_time:
            dt = time.time() - self.last_time
            if dt > 0:
                dx = x - self.last_position['x']
                dy = y - self.last_position['y']
                velocity = (dx**2 + dy**2)**0.5 / dt
                if velocity > 5.0:  # 5m/s = 18km/h
                    self.safe_log(f"WARNING: Implausible velocity: {velocity:.2f}m/s")
                    return False
        
        self.last_position = {'x': x, 'y': y, 'time': time.time()}
        return True
    
    def update_anchor_positions(self, anchor_positions):
        """Validate anchor positions against installation guidelines"""
        for anchor_id, pos in anchor_positions.items():
            if pos['z'] < 1.0:
                self.safe_log(f"WARNING: Anchor {anchor_id} height {pos['z']}m < 1m minimum")
            if min(pos['x'], pos['y']) < 0.5:
                self.safe_log(f"WARNING: Anchor {anchor_id} too close to walls")
                
    def calculate_fp_power(self, rx_diag_data):
        """Calculate First Path Power in dBm. Based on C# code."""
        try:
            if rx_diag_data['Rx_preambleCount'] == 0: return -150.0 # Avoid division by zero
            result = (rx_diag_data['Fp_amp1']**2 + rx_diag_data['Fp_amp2']**2 + rx_diag_data['Fp_amp3']**2)
            n_square = rx_diag_data['Rx_preambleCount']**2
            result /= n_square
            # Use math.log10, checking for result > 0
            result = 10 * math.log10(result) if result > 0 else -150.0
            result -= 121.74 # Constant from DW1000/3000 documentation
            return round(result, 2)
        except Exception:
            return -150.0

    def calculate_rx_power(self, rx_diag_data):
        """Calculate Receive Power in dBm. Based on C# code."""
        try:
            if rx_diag_data['Rx_preambleCount'] == 0: return -150.0 # Avoid division by zero
            pow_value = 2**17  # DW1000 constant (use 2**21 for DW3000 if you add that logic)
            n_square = rx_diag_data['Rx_preambleCount']**2
            result = rx_diag_data['Max_growthCIR'] * pow_value / n_square
            result = 10 * math.log10(result) if result > 0 else -150.0
            result -= 121.74
            return round(result, 2)
        except Exception:
            return -150.0
    
    def process_imu_data(self, data):
        """Process IMU data from device"""
        ptr = 0
        print_en = (data[ptr] << 8) | data[ptr+1]
        ptr += 2
        
        imu_data = {}
        
        if print_en & IMU_DATA_ACC_EN:
            imu_data['acc_x'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * acc_fsr
            ptr += 2
            imu_data['acc_y'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * acc_fsr
            ptr += 2
            imu_data['acc_z'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * acc_fsr
            ptr += 2
            
        if print_en & IMU_DATA_GYRO_EN:
            imu_data['gyro_x'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * gyro_fsr
            ptr += 2
            imu_data['gyro_y'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * gyro_fsr
            ptr += 2
            imu_data['gyro_z'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * gyro_fsr
            ptr += 2
            
        if print_en & IMU_DATA_EULER_EN:
            imu_data['rotation_x'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * 180.0
            ptr += 2
            imu_data['rotation_y'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * 180.0
            ptr += 2
            imu_data['rotation_z'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * 180.0
            ptr += 2
            
        if print_en & IMU_DATA_TEMP_EN:
            imu_data['temperature'] = self.decode_int16(data[ptr:ptr+2]) / 340.0 + 36.53
            ptr += 2
            
        if print_en & IMU_DATA_Q_EN:
            imu_data['q0'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0
            ptr += 2
            imu_data['q1'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0
            ptr += 2
            imu_data['q2'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0
            ptr += 2
            imu_data['q3'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0
            ptr += 2
    
        if print_en & IMU_DATA_MAGN_EN:
            imu_data['magn_x'] = self.decode_int16(data[ptr:ptr+2]) * 0.15
            ptr += 2
            imu_data['magn_y'] = self.decode_int16(data[ptr:ptr+2]) * 0.15
            ptr += 2
            imu_data['magn_z'] = self.decode_int16(data[ptr:ptr+2]) * 0.15
            ptr += 2
            
        return imu_data 
    
    def update_tag_table(self):
        """Update the GUI table with current tag data"""
        if not self.tag_table:
            return
            
        # Clear existing rows
        for row in self.tag_table.get_children():
            self.tag_table.delete(row)
            
        # Add current tag data
        for tag_id, tag in self.tags.items():
            x, y, z, angle = tag.get_position()
            self.tag_table.insert("", "end", 
                                 values=(tag_id, f"{x:.2f}", f"{y:.2f}", f"{z:.2f}", 
                                        f"{angle:.1f}", tag.status, tag.channel))
    
    def set_target(self, x, y):
        """Set target coordinates"""
        self.target_position = (x, y)
        
    def clear_target(self):
        """Clear target"""
        self.target_position = None
    def export_channel_data(self, channel, filename=None):
        """Export channel data to Excel file"""
        if not filename:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title=f"Export Channel {channel} Data"
            )
            if not filename:
                return False
                
        table_name = 'trace1' if channel == 1 else 'trace2'
        
        try:
            # Use the enhanced Excel export function
            return export_to_excel(self.tables[table_name], filename)
        except Exception as e:
            self.safe_log(f"Export failed: {str(e)}\n")
            return False
            
    def toggle_trajectory(self):
        """Toggle trajectory display on/off"""
        self.is_show_trajectory = not self.is_show_trajectory
        if self.is_show_trajectory:
            self.trajectory_button_text.set("Clear Trajectory Display")
        else:
            self.trajectory_button_text.set("Show Trajectory")
        self.safe_log(f"Trajectory display {'enabled' if self.is_show_trajectory else 'disabled'}\n")
        
    def update_ranging_mode_ui(self, mode):
        """Update UI based on selected ranging mode"""
        if mode == "HDS-TWR":
            # Disable data rate selection for HDS-TWR
            self.config_vars['Data Rate'].set('6M8' if self.module_chip == 'DW1000' else '850K')
            self.config_widgets['Data Rate'].config(state='disabled')
        else:
            self.config_widgets['Data Rate'].config(state='normal')

    def update_device_type_ui(self, device_type):
        """Update UI based on device type (Master, Sub, Tag)"""
        if device_type == "Tag":
            self.config_widgets['Positioning Mode'].config(state='disabled')
            self.config_widgets['Device ID'].config(state='normal')
            self.config_widgets['Ranging Mode'].config(state='normal')
            self.config_widgets['UWB Channel'].config(state='normal')
            self.config_widgets['Data Rate'].config(state='normal')
            # Disable Kalman filter settings for tags
            self.config_widgets['Kalman-Q'].config(state='disabled')
            self.config_widgets['Kalman-R'].config(state='disabled')
            
        elif device_type == "Sub Anchor":
            self.config_widgets['Positioning Mode'].config(state='disabled')
            self.config_widgets['Device ID'].config(state='normal')
            # Similar enable/disable logic for other controls
            
        elif device_type == "Master Anchor":
            self.config_widgets['Positioning Mode'].config(state='normal')
            self.config_widgets['Device ID'].config(state='disabled')
            # Enable all configuration options
            self.config_widgets['Kalman-Q'].config(state='normal')
            self.config_widgets['Kalman-R'].config(state='normal')

    def validate_config(self):
        """Validate configuration based on device type and ranging mode"""
        errors = []
        
        # Device ID validation
        device_type = self.config_vars['Device Type'].get()
        device_id = int(self.config_vars['Device ID'].get())
        
        if device_type == "Sub Anchor" and device_id > 14:
            errors.append("Sub Anchor ID must be 0-14 (B=0,C=1,...,P=14)")
        elif device_type == "Tag" and device_id > 99:
            errors.append("Tag ID must be 0-99")
        
        # Network ID validation
        network_id = int(self.config_vars['Network ID'].get())
        if network_id < 1 or network_id > 65535:
            errors.append("Network ID must be 1-65535")
        
        return errors

    def save_config_to_device(self):
        """Save current configuration to device"""
        try:
            # Convert configuration to MODBUS registers
            config_bytes = bytearray(122)  # Match C# buffer size
            
            # MODBUS ID (register 0)
            config_bytes[0:2] = int(self.config_vars['MODBUS-ID'].get()).to_bytes(2, 'big')
            
            # Device type (register 6)
            device_type_map = {"Master Anchor": 2, "Sub Anchor": 1, "Tag": 0}
            config_bytes[6] = device_type_map.get(self.config_vars['Device Type'].get(), 0)
            
            # Device ID (registers 8-9)
            config_bytes[8:10] = int(self.config_vars['Device ID'].get()).to_bytes(2, 'big')
            
            # Kalman filters (registers 12-15)
            config_bytes[12:14] = int(self.config_vars['Kalman-Q'].get()).to_bytes(2, 'big')
            config_bytes[14:16] = int(self.config_vars['Kalman-R'].get()).to_bytes(2, 'big')
            
            # Antenna delay (registers 16-17)
            config_bytes[16:18] = int(self.config_vars['Antenna Delay'].get()).to_bytes(2, 'big')
            
            # Send via MODBUS
            if not self.send_modbus_command(0, config_bytes):
                raise Exception("Failed to write configuration")
                
            self.safe_log("Configuration saved to device\n")
            return True
            
        except Exception as e:
            self.safe_log(f"Failed to save config: {str(e)}\n")
            return False

    def handle_ranging_mode(self, mode):
        """Apply ranging mode specific settings"""
        if mode == "HDS-TWR":
            # Set fixed parameters for HDS-TWR
            self.config_vars['Data Rate'].set('6M8' if self.module_chip == 'DW1000' else '850K')
            self.config_widgets['Data Rate'].config(state='disabled')
            
            # Additional HDS-TWR specific initialization
            if not self.send_to_major_anchor('set_hds_mode'):
                raise Exception("Failed to set HDS-TWR mode")
        else:
            self.config_widgets['Data Rate'].config(state='normal')
            
    def create_help_buttons(self, parent_frame):
        """Create help buttons in the specified frame"""
        for key, text in self.help_texts.items():
            btn = ttk.Button(parent_frame, text=f"{key} Help",
                           command=lambda t=text: self.show_help_popup(t))
            btn.pack(side=tk.LEFT, padx=2)
    
    def show_help_popup(self, text):
        """Display help text in a popup window"""
        popup = tk.Toplevel()
        popup.title("Help")
        text_widget = tk.Text(popup, wrap=tk.WORD, width=60, height=20)
        text_widget.insert(tk.END, text)
        text_widget.config(state=tk.DISABLED)
        text_widget.pack(padx=10, pady=10)
        ttk.Button(popup, text="Close", command=popup.destroy).pack(pady=5)
    
    def setup_calculation_mode(self, parent):
        """Create calculation mode controls"""
        frame = ttk.LabelFrame(parent, text="Calculation Mode")
        
        self.calc_mode = tk.BooleanVar(value=True)  # Default to hardware
        
        # Use grid instead of pack
        ttk.Checkbutton(frame, text="Hardware Calculation", 
                        variable=self.calc_mode,
                        command=self.update_calc_mode).grid(row=0, column=0, padx=2, pady=2)
    
        ttk.Button(frame, text="Help", 
                   command=self.show_calc_help).grid(row=0, column=1, padx=2, pady=2)
        
        return frame
    
    def update_calc_mode(self):
        """Send calculation mode to device"""
        mode = "hardware" if self.calc_mode.get() else "software"
        if not self.send_to_major_anchor(f"calc_mode {mode}"):
            self.safe_log("Failed to update calculation mode\n")
            # Revert on failure
            self.calc_mode.set(not self.calc_mode.get())
        
    
    def update_ranging_mode(self, mode):
        """Handle ranging mode changes"""
        if mode == "HDS-TWR":
            # Set fixed parameters
            rate = "6M8" if self.module_chip == "DW1000" else "850K"
            self.config_vars['Data Rate'].set(rate)
            self.config_widgets['Data Rate'].config(state='disabled')
            
            # Update channel options
            if self.module_chip == "DW3000":
                self.config_widgets['UWB Channel']['values'] = ["1", "2", "3", "4", "5"]
            
            self.safe_log("All devices must be set to HDS-TWR mode!\n")
        else:
            self.config_widgets['Data Rate'].config(state='normal')
            self.config_widgets['UWB Channel']['values'] = ["1", "2", "3", "4", "5", "7"]
        
        # Send command to device
        if not self.send_to_major_anchor(f"ranging {mode.lower()}"):
            self.safe_log("Failed to update ranging mode\n")
    
    def show_calc_help(self):
        """Show help information about calculation modes"""
        help_text = (
            "Calculation Mode Help\n\n"
            "Hardware Calculation: Performed by the anchor's internal processor\n"
            "Software Calculation: Performed by this application\n\n"
            "For large deployments (>10 anchors), software mode is recommended."
        )
        
        # Create a simple help popup
        help_window = tk.Toplevel()
        help_window.title("Calculation Mode Help")
        help_window.geometry("400x200")
        
        text_widget = tk.Text(help_window, wrap=tk.WORD, padx=10, pady=10)
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
        text_widget.pack(fill=tk.BOTH, expand=True)
        
        ttk.Button(help_window, text="Close", command=help_window.destroy).pack(pady=5)
    
    def start_data_transmission(self, tag_id, data, hex_mode=False, interval=0):
        """Start transmitting data to a tag"""
        try:
            if hex_mode:
                self.transmission_data = bytes.fromhex(data.replace(" ", ""))
            else:
                self.transmission_data = data.encode('ascii')[:10]  # Max 10 bytes
            
            if interval > 0:
                self.transmission_interval = interval
                self.transmission_running = True
                self.transmission_thread = threading.Thread(
                    target=self._transmission_loop, args=(tag_id,), daemon=True)
                self.transmission_thread.start()
            else:
                self._send_transmission_data(tag_id)
                
            return True
        except Exception as e:
            self.safe_log(f"Transmission error: {str(e)}\n")
            return False
            
    def _transmission_loop(self, tag_id):
        """Continuous transmission loop"""
        while self.transmission_running:
            self._send_transmission_data(tag_id)
            time.sleep(self.transmission_interval / 1000)
            
    def _send_transmission_data(self, tag_id):
        """Send data to specific tag"""
        if client_socket:
            try:
                cmd = f"AT+DataSend=\"{self.transmission_data.hex()}\",\"{tag_id}\"\n"
                client_socket.sendall(cmd.encode())
            except Exception as e:
                self.safe_log(f"Transmission send error: {str(e)}\n")
                
    def stop_data_transmission(self):
        """Stop continuous transmission"""
        self.transmission_running = False
        if self.transmission_thread:
            self.transmission_thread.join()
    
    def set_anchor_output_protocol(self, positioning=True, distances=True, 
                                 diagnostics=True, timestamps=True):
        """Configure what data the anchor outputs"""
        if not anchor_instrument:
            self.safe_log("No anchor connection\n")
            return False
            
        try:
            # Convert boolean flags to bitmask
            protocol = 0
            if positioning: protocol |= (1 << ANC_PROTOCAL_RTLS)
            if distances: protocol |= (1 << ANC_PROTOCAL_DIST) 
            if diagnostics: protocol |= (1 << ANC_PROTOCAL_RXDIAG)
            if timestamps: protocol |= (1 << ANC_PROTOCAL_TIMESTAMP)
            
            # Send via MODBUS
            with anchor_lock:
                anchor_instrument.write_register(0x30, protocol)
            return True
        except Exception as e:
            self.safe_log(f"Protocol config failed: {str(e)}\n")
            return False
    
    def set_signal_strength_tag(self, tag_id):
        """Set which tag to monitor for signal strength"""
        if tag_id in self.tags:
            self.signal_strength_tag = tag_id
            return True
        return False
        
    def apply_signal_filter(self, new_value, last_value):
        """Apply low-pass filter to signal strength"""
        return last_value * (1 - self.signal_filter_alpha) + new_value * self.signal_filter_alpha
        
    def update_signal_strength_display(self):
        """Update GUI with filtered signal strength"""
        if self.signal_strength_tag in self.tags:
            tag = self.tags[self.signal_strength_tag]
            # Get raw values from rx_diag
            fp_power = self.calculate_fp_power(self.rx_diag)
            rx_power = self.calculate_rx_power(self.rx_diag)
            
            # Apply filter
            tag.filtered_fp_power = self.apply_signal_filter(
                fp_power, getattr(tag, 'filtered_fp_power', fp_power))
            tag.filtered_rx_power = self.apply_signal_filter(
                rx_power, getattr(tag, 'filtered_rx_power', rx_power))
                
            # Update GUI
            if hasattr(self, 'signal_strength_var'):
                self.signal_strength_var.set(
                    f"Tag {tag.id}: FP={tag.filtered_fp_power:.2f}dBm, "
                    f"RX={tag.filtered_rx_power:.2f}dBm")
            
    def handle_calibration_data(self, data):
        """Process calibration data received from anchor"""
        if len(data) < 5 or data[3] != 0xDA or data[4] != 0xDA:
            return False
            
        try:
            # Parse calibration data (similar to C# code)
            anchor_id = chr(65 + data[5])  # Convert index to A-F
            x = struct.unpack('>h', bytes(data[13:14]))[0]  # Big-endian int16
            y = struct.unpack('>h', bytes(data[15:16]))[0]
            z = struct.unpack('>h', bytes(data[17:18]))[0]
            
            # Update anchor position
            if self.anchor_manager: # Check if the manager is set
                self.anchor_manager.update_anchor_position(anchor_id, x / 100.0, y / 100.0, z / 100.0)
                self.safe_log(f"Calibration update: {anchor_id} X={x}cm, Y={y}cm, Z={z}cm\n")
                return True
                
        except Exception as e:
            self.safe_log(f"Error processing calibration data: {str(e)}\n")
            return False
    
    
        
class DataBinding:
    def __init__(self, root):
        self.root = root
        self.bindings = {}
        
        # Initialize variables for data binding
        self.vars = {
            'tag_x': tk.StringVar(value="0.00"),
            'tag_y': tk.StringVar(value="0.00"),
            'tag_z': tk.StringVar(value="0.00"),
            'tag_angle': tk.StringVar(value="0.0"),
            'tag_velocity': tk.StringVar(value="0.00"),
            'tag_distance': tk.StringVar(value="0.00"),
            'tag_status': tk.StringVar(value="Inactive"),
            'tag_channel': tk.StringVar(value="2"),
            
            # Anchor diagnostics
            'max_noise': tk.StringVar(value="0"),
            'std_noise': tk.StringVar(value="0"),
            'fp_amp1': tk.StringVar(value="0"),
            'fp_amp2': tk.StringVar(value="0"),
            'fp_amp3': tk.StringVar(value="0"),
            'max_growth_cir': tk.StringVar(value="0"),
            'rx_preamble_count': tk.StringVar(value="0"),
            'fp': tk.StringVar(value="0"),
            'fp_power': tk.StringVar(value="0.00"),
            'rx_power': tk.StringVar(value="0.00"),
            
            # IMU data
            'acc_x': tk.StringVar(value="0.00"),
            'acc_y': tk.StringVar(value="0.00"),
            'acc_z': tk.StringVar(value="0.00"),
            'gyro_x': tk.StringVar(value="0.00"),
            'gyro_y': tk.StringVar(value="0.00"),
            'gyro_z': tk.StringVar(value="0.00"),
            'magn_x': tk.StringVar(value="0.00"),
            'magn_y': tk.StringVar(value="0.00"),
            'magn_z': tk.StringVar(value="0.00"),
            'rotation_x': tk.StringVar(value="0.00"),
            'rotation_y': tk.StringVar(value="0.00"),
            'rotation_z': tk.StringVar(value="0.00"),
            'temperature': tk.StringVar(value="0.00"),
        }
        
    def update_from_tag(self, tag):
        """Update all bound variables from a tag object"""
        if not tag:
            return
            
        self.vars['tag_x'].set(f"{tag.x:.2f}")
        self.vars['tag_y'].set(f"{tag.y:.2f}")
        self.vars['tag_z'].set(f"{tag.z:.2f}")
        self.vars['tag_angle'].set(f"{tag.angle:.1f}")
        self.vars['tag_velocity'].set(f"{tag.velocity:.2f}")
        self.vars['tag_distance'].set(f"{tag.distance_traveled:.2f}")
        self.vars['tag_status'].set(tag.status)
        self.vars['tag_channel'].set(str(tag.channel))
        
    def bind_widget(self, widget, var_name, widget_property='textvariable'):
        """Bind a widget to a data variable"""
        if var_name not in self.vars:
            raise ValueError(f"Unknown variable name: {var_name}")
            
        widget.configure(**{widget_property: self.vars[var_name]})
        self.bindings[(widget, widget_property)] = var_name
        
    def update_anchor_diagnostics(self, diagnostics):
        """Update anchor diagnostic variables"""
        self.vars['max_noise'].set(str(diagnostics['Max_noise']))
        self.vars['std_noise'].set(str(diagnostics['Std_noise']))
        self.vars['fp_amp1'].set(str(diagnostics['Fp_amp1']))
        self.vars['fp_amp2'].set(str(diagnostics['Fp_amp2']))
        self.vars['fp_amp3'].set(str(diagnostics['Fp_amp3']))
        self.vars['max_growth_cir'].set(str(diagnostics['Max_growthCIR']))
        self.vars['rx_preamble_count'].set(str(diagnostics['Rx_preambleCount']))
        self.vars['fp'].set(f"{diagnostics['Fp']:.2f}")
        self.vars['fp_power'].set(f"{diagnostics['Fp_power']:.2f}")
        self.vars['rx_power'].set(f"{diagnostics['Rx_power']:.2f}")
        
    def update_imu_data(self, imu_data):
        """Update IMU data variables"""
        self.vars['acc_x'].set(f"{imu_data['acc_x']:.2f}")
        self.vars['acc_y'].set(f"{imu_data['acc_y']:.2f}")
        self.vars['acc_z'].set(f"{imu_data['acc_z']:.2f}")
        self.vars['gyro_x'].set(f"{imu_data['gyro_x']:.2f}")
        self.vars['gyro_y'].set(f"{imu_data['gyro_y']:.2f}")
        self.vars['gyro_z'].set(f"{imu_data['gyro_z']:.2f}")
        self.vars['magn_x'].set(f"{imu_data['magn_x']:.2f}")
        self.vars['magn_y'].set(f"{imu_data['magn_y']:.2f}")
        self.vars['magn_z'].set(f"{imu_data['magn_z']:.2f}")
        self.vars['rotation_x'].set(f"{imu_data['rotation_x']:.2f}")
        self.vars['rotation_y'].set(f"{imu_data['rotation_y']:.2f}")
        self.vars['rotation_z'].set(f"{imu_data['rotation_z']:.2f}")
        self.vars['temperature'].set(f"{imu_data['temperature']:.2f}")

class AnchorManager:
    def __init__(self):
        self.anchors = {}  # Dictionary of anchors keyed by ID
        self.active_anchors = set()
        self.anchor_lock = threading.Lock()
        self.max_anchors = 16  # PGPlus supports up to 16 anchors
        
        # Initialize default anchor positions
        self.anchor_positions = {
            'A': {'x': 0, 'y': 0, 'z': 1.2, 'enabled': True, 'color': (0, 255, 0)},
            'B': {'x': 5, 'y': 0, 'z': 1.2, 'enabled': True, 'color': (0, 0, 255)},
            'C': {'x': 5, 'y': 5, 'z': 1.2, 'enabled': True, 'color': (255, 0, 0)},
            'D': {'x': 0, 'y': 5, 'z': 1.2, 'enabled': True, 'color': (255, 255, 0)}
        }
        
        # Anchor colors for visualization
        self.anchor_colors = [
            (255, 0, 0),    # Red - A
            (0, 255, 0),    # Green - B
            (0, 0, 255),    # Blue - C
            (255, 255, 0),  # Yellow - D
            (255, 0, 255),  # Magenta - E
            (0, 255, 255),  # Cyan - F
            (255, 128, 0),  # Orange - G
            (128, 0, 128),  # Purple - H
            # Add more colors as needed
        ]
        
        self.initialize_anchors()

    def initialize_anchors(self):
        """Initialize anchor objects with default positions"""
        with self.anchor_lock:
            for i, (anchor_id, pos) in enumerate(self.anchor_positions.items()):
                self.anchors[anchor_id] = {
                    'x': pos['x'],
                    'y': pos['y'],
                    'z': pos['z'],
                    'enabled': pos['enabled'],
                    'color': self.anchor_colors[i % len(self.anchor_colors)],
                    'distance_history': deque(maxlen=100),
                    'last_update': 0
                }
            # Enable first anchor by default
            if self.anchors:
                first_anchor = next(iter(self.anchors))
                self.active_anchors.add(first_anchor)

    def update_anchor_position(self, anchor_id, x, y, z):
        """Update anchor position with thread safety"""
        with self.anchor_lock:
            if anchor_id in self.anchors:
                self.anchors[anchor_id]['x'] = x
                self.anchors[anchor_id]['y'] = y
                self.anchors[anchor_id]['z'] = z
                self.anchors[anchor_id]['last_update'] = time.time()
                return True
            return False

    def get_anchor_positions(self):
        """Get thread-safe copy of anchor positions"""
        with self.anchor_lock:
            return {k: v.copy() for k, v in self.anchors.items() if v['enabled']}

    def toggle_anchor(self, anchor_id, enabled):
        """Enable/disable an anchor"""
        with self.anchor_lock:
            if anchor_id in self.anchors:
                self.anchors[anchor_id]['enabled'] = enabled
                if enabled:
                    self.active_anchors.add(anchor_id)
                else:
                    self.active_anchors.discard(anchor_id)
                return True
            return False

    def calibrate_anchor(self, anchor_id, x, y, z):
        """Calibrate anchor position"""
        with self.anchor_lock:
            if anchor_id in self.anchors:
                self.anchors[anchor_id]['x'] = x
                self.anchors[anchor_id]['y'] = y
                self.anchors[anchor_id]['z'] = z
                return True
            return False

class DataLogger:
    def __init__(self, max_records=6000):
        self.max_records = max_records
        self.data_lock = threading.Lock()
        self.range_analysis_tag = "0"
        self.recording = False
        
        # Initialize data tables
        self.tables = {
            'trace1': self._create_table_template(),
            'trace2': self._create_table_template(),
            'analysis': self._create_analysis_table()
        }
        
    def _create_table_template(self):
        """Create template for trace tables"""
        columns = [
            'Time', 'x', 'y', 'z', 'A', 'B', 'C', 'D', 
            'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 
            'M', 'N', 'O', 'P', 'Flag', 'Velocity'
        ]
        return {col: deque(maxlen=self.max_records) for col in columns}
        
    def _create_analysis_table(self):
        """Create analysis-specific table"""
        columns = [
            'Time', 'AnchorID', 'Distance', 'MaxNoise', 'StdNoise',
            'FirstPathAmp', 'CIR', 'PreambleCount', 'RxPower'
        ]
        return {col: deque(maxlen=self.max_records) for col in columns}
        
    def log_trace_data(self, table_name, tag_data, distances):
        """Log trace data for a tag"""
        if table_name not in ['trace1', 'trace2']:
            return
            
        with self.data_lock:
            timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
            self.tables[table_name]['Time'].append(timestamp)
            self.tables[table_name]['x'].append(f"{tag_data['x']:.2f}")
            self.tables[table_name]['y'].append(f"{tag_data['y']:.2f}")
            self.tables[table_name]['z'].append(f"{tag_data['z']:.2f}")
            
            # Log distances to each anchor
            for anchor_id in 'ABCDEFGHIJKLMNOP':
                dist = distances.get(anchor_id, '0')
                self.tables[table_name][anchor_id].append(str(dist))
                
            self.tables[table_name]['Flag'].append(tag_data['status'])
            self.tables[table_name]['Velocity'].append(f"{tag_data['velocity']:.2f}")
            
    def log_analysis_data(self, anchor_id, diagnostic_data):
        """Log anchor diagnostic data"""
        with self.data_lock:
            timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
            self.tables['analysis']['Time'].append(timestamp)
            self.tables['analysis']['AnchorID'].append(anchor_id)
            self.tables['analysis']['Distance'].append(str(diagnostic_data['distance']))
            self.tables['analysis']['MaxNoise'].append(str(diagnostic_data['Max_noise']))
            self.tables['analysis']['StdNoise'].append(str(diagnostic_data['Std_noise']))
            self.tables['analysis']['FirstPathAmp'].append(
                f"{diagnostic_data['Fp_amp1']}/{diagnostic_data['Fp_amp2']}/{diagnostic_data['Fp_amp3']}"
            )
            self.tables['analysis']['CIR'].append(str(diagnostic_data['Max_growthCIR']))
            self.tables['analysis']['PreambleCount'].append(str(diagnostic_data['Rx_preambleCount']))
            self.tables['analysis']['RxPower'].append(f"{diagnostic_data['Rx_power']:.2f}")
    
    def log_position_data(self, tag_id, x, y, z, velocity, status):
        """Log position data with timestamp"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
        with self.data_lock:
            self.position_log.append({
                'timestamp': timestamp,
                'tag_id': tag_id,
                'x': x,
                'y': y,
                'z': z,
                'velocity': velocity,
                'status': status
                })
            
    def export_to_csv(self, table_name, filename):
        """Export data table to CSV file"""
        if table_name not in self.tables:
            return False
            
        try:
            with self.data_lock:
                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)
                    # Write header
                    writer.writerow(self.tables[table_name].keys())
                    # Write data
                    for i in range(len(next(iter(self.tables[table_name].values())))):
                        row = [col_data[i] for col_data in self.tables[table_name].values()]
                        writer.writerow(row)
            return True
        except Exception as e:
            print(f"Export failed: {e}")
            return False
        
    def setup_waveform_controls(self, parent):
        """Create waveform control panel"""
        control_frame = ttk.Frame(parent)
        
        # Channel selection
        self.channel_vars = {}
        for i in [1, 2]:
            ch_frame = ttk.LabelFrame(control_frame, text=f"Channel {i}")
            ttk.Label(ch_frame, text="Tag ID:").pack()
            var = tk.StringVar(value=str(i-1))
            entry = ttk.Entry(ch_frame, textvariable=var, width=5)
            entry.bind("<Return>", lambda e, ch=i: self.set_channel_tag(ch))
            entry.pack()
            self.channel_vars[i] = var
        
        # View controls
        btn_frame = ttk.Frame(control_frame)
        ttk.Button(btn_frame, text="Reset View", 
                  command=self.reset_waveform_view).pack(pady=2)
        ttk.Button(btn_frame, text="Pause", 
                  command=self.toggle_pause).pack(pady=2)
        
        return control_frame
    
    def set_channel_tag(self, channel):
        """Change which tag is displayed on a channel"""
        tag_id = self.channel_vars[channel].get()
        if tag_id in self.tag_manager.tags:
            self.active_tags[channel] = tag_id
            self.clear_channel_data(channel)
            
    def reset_waveform_view(self):
        """Reset the waveform view to default"""
        # Add your implementation here
        print("Waveform view reset")
    
    def toggle_pause(self):
        """Toggle pause state of waveform updates"""
        # Add your implementation here
        print("Waveform pause toggled")
    
    def clear_channel_data(self, channel):
        """Clear data for a specific channel"""
        # Add your implementation here
        print(f"Channel {channel} data cleared")
    
    def start_range_recording(self, tag_id):
        """Start recording range analysis data"""
        if tag_id in tag_manager.tags:
            self.range_analysis_tag = tag_id
            self.recording = True
            return True
        return False
        
    def stop_range_recording(self):
        """Stop recording"""
        self.recording = False
        
    def export_range_data(self, filename=None):
        """Export recorded range data to Excel"""
        if not filename:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Range Analysis Data")
            if not filename:
                return False
                
        try:
            # Create DataFrame from analysis table
            df = pd.DataFrame({
                'Time': list(self.tables['analysis']['Time']),
                'AnchorID': list(self.tables['analysis']['AnchorID']),
                'Distance': list(self.tables['analysis']['Distance']),
                'FP_Power': list(self.tables['analysis']['FirstPathAmp']),
                'RX_Power': list(self.tables['analysis']['RxPower'])
            })
            
            df.to_excel(filename, index=False)
            return True
        except Exception as e:
            print(f"Export error: {e}")
            return False
        
class EnhancedDataLogger(DataLogger):
    def __init__(self):
        super().__init__()
        # Add analysis-specific columns
        self.tables['analysis']['columns'].extend([
            'Cal_Flag', 'Max_noise', 'Std_noise', 'Fp_amp1', 'Fp_amp2', 'Fp_amp3',
            'Max_growthCIR', 'Rx_preambleCount', 'Fp', 'Fp_power', 'Rx_power'
        ])
        
    def save_tag_data(self, tag_id, x, y, z, dists, cal_flag, velocity, mode=1):
        """Save tag data with enhanced fields from C# code"""
        table_name = 'trace1' if mode == 1 else 'trace2'
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        
        with self.data_lock:
            self.tables[table_name]['Time'].append(timestamp)
            self.tables[table_name]['x'].append(x)
            self.tables[table_name]['y'].append(y)
            self.tables[table_name]['z'].append(z)
            
            for i in range(16):  # ANCHOR_MAX_COUNT
                anchor_id = chr(65 + i)  # A-P
                self.tables[table_name][anchor_id].append(dists[i] if i < len(dists) else 0)
                
            self.tables[table_name]['Flag'].append(cal_flag)
            self.tables[table_name]['Velocity'].append(velocity)

class CommunicationHandler:
    def __init__(self, serial_port=None, tcp_host=None, tcp_port=None):
        self.serial_conn = serial.Serial() if serial_port else None
        self.tcp_conn = socket.socket() if tcp_host else None
        self.comm_lock = threading.Lock()
        self.recv_buffer = bytearray()
        self.buffer_lock = threading.Lock()
        self.comm_mode = 'serial' if serial_port else 'tcp' if tcp_host else None
        self.running = False
        self.data_send_periodically = None
        self.send_timer = None
        
        if serial_port:
            self._setup_serial(serial_port)
        if tcp_host:
            self._setup_tcp(tcp_host, tcp_port)
            
    def _setup_serial(self, port, baudrate=115200):
        """Configure serial connection"""
        self.serial_conn.port = port
        self.serial_conn.baudrate = baudrate
        self.serial_conn.bytesize = serial.EIGHTBITS
        self.serial_conn.parity = serial.PARITY_NONE
        self.serial_conn.stopbits = serial.STOPBITS_ONE
        self.serial_conn.timeout = 0.1
        self.serial_conn.write_timeout = 0.1
        
    def _setup_tcp(self, host, port):
        """Configure TCP connection"""
        self.tcp_conn.settimeout(1.0)
        self.tcp_host = host
        self.tcp_port = port
        
    def start(self):
        """Start communication threads"""
        if not self.running:
            self.running = True
            threading.Thread(target=self._recv_thread, daemon=True).start()
            
    def stop(self):
        """Stop communication"""
        self.running = False
        with self.comm_lock:
            if self.serial_conn and self.serial_conn.is_open:
                self.serial_conn.close()
            if self.tcp_conn:
                self.tcp_conn.close()
                
    def send_data(self, data):
        """Send data through active connection"""
        if not self.running:
            return False
            
        try:
            with self.comm_lock:
                if self.comm_mode == 'serial' and self.serial_conn.is_open:
                    self.serial_conn.write(data)
                    self._log_transmission(data, 'TX')
                    return True
                elif self.comm_mode == 'tcp' and self.tcp_conn:
                    self.tcp_conn.sendall(data)
                    self._log_transmission(data, 'TX')
                    return True
            return False
        except Exception as e:
            print(f"Send error: {e}")
            return False
            
    def _recv_thread(self):
        """Background thread for receiving data"""
        while self.running:
            try:
                if self.comm_mode == 'serial' and self.serial_conn.is_open:
                    data = self.serial_conn.read(self.serial_conn.in_waiting or 1)
                elif self.comm_mode == 'tcp' and self.tcp_conn:
                    data = self.tcp_conn.recv(1024)
                else:
                    time.sleep(0.1)
                    continue
                    
                if data:
                    with self.buffer_lock:
                        self.recv_buffer.extend(data)
                    self._log_transmission(data, 'RX')
                    
            except Exception as e:
                print(f"Receive error: {e}")
                time.sleep(0.1)
                
    def get_received_data(self):
        """Get and clear received data buffer"""
        with self.buffer_lock:
            data = bytes(self.recv_buffer)
            self.recv_buffer.clear()
            return data
            
    def _log_transmission(self, data, direction):
        """Log communication activity"""
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        hex_str = ' '.join(f"{b:02X}" for b in data)
        log_entry = f"[{timestamp}] {direction}: {hex_str}\n"
        
        # Add to GUI log or file as needed
        print(log_entry)  # Replace with your logging mechanism
            
class VisualizationSystem:
    
    def __init__(self, root, tag_manager, anchor_manager):
        self.root = root
        self.tag_manager = tag_manager
        self.anchor_manager = anchor_manager
        
        self.figure_2d, self.ax2d = plt.subplots(figsize=(8, 6), dpi=100)
        self.figure_3d = plt.figure(figsize=(8, 6), dpi=100)
        self.ax3d = self.figure_3d.add_subplot(111, projection='3d')
        
        self.show_trajectory = True
        self.show_tag_labels = True
        self.show_axes = True
        self.show_ranging_circles = False
        self.tag_size = 10
        self.map_width = 500
        self.map_height = 500
        self.scale_factor = 1.0
        self.mouse_drag_start = None
        self.view_offset = [0, 0]
        self.zoom_level = 1.0
        
        
        # --- UI Control Variables (now part of the class) ---
        self.grid2d_var = tk.BooleanVar(value=True)
        self.xlim_2d = tk.StringVar(value="10") # Sensible default
        self.ylim_2d = tk.StringVar(value="10")
        self.color_var_2d = tk.StringVar(value="white")
        self.scale_var_2d = tk.StringVar(value="linear")
        
        self.grid3d_var = tk.BooleanVar(value=True)
        self.xlim_3d = tk.StringVar(value="10")
        self.ylim_3d = tk.StringVar(value="10")
        self.zlim_3d = tk.StringVar(value="5")
        self.color_var_3d = tk.StringVar(value="white")
        self.scale_var_3d_x = tk.StringVar(value="linear")
        self.scale_var_3d_y = tk.StringVar(value="linear")
        self.scale_var_3d_z = tk.StringVar(value="linear")
        
        self.bg_colors = ["white", "black", "lightgray", "lightblue", "beige"]
        self._init_artists()
        
    def _init_artists(self):
        """Initializes all the Matplotlib artists (lines, scatter plots, etc.)."""
        # --- 2D Artists ---
        self.line2d, = self.ax2d.plot([], [], 'bo-', label='Tag Path', markersize=4)
        self.tag_ab = AnnotationBbox(car_img, (0, 0), frameon=False, zorder=10)
        self.tag_ab.set_visible(False)
        self.ax2d.add_artist(self.tag_ab)
        self.target_ab = AnnotationBbox(target_img, (0, 0), frameon=False, zorder=5)
        self.target_ab.set_visible(False)
        self.ax2d.add_artist(self.target_ab)

        self.anchor_artists_2d = {}
        for anchor_id, pos in self.anchor_manager.anchor_positions.items():
            ab = AnnotationBbox(anchor_img, (pos['x'], pos['y']), frameon=False)
            text = self.ax2d.text(pos['x'], pos['y'] + 0.2, f"{anchor_id}", ha='center', va='bottom', fontsize=8)
            self.anchor_artists_2d[anchor_id] = {'ab': ab, 'text': text}
            self.ax2d.add_artist(ab)

        # --- 3D Artists ---
        self.line3d, = self.ax3d.plot([], [], [], 'bo-', label='Tag Path', alpha=0.5, markersize=3)
        self.tag_scatter_3d = self.ax3d.scatter([], [], [], c='red', s=100, label='Tag', zorder=10)
        anchor_ids = list(self.anchor_manager.anchor_positions.keys())
        self.anchor_scatter_3d = self.ax3d.scatter([], [], [], c='green', s=100, marker='s', label='Anchors')
        self.anchor_texts_3d = [self.ax3d.text(0, 0, 0, f"{anchor_id}") for anchor_id in anchor_ids]    
    
    def setup_gui_frames(self, frame2d_parent, frame3d_parent):
        """Creates and places the plot canvases and all controls into the parent Tkinter frames."""
        
        #======================================================================
        # 2D FRAME SETUP
        #======================================================================
                
        canvas2d = FigureCanvasTkAgg(self.figure_2d, master=frame2d_parent)
        canvas2d.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        canvas2d.mpl_connect('button_press_event', self.handle_target_click)
        
        toolbar2d = NavigationToolbar2Tk(canvas2d, frame2d_parent)
        toolbar2d.update()

        controls_2d_frame = ttk.Frame(frame2d_parent)
        controls_2d_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # --- 2D Controls ---
        ttk.Checkbutton(controls_2d_frame, text="Show Grid", variable=self.grid2d_var).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(controls_2d_frame, text="Scale:").pack(side=tk.LEFT)
        scale_combo_2d = ttk.Combobox(controls_2d_frame, textvariable=self.scale_var_2d, values=["linear", "log"], width=6)
        scale_combo_2d.pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(controls_2d_frame, text="BG Color:").pack(side=tk.LEFT)
        color_combo_2d = ttk.Combobox(controls_2d_frame, textvariable=self.color_var_2d, values=self.bg_colors, width=10)
        color_combo_2d.pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(controls_2d_frame, text="X Limit:").pack(side=tk.LEFT)
        ttk.Entry(controls_2d_frame, textvariable=self.xlim_2d, width=5).pack(side=tk.LEFT, padx=2)
        
        ttk.Label(controls_2d_frame, text="Y Limit:").pack(side=tk.LEFT)
        ttk.Entry(controls_2d_frame, textvariable=self.ylim_2d, width=5).pack(side=tk.LEFT, padx=2)

        ttk.Button(controls_2d_frame, text="Autoscale", command=self.toggle_2d_autoscale).pack(side=tk.LEFT, padx=5)
        
        # Bind callbacks for interactive updates
        self.xlim_2d.trace_add("write", self.update_2d_view_settings)
        self.ylim_2d.trace_add("write", self.update_2d_view_settings)
        scale_combo_2d.bind("<<ComboboxSelected>>", self.update_2d_view_settings)
        color_combo_2d.bind("<<ComboboxSelected>>", self.update_2d_view_settings)

        #======================================================================
        # 3D FRAME SETUP
        #======================================================================
            
        canvas3d = FigureCanvasTkAgg(self.figure_3d, master=frame3d_parent)
        canvas3d.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        toolbar3d = NavigationToolbar2Tk(canvas3d, frame3d_parent)
        toolbar3d.update()
        
        # --- 3D Controls (View Angle) ---
        view_frame_3d = ttk.LabelFrame(frame3d_parent, text="3D View Angle")
        view_frame_3d.pack(fill=tk.X, padx=10, pady=5)
        
        self.elev_slider = tk.Scale(view_frame_3d, from_=0, to=90, orient=tk.HORIZONTAL, label="Elevation")
        self.elev_slider.set(20)
        self.elev_slider.pack(fill=tk.X, expand=True)
        
        self.azim_slider = tk.Scale(view_frame_3d, from_=-180, to=180, orient=tk.HORIZONTAL, label="Azimuth")
        self.azim_slider.set(120)
        self.azim_slider.pack(fill=tk.X, expand=True)

        # --- 3D Controls (Limits, Color) ---
        controls_3d_frame = ttk.Frame(frame3d_parent)
        controls_3d_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Checkbutton(controls_3d_frame, text="Show Grid", variable=self.grid3d_var).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(controls_3d_frame, text="BG Color:").pack(side=tk.LEFT)
        color_combo_3d = ttk.Combobox(controls_3d_frame, textvariable=self.color_var_3d, values=self.bg_colors, width=10)
        color_combo_3d.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(controls_3d_frame, text="X Scale:").pack(side=tk.LEFT)
        scale_combo_3d_x = ttk.Combobox(controls_3d_frame, textvariable=self.scale_var_3d_x, values=["linear", "log"], width=6)
        scale_combo_3d_x.pack(side=tk.LEFT)
        ttk.Label(controls_3d_frame, text="Y Scale:").pack(side=tk.LEFT)
        scale_combo_3d_y = ttk.Combobox(controls_3d_frame, textvariable=self.scale_var_3d_y, values=["linear", "log"], width=6)
        scale_combo_3d_y.pack(side=tk.LEFT)
        ttk.Label(controls_3d_frame, text="Z Scale:").pack(side=tk.LEFT)
        scale_combo_3d_z = ttk.Combobox(controls_3d_frame, textvariable=self.scale_var_3d_z, values=["linear", "log"], width=6)
        scale_combo_3d_z.pack(side=tk.LEFT)
        
        ttk.Label(controls_3d_frame, text="X:").pack(side=tk.LEFT)
        xlim_entry_3d = ttk.Entry(controls_3d_frame, textvariable=self.xlim_3d, width=5)
        xlim_entry_3d.pack(side=tk.LEFT)
        ttk.Label(controls_3d_frame, text="Y:").pack(side=tk.LEFT)
        ylim_entry_3d = ttk.Entry(controls_3d_frame, textvariable=self.ylim_3d, width=5)
        ylim_entry_3d.pack(side=tk.LEFT)
        ttk.Label(controls_3d_frame, text="Z:").pack(side=tk.LEFT)
        zlim_entry_3d = ttk.Entry(controls_3d_frame, textvariable=self.zlim_3d, width=5)
        zlim_entry_3d.pack(side=tk.LEFT)
        
        ttk.Button(controls_3d_frame, text="Autoscale", command=self.toggle_3d_autoscale).pack(side=tk.LEFT, padx=5)

        # Bind callbacks for interactive updates
        self.xlim_3d.trace_add("write", self.update_3d_view_settings)
        self.ylim_3d.trace_add("write", self.update_3d_view_settings)
        self.zlim_3d.trace_add("write", self.update_3d_view_settings)
        color_combo_3d.bind("<<ComboboxSelected>>", self.update_3d_view_settings)
        
    def start_animation(self):
        """Starts the animation loops for both plots."""
        self.ani2d = FuncAnimation(self.figure_2d, self._update_plots, init_func=self._init_plots, interval=100, blit=False)
        self.ani3d = FuncAnimation(self.figure_3d, self._update_plots, init_func=self._init_plots, interval=100, blit=False)
        
    def _init_plots(self):
        """Initializes the static elements of the plots for the animation."""
        self.ax2d.set_title("2D Tracking View")
        self.ax2d.set_xlabel("X (m)")
        self.ax2d.set_ylabel("Y (m)")
        self.ax2d.grid(True)
        self.ax2d.legend(loc='upper right', fontsize=8)

        self.ax3d.set_title("3D Tracking View")
        self.ax3d.set_xlabel("X (m)")
        self.ax3d.set_ylabel("Y (m)")
        self.ax3d.set_zlabel("Z (m)")
        self.ax3d.grid(True)
        self.ax3d.legend(loc='upper right', fontsize=8)

        # Return all dynamic artists
        artists = [self.line2d, self.tag_ab, self.target_ab, self.line3d, self.tag_scatter_3d, self.anchor_scatter_3d]
        artists.extend(a['ab'] for a in self.anchor_artists_2d.values())
        artists.extend(a['text'] for a in self.anchor_artists_2d.values())
        artists.extend(self.anchor_texts_3d)
        return artists    
    
    def _update_plots(self, i):
        """The main animation callback function, now with fixes for anchors and 3D plotting."""
        # --- 1. Update Anchors (This was missing) ---
        # This needs to run every frame to reflect any changes in anchor positions.
        current_anchors = self.anchor_manager.get_anchor_positions()
        anchor_x = [pos['x'] for pos in current_anchors.values()]
        anchor_y = [pos['y'] for pos in current_anchors.values()]
        anchor_z = [pos['z'] for pos in current_anchors.values()]

        # Update 2D Anchor Artists
        for anchor_id, pos in current_anchors.items():
            if anchor_id in self.anchor_artists_2d:
                self.anchor_artists_2d[anchor_id]['ab'].xybox = (pos['x'], pos['y'])
                self.anchor_artists_2d[anchor_id]['text'].set_position((pos['x'], pos['y'] + 0.2))

        # Update 3D Anchor Artists
        self.anchor_scatter_3d._offsets3d = (anchor_x, anchor_y, anchor_z)
        for i, (anchor_id, pos) in enumerate(current_anchors.items()):
            if i < len(self.anchor_texts_3d):
                self.anchor_texts_3d[i].set_position((pos['x'], pos['y'], pos['z']))
                self.anchor_texts_3d[i].set_text(anchor_id)
        
        # --- 2. Update Active Tag ---
        # Get data from the managers, NOT global variables
        active_tag_id = self.tag_manager.active_tag_id # Use the correct attribute
        active_tag = self.tag_manager.find_tag(active_tag_id)
        
        # If there's no active tag or it has no history, we can't draw it.
        if not active_tag or not active_tag.movement_history:
            return self._init_plots()
        
        # We must return all the artists for blitting to work correctly.
        all_artists = [self.line2d, self.tag_ab, self.target_ab, self.line3d, self.tag_scatter_3d, self.anchor_scatter_3d]
        all_artists.extend(a['ab'] for a in self.anchor_artists_2d.values())
        all_artists.extend(a['text'] for a in self.anchor_artists_2d.values())
        all_artists.extend(self.anchor_texts_3d)
        return all_artists
        
        # --- 3. Get Tag Data ---
        history = list(active_tag.movement_history)
        hist_x = [pos[0] for pos in history]
        hist_y = [pos[1] for pos in history]
        hist_z = [pos[2] for pos in history]
    
        # This check is important! If history is empty, the next lines will error.
        if not hist_x:
            return self._init_plots()
    
        current_x, current_y, current_z = hist_x[-1], hist_y[-1], hist_z[-1]

        # --- 4. --- Update 2D Plot ---
        self.line2d.set_data(hist_x, hist_y)
        self.tag_ab.xybox = (current_x, current_y)
        self.tag_ab.set_visible(True)
        
        if self.tag_manager.target_position:
            self.target_ab.xybox = self.tag_manager.target_position
            self.target_ab.set_visible(True)
        else:
            self.target_ab.set_visible(False)
    
        # --- 5. --- Update 3D Plot ---
        self.line3d.set_data(hist_x, hist_y) 
        self.line3d.set_3d_properties(hist_z) 
        self.tag_scatter_3d._offsets3d = ([current_x], [current_y], [current_z])
        
        # --- 6. Handle Autoscaling ---
        if self.ax2d.get_autoscalex_on():
            self.ax2d.relim()
            self.ax2d.autoscale_view()
            self.ax2d.grid(self.grid2d_var.get())
            
        if self.ax3d.get_autoscalex_on():
            self.ax3d.relim()
            self.ax3d.autoscale_view()
            self.ax3d.grid(self.grid3d_var.get())
            self.ax3d.view_init(elev=self.elev_slider.get(), azim=self.azim_slider.get())

        return self._init_plots()
        
    def setup_3d_controls(self, parent):
        """Create 3D control panel"""
        control_frame = ttk.LabelFrame(parent, text="3D View Controls")
        
        # Axis ranges
        self.x_range = tk.Scale(control_frame, from_=1, to=100, 
                               orient=tk.HORIZONTAL, label="X Range (m)")
        self.x_range.set(50)
        self.x_range.pack(fill=tk.X)
        self.y_range = tk.Scale(control_frame, from_=1, to=100, 
                               orient=tk.HORIZONTAL, label="Y Range (m)")
        self.y_range.set(50)
        self.y_range.pack(fill=tk.Y)
        self.z_range = tk.Scale(control_frame, from_=1, to=100, 
                               orient=tk.HORIZONTAL, label="Z Range (m)")
        self.z_range.set(50)
        self.z_range.pack(fill=tk.Z)
        
        # Grid steps
        self.x_step = tk.Scale(control_frame, from_=0.1, to=10, resolution=0.1,
                              orient=tk.HORIZONTAL, label="X Grid Step (m)")
        self.x_step.set(1)
        self.x_step.pack(fill=tk.X)
        self.y_step = tk.Scale(control_frame, from_=0.1, to=10, resolution=0.1,
                              orient=tk.HORIZONTAL, label="Y Grid Step (m)")
        self.y_step.set(1)
        self.y_step.pack(fill=tk.Y)
        self.z_step = tk.Scale(control_frame, from_=0.1, to=10, resolution=0.1,
                              orient=tk.HORIZONTAL, label="Z Grid Step (m)")
        self.z_step.set(1)
        self.z_step.pack(fill=tk.Z)

        
        # Trajectory length
        self.traj_length = tk.Scale(control_frame, from_=10, to=1000,
                                   orient=tk.HORIZONTAL, label="Trajectory Points")
        self.traj_length.set(100)
        self.traj_length.pack(fill=tk.X)
        
        # Apply button
        ttk.Button(control_frame, text="Apply Settings",
                  command=self.apply_3d_settings).pack(pady=5)
        return control_frame
    
    def apply_3d_settings(self):
        """Apply 3D view settings"""
        self.ax_3d.set_xlim(0, self.x_range.get())
        self.ax_3d.set_ylim(0, self.y_range.get())
        self.ax_3d.set_zlim(0, self.z_range.get())
        
        # Update grid
        self.ax_3d.set_xticks(np.arange(0, self.x_range.get()+1, self.x_step.get()))
        self.ax_3d.set_yticks(np.arange(0, self.y_range.get()+1, self.y_step.get()))
        self.ax_3d.set_zticks(np.arange(0, self.z_range.get()+1, self.z_step.get()))
        
        # Update trajectory length
        for tag in self.tag_manager.tags.values():
            tag.movement_history = deque(maxlen=self.traj_length.get())
        
        self.figure_3d.canvas.draw_idle()
    
    def update_2d_view_settings(self, *args):
        """Callback to apply 2D view settings from UI controls."""
        try:
            # Stop autoscaling to allow manual limits
            self.ax2d.autoscale(enable=False)
            self.ax2d.set_xlim(0, float(self.xlim_2d.get()))
            self.ax2d.set_ylim(0, float(self.ylim_2d.get()))
            self.ax2d.set_xscale(self.scale_var_2d.get())
            self.ax2d.set_yscale(self.scale_var_2d.get())
            self.ax2d.set_facecolor(self.color_var_2d.get())
            # The grid is already handled by the animation loop reading the checkbox var
            self.fig2d.canvas.draw_idle()
        except (ValueError, TclError):
            # Ignore errors from invalid/empty entry widgets during typing
            pass

    def update_3d_view_settings(self, *args):
        """Callback to apply 3D view settings from UI controls."""
        try:
            # Stop autoscaling to allow manual limits
            self.ax3d.autoscale(enable=False)
            self.ax3d.set_xlim(0, int(self.xlim_3d.get()))
            self.ax3d.set_ylim(0, int(self.ylim_3d.get()))
            self.ax3d.set_zlim(0, int(self.zlim_3d.get()))
            self.ax3d.set_xscale(self.scale_var_3d_x.get())
            self.ax3d.set_yscale(self.scale_var_3d_y.get())
            self.ax3d.set_zscale(self.scale_var_3d_z.get())
            self.ax3d.set_facecolor(self.color_var_3d.get())
            # The grid and view angle are handled by the animation loop
            self.fig3d.canvas.draw_idle()
        except (ValueError, TclError):
            # Ignore errors from invalid/empty entry widgets during typing
            pass

    def toggle_2d_autoscale(self, *args):
        """Enable or disable autoscaling on the 2D plot."""
        self.ax2d.autoscale(enable=True)
        self.ax2d.relim()
        self.ax2d.autoscale_view()
        self.fig2d.canvas.draw_idle()

    def toggle_3d_autoscale(self, *args):
        """Enable or disable autoscaling on the 3D plot."""
        self.ax3d.autoscale(enable=True)
        self.ax3d.relim()
        self.ax3d.autoscale_view()
        self.fig3d.canvas.draw_idle()
    
    def setup_mouse_handlers(self, canvas):
        """Set up mouse event handlers"""
        canvas.mpl_connect('button_press_event', self.on_mouse_press)
        canvas.mpl_connect('button_release_event', self.on_mouse_release)
        canvas.mpl_connect('motion_notify_event', self.on_mouse_move)
        canvas.mpl_connect('scroll_event', self.on_mouse_scroll)
        
    def on_mouse_press(self, event):
        """Handle mouse button press"""
        if event.button == 1:  # Left click
            self.mouse_drag_start = (event.xdata, event.ydata)
            
    def on_mouse_move(self, event):
        """Handle mouse movement"""
        if event.button == 1 and self.mouse_drag_start:
            dx = event.xdata - self.mouse_drag_start[0]
            dy = event.ydata - self.mouse_drag_start[1]
            self.view_offset[0] += dx
            self.view_offset[1] += dy
            self.mouse_drag_start = (event.xdata, event.ydata)
            self._update_2d_plot()
            
    def on_mouse_release(self, event):
        """Handle mouse button release"""
        if event.button == 1:  # Left click
            self.mouse_drag_start = None
            
    def on_mouse_scroll(self, event):
        """Handle mouse scroll for zooming"""
        if event.button == 'up':
            self.zoom_level *= 1.1
        elif event.button == 'down':
            self.zoom_level /= 1.1
        self._update_2d_plot()   
    
    def handle_target_click(self, event):
        """Handle Matplotlib mouse clicks on the 2D plot to set target coordinates."""
        # Check if the click was inside the plot axes
        if event.inaxes == self.ax2d:
            x = event.xdata
            y = event.ydata
            
            # Update the target in the tag_manager and log it
            tag_manager.set_target(x, y)
            safe_log_insert(log_area, f"Target set to: ({x:.1f}, {y:.1f})\n")
            
            # Update target entry fields if they exist
            if hasattr(root, 'target_x_entry') and hasattr(root, 'target_y_entry'):
                # This check is important because the navigation window might not be open
                try:
                    root.target_x_entry.delete(0, tk.END)
                    root.target_x_entry.insert(0, f"{x:.1f}")
                    root.target_y_entry.delete(0, tk.END)
                    root.target_y_entry.insert(0, f"{y:.1f}")
                except tk.TclError:
                    # This handles the case where the nav window has been closed
                    print("Navigation window not available to set target.")

class NavigationController:
    def __init__(self, tag_manager):
        self.tag_manager = tag_manager
        self.target_position = None
        self.navigation_mode = "manual"  # or "auto"
        self.max_speed = 0.5  # m/s
        self.current_speed_level = 3  # 1-5
        
    def set_navigation_target(self, x, y, tag_id=None):
        """Set target position for navigation"""
        if tag_id is None:
            tag_id = self.tag_manager.active_tag
        if tag_id not in self.tag_manager.tags:
            return False
            
        self.target_position = (x, y)
        tag = self.tag_manager.tags[tag_id]
        tag.target = self.target_position
        return True
    
    def start_auto_navigation(self, tag_id=None):
        """Start autonomous navigation to target"""
        if tag_id is None:
            tag_id = self.tag_manager.active_tag
        if tag_id not in self.tag_manager.tags:
            return False
            
        tag = self.tag_manager.tags[tag_id]
        if not tag.target:
            return False
            
        self.navigation_mode = "auto"
        # Calculate path and send commands to move tag
        # This would interface with your actual hardware control
        
    def send_movement_command(self, direction, tag_id=None):
        """Send directional movement command"""
        if tag_id is None:
            tag_id = self.tag_manager.active_tag
        if tag_id not in self.tag_manager.tags:
            return False
            
        direction_map = {
            'forward': {'x': 0, 'y': 1},
            'backward': {'x': 0, 'y': -1},
            'left': {'x': -1, 'y': 0},
            'right': {'x': 1, 'y': 0},
            'forward_left': {'x': -0.7, 'y': 0.7},
            'forward_right': {'x': 0.7, 'y': 0.7},
            'backward_left': {'x': -0.7, 'y': -0.7},
            'backward_right': {'x': 0.7, 'y': -0.7}
        }
        
        if direction in direction_map:
            # This would interface with your actual hardware control
            # For now just update the tag's simulated position
            tag = self.tag_manager.tags[tag_id]
            move = direction_map[direction]
            speed_factor = self.current_speed_level / 5.0
            tag.x += move['x'] * speed_factor * 0.1
            tag.y += move['y'] * speed_factor * 0.1
            return True
        return False
    
    def calculate_path(self, tag):
        """Calculate path to target using A* or simple vector"""
        if not tag.target:
            return
            
        # Simple vector approach for demo
        dx = tag.target[0] - tag.x
        dy = tag.target[1] - tag.y
        distance = (dx**2 + dy**2)**0.5
    
        if distance < 0.1:  # Reached target
            tag.status = "At Target"
        return
        
        # Normalize direction vector
        tag.nav_direction = (dx/distance, dy/distance)
        tag.status = "Navigating"
    
class EnhancedNavigationController(NavigationController):
    def __init__(self, tag_manager):
        super().__init__(tag_manager)
        self.navigation_state = "idle"  # idle, manual, auto
        self.target_position = None
        self.current_speed = 50  # cm/s
        self.movement_duration = 200  # ms
        
    def set_navigation_target(self, x, y, tag_id=None):
        """Set target coordinates in cm"""
        if tag_id is None:
            tag_id = self.tag_manager.active_tag
            
        tag = self.tag_manager.find_tag(tag_id)
        if not tag:
            return False
            
        self.target_position = (x, y)
        tag.target = self.target_position
        self.navigation_state = "auto"
        return True
        
    def send_movement_command(self, direction, tag_id=None):
        """Enhanced movement command handling"""
        if tag_id is None:
            tag_id = self.tag_manager.active_tag
            
        direction_map = {
            'FL': {'x': -0.7, 'y': 0.7},   # Forward-Left
            'FS': {'x': 0, 'y': 1},        # Forward-Straight
            'FR': {'x': 0.7, 'y': 0.7},    # Forward-Right
            'BL': {'x': -0.7, 'y': -0.7},  # Backward-Left
            'BS': {'x': 0, 'y': -1},       # Backward-Straight
            'BR': {'x': 0.7, 'y': -0.7}    # Backward-Right
        }
        
        if direction in direction_map:
            movement = direction_map[direction]
            movement['duration'] = self.movement_duration
            movement['speed'] = self.current_speed
            
            if self.send_nav_command(tag_id, 'move', movement):
                self.safe_log(f"Sent {direction} to tag {tag_id}\n")
                return True
        return False

class ConnectionState:
    DISCONNECTED = 0
    CONNECTING = 1
    CONNECTED = 2
    WRONG_VERSION = 3

class WorkState:
    IDLE = 0
    RTLS_START = 1
    RTLSING = 2

class IMUState:
    NO_CONNECT = 0
    RUNNING = 1
    CALIBING = 2

class UIStateManager:
    def __init__(self, root):
        self.root = root
        self.connection_state = ConnectionState.DISCONNECTED
        self.work_state = WorkState.IDLE
        self.imu_state = IMUState.NO_CONNECT
        
    def update_ui(self):
        """Update UI elements based on current state"""
        # Connection state
        if self.connection_state == ConnectionState.DISCONNECTED:
            self.disable_controls()
            self.status_label.config(text="Disconnected", bg='red')
        elif self.connection_state == ConnectionState.CONNECTED:
            self.enable_basic_controls()
            self.status_label.config(text="Connected", bg='green')
        
        # Work state
        if self.work_state == WorkState.IDLE:
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
        elif self.work_state == WorkState.RTLSING:
            self.start_button.config(state=tk.DISABLED)
            self.stop_button.config(state=tk.NORMAL)

    def disable_controls(self):
        """Disable all controls"""
        for widget in self.control_widgets:
            widget.config(state=tk.DISABLED)

class Anchor:
    def __init__(self, id, x=0, y=0, z=0, enabled=True):
        self.id = id
        self.x = x
        self.y = y
        self.z = z
        self.enabled = enabled
        self.distance = 0
        self.last_update = 0
        self.distance_history = deque(maxlen=100)
        
    def update_position(self, x, y, z):
        self.x = x
        self.y = y
        self.z = z
        self.last_update = time.time()
        
    def update_distance(self, distance):
        self.distance = distance
        self.distance_history.append(distance)
        self.last_update = time.time()

class SerialHandler:
    def __init__(self, port=None, baudrate=115200):
        self.serial_port = serial.Serial()
        self.port = port
        self.baudrate = baudrate
        self.timeout = 0.075  # 75ms timeout like in C# code
        self.write_timeout = 0.1  # 100ms write timeout
        self.is_open = False
        self.buffer = bytearray()
        self.lock = threading.Lock()
        
    def open(self):
        try:
            if not self.port:
                raise ValueError("No port specified")
                
            self.serial_port.port = self.port
            self.serial_port.baudrate = self.baudrate
            self.serial_port.bytesize = serial.EIGHTBITS
            self.serial_port.parity = serial.PARITY_NONE
            self.serial_port.stopbits = serial.STOPBITS_ONE
            self.serial_port.timeout = self.timeout
            self.serial_port.write_timeout = self.write_timeout
            
            self.serial_port.open()
            self.is_open = True
            return True
        except Exception as e:
            print(f"Serial open failed: {e}")
            return False
            
    def close(self):
        try:
            if self.is_open:
                self.serial_port.close()
            self.is_open = False
            return True
        except Exception as e:
            print(f"Serial close failed: {e}")
            return False
            
    def read_data(self):
        if not self.is_open:
            return None
            
        try:
            with self.lock:
                # Read all available data
                data = self.serial_port.read(self.serial_port.in_waiting or 1)
                if data:
                    self.buffer.extend(data)
                    return bytes(data)
            return None
        except Exception as e:
            print(f"Serial read error: {e}")
            return None
            
    def write_data(self, data):
        if not self.is_open:
            return False
            
        try:
            with self.lock:
                self.serial_port.write(data)
                return True
        except Exception as e:
            print(f"Serial write error: {e}")
            return False
            
    def clear_buffers(self):
        try:
            with self.lock:
                self.serial_port.reset_input_buffer()
                self.serial_port.reset_output_buffer()
                self.buffer = bytearray()
            return True
        except Exception as e:
            print(f"Clear buffers failed: {e}")
            return False

class RTLSManager:
    def __init__(self):
        self.anchors = [Anchor(chr(65 + i)) for i in range(16)]  # Anchors A-P
        self.tags = []
        self.active_tag_index = 0
        self.work_state = "Idle"  # Idle, RtlsStart, Rtlsing, RtlsStop
        self.connect_state = "Disconnected"  # Disconnected, Connecting, Connected
        self.connect_mode = "Unknown"  # USB, TCP, Unknown
        
    def init_rtls(self):
        """Initialize RTLS system like in C# code"""
        # Clear existing data
        for anchor in self.anchors:
            anchor.distance_history.clear()
            
        self.tags.clear()
        self.work_state = "RtlsStart"
        
    def start_rtls(self):
        """Start RTLS positioning"""
        if self.work_state != "Idle":
            return False
            
        # Send start command (MODBUS function code 0x10)
        modbus_data = {
            'id': self.active_tag_index,
            'function': 0x10,
            'addr': 0x3B,  # Module mode address
            'reg_num': 1,
            'data': [0x00, 0x04]  # Continuous positioning
        }
        
        if self.send_modbus_command(modbus_data):
            self.work_state = "Rtlsing"
            return True
        return False
        
    def stop_rtls(self):
        """Stop RTLS positioning"""
        if self.work_state != "Rtlsing":
            return False
            
        modbus_data = {
            'id': self.active_tag_index,
            'function': 0x10,
            'addr': 0x3B,
            'reg_num': 1,
            'data': [0x00, 0x00]  # Stop positioning
        }
        
        if self.send_modbus_command(modbus_data):
            self.work_state = "Idle"
            return True
        return False

class ModbusRTU:
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.modbus_id = 1
            cls._instance.function_code = 0
            cls._instance.addr = 0
            cls._instance.reg_num = 0
        return cls._instance
        
    def create_read_command(self, modbus_id, function_code, addr, reg_num):
        """Create MODBUS read command (function code 0x03)"""
        self.modbus_id = modbus_id
        self.function_code = function_code
        self.addr = addr
        self.reg_num = reg_num
        
        cmd = bytearray(6)
        cmd[0] = modbus_id
        cmd[1] = function_code
        cmd[2] = (addr >> 8) & 0xFF
        cmd[3] = addr & 0xFF
        cmd[4] = (reg_num >> 8) & 0xFF
        cmd[5] = reg_num & 0xFF
        
        crc = self.calculate_crc(cmd)
        cmd.extend(crc)
        return cmd
        
    def create_write_command(self, modbus_id, function_code, addr, data):
        """Create MODBUS write command (function code 0x10)"""
        self.modbus_id = modbus_id
        self.function_code = function_code
        self.addr = addr
        self.reg_num = len(data) // 2
        
        cmd = bytearray(7 + len(data))
        cmd[0] = modbus_id
        cmd[1] = function_code
        cmd[2] = (addr >> 8) & 0xFF
        cmd[3] = addr & 0xFF
        cmd[4] = (self.reg_num >> 8) & 0xFF
        cmd[5] = self.reg_num & 0xFF
        cmd[6] = len(data)
        cmd[7:7+len(data)] = data
        
        crc = self.calculate_crc(cmd)
        cmd.extend(crc)
        return cmd
        
    def calculate_crc(self, data):
        """Calculate MODBUS CRC-16 checksum"""
        crc = 0xFFFF
        for byte in data:
            crc ^= byte
            for _ in range(8):
                if crc & 0x0001:
                    crc >>= 1
                    crc ^= 0xA001
                else:
                    crc >>= 1
        return bytes([crc & 0xFF, (crc >> 8) & 0xFF]) 

class IMUConfig:
    def __init__(self):
        self.accel_range = 16  # ±16g
        self.gyro_range = 2000  # ±2000dps
        self.sample_rate = 100  #Hz
        self.output_rate = 100  #Hz
        self.orientation = "horizontal"  # or "vertical"
        self.output_flags = {
            'accel': True,
            'gyro': True,
            'euler': True,
            'temp': True,
            'quat': True,
            'mag': False
        }
        
    def send_config_to_device(self):
        """Send IMU configuration to device"""
        if not client_socket:
            return False
            
        try:
            # Convert config to command string
            cmd = f"IMU_CFG {self.accel_range} {self.gyro_range} {self.sample_rate} "
            cmd += f"{self.output_rate} {self.orientation} "
            cmd += " ".join(["1" if v else "0" for v in self.output_flags.values()])
            
            client_socket.sendall(cmd.encode())
            return True
        except Exception as e:
            print(f"IMU config error: {e}")
            return False
            
    def calibrate(self):
        """Start IMU calibration"""
        if client_socket:
            client_socket.sendall(b"IMU_CALIBRATE")
            return True
        return False        

class AnchorCalibrationWindow(tk.Toplevel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.title("Auto Calibration")
        self.transient(parent)  # Keep window on top of the main app
        self.grab_set()         # Modal behavior
        self.geometry("800x600")

        # Initialize with default anchor positions
        self.anchor_positions = {
            'A': {'x': 0.0, 'y': 0.0, 'z': 1.2, 'enabled': tk.BooleanVar(value=True)},
            'B': {'x': 5.0, 'y': 0.0, 'z': 1.2, 'enabled': tk.BooleanVar(value=True)},
            'C': {'x': 5.0, 'y': 5.0, 'z': 1.2, 'enabled': tk.BooleanVar(value=True)},
            'D': {'x': 0.0, 'y': 5.0, 'z': 1.2, 'enabled': tk.BooleanVar(value=True)},
            'E': {'x': 0.0, 'y': 0.0, 'z': 0.0, 'enabled': tk.BooleanVar(value=False)},
            'F': {'x': 0.0, 'y': 0.0, 'z': 0.0, 'enabled': tk.BooleanVar(value=False)}
        }

        self.calibration_state = "idle"  # idle, calibrating, done
        self.calibration_results = {}
        self.init_ui()
        self.start_position_updater()

        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
    def init_ui(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Control buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        self.start_btn = ttk.Button(btn_frame, text="Start Calibration", command=self.start_calibration)
        self.stop_btn = ttk.Button(btn_frame, text="Stop Calibration", command=self.stop_calibration)
        self.confirm_btn = ttk.Button(btn_frame, text="Confirm Changes", command=self.confirm_changes)
        self.reset_btn = ttk.Button(btn_frame, text="Reset Coordinates", command=self.reset_coordinates)
        self.start_btn.pack(side=tk.LEFT, padx=2)
        self.stop_btn.pack(side=tk.LEFT, padx=2)
        self.confirm_btn.pack(side=tk.LEFT, padx=2)
        self.reset_btn.pack(side=tk.LEFT, padx=2)

        # Base station selection
        station_frame = ttk.Frame(main_frame)
        station_frame.pack(fill=tk.X, pady=5)
        ttk.Label(station_frame, text="Origin BS:").pack(side=tk.LEFT, padx=(0, 2))
        self.origin_combo = ttk.Combobox(station_frame, values=['A', 'B', 'C', 'D', 'E', 'F'], width=4)
        self.origin_combo.set('A')
        self.origin_combo.pack(side=tk.LEFT, padx=2)

        ttk.Label(station_frame, text="Direction:").pack(side=tk.LEFT, padx=(10, 2))
        self.direction_combo = ttk.Combobox(station_frame, values=['+X', '+Y', '-X', '-Y'], width=4)
        self.direction_combo.set('+X')
        self.direction_combo.pack(side=tk.LEFT, padx=2)

        ttk.Label(station_frame, text="Aux Point BS:").pack(side=tk.LEFT, padx=(10, 2))
        self.aux_combo = ttk.Combobox(station_frame, values=['A', 'B', 'C', 'D', 'E', 'F'], width=4)
        self.aux_combo.set('B')
        self.aux_combo.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(station_frame, text="Aux Point Axis:").pack(side=tk.LEFT, padx=(10, 2))
        self.axis_combo = ttk.Combobox(station_frame, values=['X', 'Y'], width=3)
        self.axis_combo.set('X')
        self.axis_combo.pack(side=tk.LEFT, padx=2)

        # Anchor position table (using a Treeview)
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        columns = ("#1", "#2", "#3", "#4") # Columns for data, #0 is for the checkbox tree
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        self.tree.heading("#0", text="Enable")
        self.tree.heading("#1", text="Base Station")
        self.tree.heading("#2", text="X (cm)")
        self.tree.heading("#3", text="Y (cm)")
        self.tree.heading("#4", text="Z (cm)")
        self.tree.column("#0", width=50, stretch=False)
        self.tree.column("#1", width=100, anchor=tk.CENTER)
        self.tree.column("#2", width=100, anchor=tk.CENTER)
        self.tree.column("#3", width=100, anchor=tk.CENTER)
        self.tree.column("#4", width=100, anchor=tk.CENTER)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.update_table()

        # Log display
        ttk.Label(main_frame, text="Log Display").pack(fill=tk.X)
        self.log_display = scrolledtext.ScrolledText(main_frame, height=8, wrap=tk.WORD)
        self.log_display.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        self.log_display.insert(tk.END, "Calibration logs will appear here...\n")
        self.log_display.config(state=tk.DISABLED)

        self.update_button_states()
        
    def update_table(self):
        # Clear existing rows
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Add current data
        for row, (anchor_id, data) in enumerate(self.anchor_positions.items()):
            # Using the 'text' property for the checkbox state visualization
            checked = "☑" if data['enabled'].get() else "☐"
            self.tree.insert("", "end", iid=anchor_id, text=checked, values=(
                f"Station {anchor_id}",
                f"{data['x']:.1f}",
                f"{data['y']:.1f}",
                f"{data['z']:.1f}"
            ))
        self.tree.bind("<Button-1>", self.toggle_anchor_checkbox)
    
    def toggle_anchor_checkbox(self, event):
        """Handle clicking on the checkbox area of the tree."""
        region = self.tree.identify("region", event.x, event.y)
        if region == "tree": # The checkbox area is the 'tree' region
            item_id = self.tree.identify_row(event.y)
            if item_id in self.anchor_positions:
                current_state = self.anchor_positions[item_id]['enabled'].get()
                self.anchor_positions[item_id]['enabled'].set(not current_state)
                self.log(f"Station {item_id} {'enabled' if not current_state else 'disabled'}")
                self.update_table()
    
    def start_position_updater(self):
        """Update positions from actual anchor data."""
        if self.calibration_state == "calibrating":
            for anchor_id in self.anchor_positions:
                if anchor_id in tag_manager.anchor_positions:
                    pos = tag_manager.anchor_positions[anchor_id]
                    self.anchor_positions[anchor_id]['x'] = pos['x'] * 100  # m to cm
                    self.anchor_positions[anchor_id]['y'] = pos['y'] * 100
                    self.anchor_positions[anchor_id]['z'] = pos['z'] * 100
            self.update_table()
        # Schedule the next update
        self.after(100, self.start_position_updater)
    
    def update_positions(self):
        """Update positions from actual anchor data"""
        if self.calibration_state == "calibrating":
            # Get live data from tag_manager
            for anchor_id in self.anchor_positions:
                if anchor_id in tag_manager.anchor_positions:
                    pos = tag_manager.anchor_positions[anchor_id]
                    self.anchor_positions[anchor_id]['x'] = pos['x'] * 100  # Convert m to cm
                    self.anchor_positions[anchor_id]['y'] = pos['y'] * 100
                    self.anchor_positions[anchor_id]['z'] = pos['z'] * 100
        
        self.update_table()
    
    def start_calibration(self):
        """Start auto-calibration process by writing parameters to MODBUS registers."""
        if not anchor_instrument:
            self.log("Error: No anchor connection")
            return

        try:
            # --- 1. Get and Validate GUI Selections ---
            origin_anchor_str = self.origin_combo.get()
            direction_str = self.direction_combo.get()
            aux_anchor_str = self.aux_combo.get()
            aux_axis_str = self.axis_combo.get()
            
            if origin_anchor_str == aux_anchor_str:
                self.log("Error: Origin and Auxiliary anchors must be different")
                return

            # --- 2. Translate Selections to Integer Values for MODBUS ---
            # From the global scope of your Python file:
            # MODBUS_ADDR_POSITIONING_CONTROL = 59  (which is 0x3B)
            # MODBUS_VAL_START_POSITIONING = 4      (which is 0x04)
            
            # --- 3. Send MODBUS Commands ---
            with anchor_lock:
                self.log(f"Sending 'Start Positioning' command to register {MODBUS_ADDR_POSITIONING_CONTROL}...")
            
                # Use write_register, which typically uses function code 6 or 16.
                # This is the most common way to write a single value.
                anchor_instrument.write_register(
                    registeraddress=MODBUS_ADDR_POSITIONING_CONTROL,
                    value=MODBUS_VAL_START_POSITIONING,
                    functioncode=6  # Explicitly use function code 6 (Write Single Register)
                    )
            
                self.log("Start command sent successfully.")

            # --- 4. Update GUI State ---
            self.calibration_state = "calibrating"
            self.update_button_states()
            self.log(f"Calibration initiated with Origin: {origin_anchor_str}, Aux: {aux_anchor_str}.")
            self.log("Monitoring for results...")


        except minimalmodbus.NoResponseError:
            self.log("Error: No response from the device. Check wiring, MODBUS ID, and power.")
            self.log("Ensure the device is not in a mode that prevents communication.")
            self.calibration_state = "idle"
            self.update_button_states()
        except Exception as e:
            self.log(f"Calibration start failed: {str(e)}")
            self.calibration_state = "idle"
            self.update_button_states()
    
    def stop_calibration(self):
        """Stop the calibration process"""
        try:
            with anchor_lock:
                # Send MODBUS command to stop calibration (register 0x3B with value 0x00)
                anchor_instrument.write_register(0x3B, 0x00)
                self.calibration_state = "done"
                self.log("Calibration stopped")
                self.update_button_states()
                
                # Process results
                self.process_calibration_results()
                
        except Exception as e:
            self.log(f"Error stopping calibration: {str(e)}")
    
    def process_calibration_results(self):
        """Process and validate the calibration results"""
        valid_count = 0
        for anchor_id, pos in self.anchor_positions.items():
            if pos['enabled']:
                # Basic validation
                if pos['x'] != 0 or pos['y'] != 0:
                    valid_count += 1
                    self.log(f"Station {anchor_id}: X={pos['x']:.1f}, Y={pos['y']:.1f}")
                else:
                    self.log(f"Warning: Station {anchor_id} has zero coordinates")
        
        if valid_count >= 3:  # Minimum 3 anchors needed for positioning
            self.log("Calibration completed successfully")
        else:
            self.log("Warning: Insufficient valid anchor positions")
    
    def confirm_changes(self):
        """Save the calibrated positions to the system"""
        for anchor_id, data in self.anchor_positions.items():
            if anchor_id in tag_manager.anchor_positions:
                # Convert cm back to meters for main system
                tag_manager.anchor_positions[anchor_id]['x'] = data['x'] / 100
                tag_manager.anchor_positions[anchor_id]['y'] = data['y'] / 100
                tag_manager.anchor_positions[anchor_id]['z'] = data['z'] / 100
        
        self.log("Anchor positions updated in system")
        self.calibration_state = "idle"
        self.update_button_states()
    
    def reset_coordinates(self):
        """Reset to default positions"""
        self.anchor_positions = {
            'A': {'x': 0.0, 'y': 0.0, 'z': 1.2, 'enabled': True},
            'B': {'x': 5.0, 'y': 0.0, 'z': 1.2, 'enabled': True},
            'C': {'x': 5.0, 'y': 5.0, 'z': 1.2, 'enabled': True},
            'D': {'x': 0.0, 'y': 5.0, 'z': 1.2, 'enabled': True},
            'E': {'x': 0.0, 'y': 0.0, 'z': 0.0, 'enabled': False},
            'F': {'x': 0.0, 'y': 0.0, 'z': 0.0, 'enabled': False}
        }
        self.update_table()
        self.log("Coordinates reset to defaults")
        self.calibration_state = "idle"
        self.update_button_states()
    
    def update_button_states(self):
        """Enable/disable buttons based on calibration state."""
        self.start_btn.config(state="normal" if self.calibration_state == "idle" else "disabled")
        self.stop_btn.config(state="normal" if self.calibration_state == "calibrating" else "disabled")
        self.confirm_btn.config(state="normal" if self.calibration_state == "done" else "disabled")
        self.reset_btn.config(state="normal")
    
    def log(self, message):
        """Add timestamped message to log."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_display.config(state=tk.NORMAL)
        self.log_display.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_display.see(tk.END) # Auto-scroll
        self.log_display.config(state=tk.DISABLED)
    
    def on_close(self):
        """Handle window closing."""
        if self.calibration_state == "calibrating":
            if messagebox.askyesno("Calibration in Progress",
                                "Calibration is running. Are you sure you want to quit?"):
                self.stop_calibration()
                self.destroy()
        else:
            self.destroy()


# === Functions ===
# def is_admin():
#     try:
#         return ctypes.windll.shell32.IsUserAnAdmin()
#     except:
#         return False

# if not is_admin():
#     ctypes.windll.shell32.ShellExecuteW(
#         None, "runas", sys.executable, " ".join(sys.argv), None, 1)
#     sys.exit()
    
def safe_log_insert(log_area, text):
    if not shutdown_flag:
        log_area.after(0, lambda: log_area.insert(tk.END, text))

def handle_client(conn, addr, log_area):
    global client_socket
    client_socket = conn
    
    # Initialize log file
    if not os.path.exists("log.csv"):
        with open("log.csv", "w", newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["Time", "X", "Y", "Z", "Angle", "Channel"])  # Added Channel column
    
    with conn:
        safe_log_insert(log_area, f"Connected to Raspberry Pi at {addr}\n")
        current_channel = 2  # Default channel
        
        while True:
            try:
                data = conn.recv(1024)
                if not data:
                    break
                
                if tag_manager.process_modbus_data(data):
                    # Update tag position if we have valid MODBUS data
                    if tag_manager.last_cal_data['x'] != 0 or tag_manager.last_cal_data['y'] != 0:
                        tag_data = {
                            'id': str(tag_manager.last_cal_data['Tag_ID']),
                            'x': tag_manager.last_cal_data['x'] / 100.0,  # Convert cm to m
                            'y': tag_manager.last_cal_data['y'] / 100.0,
                            'z': tag_manager.last_cal_data['z'] / 100.0,
                            'angle': 0,  # You may need to calculate this
                            'channel': current_channel
                        }
                        tag_manager.update_tag(tag_data)
                    continue  # Skip JSON processing if MODBUS was successful
                
                              
                # Handle channel switching command
                if data.decode().strip().lower().startswith("set_channel"):
                    try:
                        msg = data.decode().strip()
                        channel = int(msg.split()[1])
                        if ser_uwb and ser_uwb.is_open:
                            ser_uwb.write(f"chan {channel}\r\n".encode())
                        else:
                            safe_log_insert(log_area, "UWB serial not connected!\n")
                        if channel in [1, 2, 3, 4, 5, 7]:  # Valid UWB channels
                            ser_uwb.write(f"chan {channel}\r\n".encode())
                            time.sleep(0.2)  # Increased delay for channel switching
                            current_channel = channel
                            conn.sendall(f"Channel set to {channel}\n".encode())
                            safe_log_insert(log_area, f"UWB channel changed to {channel}\n")
                        else:
                            conn.sendall(b"Invalid channel (valid: 1-5,7)\n")
                    except Exception as e:
                        conn.sendall(f"Channel error: {str(e)}\n".encode())
                    continue  # Skip normal processing for commands
                
                if data.startswith(b'IMU_DATA'):  # Or whatever prefix your IMU data uses
                    imu_data = tag_manager.process_imu_data(data)
                    tag_id = "default"  # Or extract from data if available
                    if tag_id in tag_manager.tags:
                        tag_manager.tags[tag_id].update_imu_data(imu_data)    
                
                # Normal position data processing
                try:
                    payload = json.loads(data.decode())
                    ts = datetime.now().strftime("%H:%M:%S")
                    
                    tag_id = payload.get('tag_id', 'default')
                    x = float(payload.get('x', 0))
                    y = float(payload.get('y', 0))
                    z = float(payload.get('z', 0))
                    angle = float(payload.get('angle', 0))
                    channel = int(payload.get('channel', current_channel)) # Get channel from payload
                    
                    # Extract position data (with channel info)
                    tag_data = {
                        'id': tag_id,
                        'x': x,
                        'y': y,
                        'z': z,
                        'angle': angle,
                        'channel': channel
                        }   
                    tag_manager.update_tag(tag_data)
                                        
                    if 'type' in payload and payload['type'] == 'nav_response':
                        if payload['command'] == 'get_speed':
                            print(f"Tag {payload['tag_id']} speed: {payload['speed']} cm/s")
                        elif payload['command'] == 'position_update':
                            # Update your tag position display
                            pass
                        return  # Skip normal processing for navigation responses
                    
                    if not tag_manager.filter_multipath(x, y, z):
                        safe_log_insert(log_area, f"Filtered multipath position: ({x},{y},{z})\n")
                        continue

                    with data_lock:
                        # Get or create tag object
                        tag_id = payload.get('tag_id', 'default')  # Make sure your payload includes tag_id
                        tag = tag_manager.tags.setdefault(tag_id, UWBTag(id=tag_id))
                        
                        # Store previous position for calculations
                        prev_position = (tag.x, tag.y, tag.z)
                        
                        # Update tag data
                        tag.x = x
                        tag.y = y
                        tag.z = z
                        tag.angle = angle
                        tag.channel = channel
                        tag.last_update = time.time()
                        tag.movement_history.append((x, y, z, angle, time.time()))
                        
                        validate_position(x, y, z)
                        # Update data buffers for visualization (optional - you might want per-tag buffers)
                        tag_manager.update_or_create_tag(tag_data)
                        angle_data.append(angle)
                        time_stamps.append(ts)
                        
                        # Update statistics
                        global last_position, last_time, total_distance, total_time
                        now = datetime.now()
                        
                        if prev_position[0] is not None:  # Check if we had a previous position
                            dx = x - prev_position[0]
                            dy = y - prev_position[1]
                            dz = z - prev_position[2]
                            dt = (now - last_time).total_seconds()
                            distance_step = (dx**2 + dy**2 + dz**2) ** 0.5
                            
                            if dt > 0:
                                tag.velocity = distance_step / dt
                                total_time += dt
                            else:
                                tag.velocity = 0.0
            
                            tag.distance_traveled += distance_step
                            total_distance += distance_step
                            avg_speed = total_distance / total_time if total_time > 0 else 0.0
                            heading = np.degrees(np.arctan2(dy, dx)) % 360
                            
                            # Update GUI variables - you might want to make these tag-specific
                            root.after(0, lambda: speed_var.set(f"Speed: {tag.velocity:.2f} u/s (Ch{channel})"))
                            avg_speed_var.set(f"Avg Speed: {avg_speed:.2f} u/s")
                            distance_var.set(f"Distance: {tag.distance_traveled:.2f} u")
                            heading_var.set(f"Heading: {heading:.1f}°")
                            
                            # Update tag status based on movement
                            tag.status = "Moving" if tag.velocity > 0.1 else "Stationary"
    
                        last_position = (x, y, z)
                        last_time = now
    
                        # Update position displays
                        x_var.set(f"X: {x:.2f}")
                        y_var.set(f"Y: {y:.2f}")
                        z_var.set(f"Z: {z:.2f}")
                        channel_var.set(f"Channel: {channel}")
    
                        # Update tag-specific GUI elements if you have them
                        if hasattr(tag_manager, 'update_tag_display'):
                            tag_manager.update_tag_display(tag_id)

                    # Log the data
                    log_line = f"{ts} - [Ch{channel}] X: {x}, Y: {y}, Angle: {angle}°\n"
                    safe_log_insert(log_area, log_line)
                    
                    # Write to CSV (now includes channel)
                    with open("log.csv", "a", newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow([ts, x, y, z, angle, channel])

                except (json.JSONDecodeError, ValueError) as e:
                    safe_log_insert(log_area, f"Data error: {e}\n")
                    
            except (ConnectionResetError, BrokenPipeError) as e:
                safe_log_insert(log_area, f"Client disconnected: {str(e)}\n")
                break
                
    client_socket = None
    
def start_server(log_area):
    global shutdown_flag
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    server.bind((HOST, PORT))
    server.listen(1)
    safe_log_insert(log_area, f"Server listening on {HOST}:{PORT}\n")

    while not shutdown_flag:
        try:
            server.settimeout(1)
            conn, addr = server.accept()
            threading.Thread(target=handle_client, args=(conn, addr, log_area), daemon=True).start()
        except socket.timeout:
            continue
    server.close()

def send_command(entry_field):
    global client_socket
    if client_socket:
        msg = entry_field.get()
        if msg:
            try:
                client_socket.sendall(msg.encode())
                entry_field.delete(0, tk.END)
            except:
                pass

def scan_com_ports():
    """Enhanced COM port scanning with device identification"""
    ports = []
    available_ports = serial.tools.list_ports.comports()
    
    if not available_ports:
        safe_log_insert(log_area, "No serial ports detected on this system.\n")
        com_dropdown['values'] = []
        com_port_var.set("")
        return

    safe_log_insert(log_area, f"Found ports: {[port.device for port in available_ports]}. Checking for PGPlus devices...\n")
    
    for port in available_ports:
        # The is_pgplus_device function now provides feedback if it fails
        if is_pgplus_device(port.device):
            ports.append(port.device)
            
    if ports:
        com_dropdown['values'] = ports
        com_port_var.set(ports[0])
        safe_log_insert(log_area, f"Found compatible PGPlus devices: {', '.join(ports)}\n")
    else:
        safe_log_insert(log_area, "Scan complete. No responsive PGPlus devices found.\n")
        com_dropdown['values'] = []
        com_port_var.set("")    
    
def is_pgplus_device(port):
    """
    Try to identify PGPlus device by sending a MODBUS query.
    Returns True if successful, False otherwise.
    Includes detailed error logging.
    """
    instrument = None
    try:
        # Initialize the modbus instrument for the check
        instrument = minimalmodbus.Instrument(port, 1) # slaveaddress 1
        instrument.serial.baudrate = 115200
        instrument.serial.timeout = 0.2  # 200ms timeout
        instrument.mode = minimalmodbus.MODE_RTU
        instrument.clear_buffers_before_each_transaction = True

        # Attempt to read a register. Register 0 is common for version info.
        # This is the line that probes the device.
        version = instrument.read_register(0, functioncode=3) 
        
        safe_log_insert(log_area, f"Device on {port} responded. Version: {version}.\n")
        return True # The device responded successfully
        
    except minimalmodbus.NoResponseError:
        # This is the most common error: the port is valid but the device didn't answer the MODBUS query.
        # It's not a PGPlus device or it's in a non-responsive state.
        # No log is needed here as it's expected for non-target devices.
        return False
    except serial.SerialException as e:
        # This error means the port could not be opened (e.g., access denied, already in use).
        safe_log_insert(log_area, f"Could not open port {port}: {e}\n")
        return False
    except Exception as e:
        # Catch any other unexpected errors during the check.
        safe_log_insert(log_area, f"An unexpected error occurred while checking {port}: {e}\n")
        return False
    finally:
        # Ensure the serial port is closed after the check
        if instrument and instrument.serial.is_open:
            instrument.serial.close()
           
def update_anchor_state(key, value):
    global anchor_state
    with anchor_state_lock:
        anchor_state[key] = value
    # This needs to be scheduled on the main thread
    root.after(0, update_status_display)

def update_status_display():
    status_text = (f"Connected: {'Yes' if anchor_state['connected'] else 'No'} | "
                  f"Positioning: {'Active' if anchor_state['positioning'] else 'Inactive'} | "
                  f"Network: {anchor_state['network_id'] or 'Unknown'} | "
                  f"Rate: {anchor_state['update_rate'] or 'Unknown'}ms")
    anchor_status_var.set(status_text)

def toggle_serial_connection():
    """Connects/disconnects to the anchor using Modbus RTU."""
    global anchor_instrument
    with anchor_lock:
        if anchor_instrument and anchor_instrument.serial.is_open:
            try:
                send_to_major_anchor('stop_pos')
                anchor_instrument.serial.close()
                safe_log_insert(log_area, f"Disconnected from {com_port_var.get()}\n")
            except Exception as e:
                safe_log_insert(log_area, f"Error during disconnect: {e}\n")
            finally:
                anchor_instrument = None
                update_anchor_state('connected', False)
                update_anchor_state('positioning', False)
                connect_button.config(text="Connect")
                com_status_var.set("Disconnected")
        else:
            try:
                port = com_port_var.get()
                slave_id = int(config_vars['MODBUS-ID'].get())

                # Create the Modbus instrument
                anchor_instrument = minimalmodbus.Instrument(
                    port=port,
                    slaveaddress=slave_id,
                    close_port_after_each_call=False,
                    debug=False
                    )

                # Configure serial settings based on C# analysis
                anchor_instrument.serial.baudrate = 115200
                anchor_instrument.serial.bytesize = 8
                anchor_instrument.serial.parity = serial.PARITY_NONE
                anchor_instrument.serial.stopbits = 1
                anchor_instrument.serial.timeout = 0.1  # 100 ms timeout
                anchor_instrument.mode = minimalmodbus.MODE_RTU # Explicitly set RTU mode
                anchor_instrument.clear_buffers_before_each_transaction = True

                # Test connection by reading a register (e.g., a version register, address 0 is common)
                version = anchor_instrument.read_register(0, functioncode=3)
                safe_log_insert(log_area, f"Successfully connected to anchor on {port}.\n")
                safe_log_insert(log_area, f"Anchor Firmware Version (Reg 0): {version}\n")
                
                update_anchor_state('connected', True)
                connect_button.config(text="Disconnect")
                com_status_var.set(f"Connected to {port}")
                
                # Read initial configuration
                threading.Thread(target=read_anchor_config, daemon=True).start()
                
            except minimalmodbus.NoResponseError:
                safe_log_insert(log_area, "No response from anchor - check wiring/power\n")
                com_status_var.set("Connection Failed (No Response)")
            except serial.SerialException as e:
                safe_log_insert(log_area, f"Serial error: {str(e)}\n")
                com_status_var.set("Connection Failed")
            except Exception as e:
                safe_log_insert(log_area, f"MODBUS connection failed: {e}\n")
                anchor_instrument = None
                com_status_var.set("Connection Failed")

def query_anchor_info():
    if send_to_major_anchor('get_info'):
        # Parse response and update state
        # (Implementation depends on your anchor's response format)
        pass

def send_to_major_anchor(command, params=None, retries=3, timeout=0.5):
    """Improved anchor command sending with retries"""
    if not anchor_instrument or not anchor_instrument.serial.is_open:
        safe_log_insert(log_area, "Serial connection not established\n")
        return False
        
    try:
        for attempt in range(retries):
            try:
                with anchor_lock:
                    if command == 'start_pos':
                        anchor_instrument.write_register(
                            MODBUS_ADDR_POSITIONING_CONTROL,
                            MODBUS_VAL_START_POSITIONING
                        )
                    elif command == 'stop_pos':
                        anchor_instrument.write_register(
                            MODBUS_ADDR_POSITIONING_CONTROL,
                            MODBUS_VAL_STOP_POSITIONING
                        )
                    elif command == 'set_network':
                        anchor_instrument.write_register(
                            0x20,  # Network ID register
                            params['id']
                        )
                    elif command == 'set_rate':
                        anchor_instrument.write_register(
                            0x82,  # Update rate register
                            params['rate']
                        )
                    # Add other commands as needed
                    
                    return True
                    
            except minimalmodbus.ModbusException as e:
                if attempt == retries - 1:
                    safe_log_insert(log_area, f"Failed: {str(e)}\n")
                    raise
                time.sleep(0.1)
                
    except Exception as e:
        safe_log_insert(log_area, f"Command '{command}' failed: {str(e)}\n")
        return False

def send_nav_command(tag_id, command, params=None):
    """Send a navigation command to the Raspberry Pi"""
    if not client_socket:
        print("No client connected")
        return False
    
    try:
        message = {
            'type': 'nav_command',
            'tag_id': tag_id,
            'command': command,
            'params': params or {}
        }
        client_socket.sendall(json.dumps(message).encode())
        return True
    except Exception as e:
        print(f"Command failed: {str(e)}")
        return False

def get_tag_info(tag_id):
    """Get current info for a specific tag"""
    with data_lock:
        if tag_id in tag_manager.tags:
            tag = tag_manager.tags[tag_id]
            return {
                'x': tag.x,
                'y': tag.y,
                'z': tag.z,
                'angle': tag.angle,
                'status': 'Active' if time.time() - tag.last_update < 5 else 'Inactive'
            }
    return None

def add_tooltip(widget, text):
    """Add a tooltip to a widget."""
    tooltip = tk.Toplevel(widget)
    tooltip.withdraw()
    tooltip.wm_overrideredirect(True)
    
    def enter(event):
        x, y, _, _ = widget.bbox("insert")
        x += widget.winfo_rootx() + 25
        y += widget.winfo_rooty() + 25
        tooltip.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tooltip, text=text, bg="lightyellow", relief="solid", borderwidth=1)
        label.pack()
        tooltip.deiconify()
    
    def leave(event):
        tooltip.withdraw()
    
    widget.bind("<Enter>", enter)
    widget.bind("<Leave>", leave)

def on_close():
    global shutdown_flag
    shutdown_flag = True
    if client_socket:
        client_socket.close()
    if anchor_instrument:
        anchor_instrument.serial.close()
    root.destroy()
    
def send_channel_command(self):
    """Send channel selection command to Raspberry Pi"""
    channel = self.channel_var.get()
    if client_socket:
        try:
            client_socket.sendall(f"set_channel {channel}".encode())
            self.log_area.insert(tk.END, f"Set UWB channel to {channel}\n")
        except Exception as e:
            self.log_area.insert(tk.END, f"Failed to set channel: {e}\n")
            
def init_uwb_serial(port, baudrate=115200):
    global ser_uwb
    try:
        ser_uwb = serial.Serial(port=port, baudrate=baudrate, timeout=1)
        safe_log_insert(log_area, f"UWB serial initialized on {port}\n")
        return True
    except Exception as e:
        safe_log_insert(log_area, f"UWB serial error: {e}\n")
        return False
    
def refresh_serial():
    """Refresh the serial connection by re-scanning ports and reconnecting if needed."""
    global serial_connection, ser_uwb
    
    # 1. Re-scan COM ports
    scan_com_ports()
    
    # 2. If already connected, attempt to reconnect
    if serial_connection and serial_connection.is_open:
        try:
            serial_connection.close()
            time.sleep(0.5)  # Small delay before reconnecting
            serial_connection = serial.Serial(port=com_port_var.get(), baudrate=115200, timeout=1)
            safe_log_insert(log_area, "Serial connection refreshed.\n")
        except Exception as e:
            safe_log_insert(log_area, f"Failed to refresh serial: {e}\n")
            serial_connection = None
    
    # 3. Optional: Clear UWB serial buffer
    if ser_uwb and ser_uwb.is_open:
        ser_uwb.reset_input_buffer()
        safe_log_insert(log_area, "UWB serial buffer cleared.\n")

def clear_serial_buffer():
    """Clear the input buffer of the serial connection."""
    if serial_connection and serial_connection.is_open:
        serial_connection.reset_input_buffer()
        safe_log_insert(log_area, "Serial buffer cleared.\n")
    else:
        safe_log_insert(log_area, "No active serial connection to clear.\n")

def create_anchor_config_frame(self):
    """Create frame for anchor configuration"""
    config_frame = ttk.LabelFrame(self.root, text="Anchor Configuration")
    config_frame.pack(padx=10, pady=5, fill=tk.X)

    # Serial Port Settings
    ttk.Label(config_frame, text="Serial Port:").grid(row=0, column=0, sticky="w")
    self.port_entry = ttk.Entry(config_frame, width=10)
    self.port_entry.grid(row=0, column=1)
    self.port_entry.insert(0, "COM4")

    ttk.Label(config_frame, text="Baudrate:").grid(row=0, column=2, padx=5)
    self.baudrate_combo = ttk.Combobox(config_frame, values=[9600, 19200, 38400, 57600, 115200], width=7)
    self.baudrate_combo.grid(row=0, column=3)
    self.baudrate_combo.set(115200)

    # Anchor Settings
    settings = [
        ("MODBUS-ID", "0x01", 1, 1),
        ("Kalman-Q", "3", 2, 1),
        ("Kalman-R", "10", 2, 3),
        ("Antenna Delay", "33040", 3, 1),
        ("Device Type", ["Master", "Tag"], 3, 3),
        ("Device ID", "0", 4, 1),
        ("Positioning Mode", ["2D", "3D"], 4, 3),
        ("Ranging Mode", ["TWR", "HDS-TWR"], 5, 1),
        ("UWB Channel", ["1", "2", "3", "4", "5", "7"], 5, 3),
        ("Data Rate", ["110K", "850K", "6M8"], 6, 1)
    ]

    self.config_vars = {}
    for i, (label, default, row, col) in enumerate(settings):
        ttk.Label(config_frame, text=label+":").grid(row=row, column=col*2-2, sticky="w")
        if isinstance(default, list):  # Dropdown
            var = tk.StringVar()
            combo = ttk.Combobox(config_frame, textvariable=var, values=default, width=7)
            combo.grid(row=row, column=col*2-1)
            combo.set(default[0])
        else:  # Entry field
            var = tk.StringVar(value=default)
            entry = ttk.Entry(config_frame, textvariable=var, width=7)
            entry.grid(row=row, column=col*2-1)
        self.config_vars[label] = var

    # Action Buttons
    ttk.Button(config_frame, text="Load Config", command=self.load_config).grid(row=7, column=0, pady=5)
    ttk.Button(config_frame, text="Save Config", command=self.save_config).grid(row=7, column=1)
    ttk.Button(config_frame, text="Start Positioning", command=self.start_positioning).grid(row=7, column=2)
    ttk.Button(config_frame, text="Stop Positioning", command=self.stop_positioning).grid(row=7, column=3)

def load_config(self):
    """Load configuration from anchor"""
    try:
        # Example MODBUS read - implement based on your protocol
        response = send_to_major_anchor("nvm read 0x80")  # Sample command
        if response:
            # Parse response and update GUI
            self.config_vars["MODBUS-ID"].set(response.get("modbus_id", "0x01"))
            self.log_area.insert(tk.END, "Config loaded successfully\n")
    except Exception as e:
        self.log_area.insert(tk.END, f"Load failed: {str(e)}\n")

def save_config(self, config_vars=None, log_area=None):
    """Handle both GUI and programmatic config saves"""
    config_vars = config_vars or self.config_vars
    log_area = log_area or self.log_area
    
    try:
        device_type = config_vars['Device Type'].get().lower()
        type_map = {'major anchor': '1', 'sub anchor': '2', 'tag': '3'}
        device_type_value = type_map.get(device_type, '1')

        commands = [
            f"nvm 0x80 {config_vars['MODBUS-ID'].get()}",
            f"kalman_q {config_vars['Kalman-Q'].get()}",
            f"kalman_r {config_vars['Kalman-R'].get()}",
            f"antdly {config_vars['Antenna Delay'].get()}",
            f"mode {config_vars['Positioning Mode'].get().lower()}",
            f"chan {config_vars['UWB Channel'].get()}",
            f"rate {config_vars['Data Rate'].get()}",
            f"nvm 0x84 {device_type_value}",
            f"nvm 0x85 {config_vars['Device ID'].get()}",
            f"ranging {config_vars['Ranging Mode'].get().lower()}",
            f"nvm 0x86 {config_vars['Network ID'].get()}"
        ]

        for cmd in commands:
            if not send_to_major_anchor(cmd):
                raise Exception(f"Command failed: {cmd}")
                
        log_area.insert(tk.END, "Config saved to anchor\n")
        return True
        
    except Exception as e:
        log_area.insert(tk.END, f"Save failed: {str(e)}\n")
        return False

def start_positioning():
    """Start positioning cycle with multi-anchor coordination"""
    try:
        # Set network ID first (must match all anchors)
        if not send_to_major_anchor('set_network', {'id': config_vars['Network ID'].get()}):
            raise Exception("Failed to set network ID")
        
        # Set optimal data rate based on coverage area
        coverage_area = float(config_vars['Coverage Area'].get())  # Add this to your GUI config
        if coverage_area > 15:
            send_to_major_anchor('set_rate', {'rate': 110})  # 110kbps for long range
        else:
            send_to_major_anchor('set_rate', {'rate': 6800})  # 6.8Mbps for high precision

        # Start anchor sequence per HDS-TWR protocol
        commands = [
            ('start_pos', None),    # Anchor A
            ('trigger_B', 0.003),   # 3ms delay
            ('trigger_C', 0.003),
            ('trigger_D', 0.003)
        ]
        
        for cmd, delay in commands:
            if not send_to_major_anchor(cmd):
                raise Exception(f"Command {cmd} failed")
            if delay:
                time.sleep(delay)
                
        update_anchor_state('positioning', True)
        safe_log_insert(log_area, "HDS-TWR positioning started\n")
        return True
        
    except Exception as e:
        safe_log_insert(log_area, f"Start positioning failed: {str(e)}\n")
        return False

def stop_positioning():
    """Stop positioning"""
    if send_to_major_anchor('stop_pos'):
        update_anchor_state('positioning', False)
        safe_log_insert(log_area, "Positioning stopped\n")
    else:
        safe_log_insert(log_area, "Stop command failed\n")
        
def validate_position(x, y, z):
    """Validate position against physical constraints"""
    # Ground/obstruction check
    if z < 1.0:  # Minimum height per docs
        raise ValueError(f"Invalid height {z}m - must be ≥1m")
    
    # Origin rejection (common error state)
    if abs(x) < 0.1 and abs(y) < 0.1:
        raise ValueError("Position at origin - likely error")
    
    # Velocity sanity check (if available)
    global last_position, last_time
    if last_position and last_time:
        dt = (datetime.now() - last_time).total_seconds()
        if dt > 0:
            velocity = ((x-last_position[0])**2 + (y-last_position[1])**2)**0.5 / dt
            if velocity > 5.0:  # 5m/s = 18km/h
                raise ValueError(f"Implausible velocity: {velocity:.2f}m/s")
        
def import_config():
    """Import configuration from file"""
    filepath = filedialog.askopenfilename(filetypes=[("Config Files", "*.cfg"), ("All Files", "*.*")])
    if filepath:
        try:
            with open(filepath, 'r') as f:
                config = json.load(f)
                for key, var in config_vars.items():
                    if key in config:
                        var.set(config[key])
                safe_log_insert(log_area, f"Configuration imported from {filepath}\n")
        except Exception as e:
            safe_log_insert(log_area, f"Import failed: {str(e)}\n")

def show_navigation_controls():
    """Create the navigation control window"""
    nav_window = tk.Toplevel(root)
    nav_window.title("Navigation Control")
    nav_window.geometry("800x600")
    
    # Search Tags section
    search_frame = ttk.LabelFrame(nav_window, text="Search Tags")
    search_frame.pack(fill=tk.X, padx=10, pady=5)
    
    # Store entry widgets as attributes
    root.target_x_entry = ttk.Entry(nav_window, width=8)
    root.target_y_entry = ttk.Entry(nav_window, width=8)
    
    # Tags table
    columns = ("Tag ID", "x(cm)", "y(cm)", "z(cm)", "Status", "Angle", "Magnetic Field(nT)")
    tree = ttk.Treeview(search_frame, columns=columns, show="headings", height=5)
    
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=100, anchor=tk.CENTER)
    
    tree.pack(fill=tk.X, padx=5, pady=5)
    
    # Navigation Control section
    control_frame = ttk.LabelFrame(nav_window, text="Navigation Control")
    control_frame.pack(fill=tk.X, padx=10, pady=5)
    
    # Tag control row
    tag_frame = ttk.Frame(control_frame)
    tag_frame.pack(fill=tk.X, pady=5)
    
    ttk.Label(tag_frame, text="Tag ID:").pack(side=tk.LEFT, padx=2)
    tag_id_entry = ttk.Entry(tag_frame, width=8)
    tag_id_entry.pack(side=tk.LEFT, padx=2)
    
    ttk.Button(tag_frame, text="Change Work State", 
              command=lambda: change_work_state(tag_id_entry.get())).pack(side=tk.LEFT, padx=2)
    ttk.Button(tag_frame, text="Stop Movement", 
              command=lambda: stop_movement(tag_id_entry.get())).pack(side=tk.LEFT, padx=2)
    
    # Navigation config row
    config_frame = ttk.Frame(control_frame)
    config_frame.pack(fill=tk.X, pady=5)
    
    ttk.Button(config_frame, text="Navigation Config", 
              command=show_nav_config).pack(side=tk.LEFT, padx=2)
    ttk.Button(config_frame, text="Start Auto Navigation", 
              command=start_auto_nav).pack(side=tk.LEFT, padx=2)
    ttk.Button(config_frame, text="Start Control", 
              command=start_manual_control).pack(side=tk.LEFT, padx=2)
    
    # Speed control row
    speed_frame = ttk.Frame(control_frame)
    speed_frame.pack(fill=tk.X, pady=5)
    
    ttk.Label(speed_frame, text="Speed Level:").pack(side=tk.LEFT, padx=2)
    speed_level = ttk.Combobox(speed_frame, values=["1", "2", "3", "4", "5"], width=3)
    speed_level.pack(side=tk.LEFT, padx=2)
    speed_level.set("3")
    
    ttk.Label(speed_frame, text="Movement Duration(ms):").pack(side=tk.LEFT, padx=2)
    duration_entry = ttk.Entry(speed_frame, width=6)
    duration_entry.pack(side=tk.LEFT, padx=2)
    duration_entry.insert(0, "200")
    
    ttk.Button(speed_frame, text="Read Speed", 
              command=lambda: read_speed(tag_id_entry.get())).pack(side=tk.LEFT, padx=2)
    ttk.Button(speed_frame, text="Change Speed", 
              command=lambda: change_speed(tag_id_entry.get(), speed_level.get())).pack(side=tk.LEFT, padx=2)
    
    # Movement buttons
    move_frame = ttk.Frame(control_frame)
    move_frame.pack(fill=tk.X, pady=5)
    
    movement_buttons = [
        ("Forward-Left", "FL"),
        ("Forward-Straight", "FS"),
        ("Forward-Right", "FR"),
        ("Backward-Left", "BL"),
        ("Backward-Straight", "BS"),
        ("Backward-Right", "BR")
    ]
    
    for i, (text, cmd) in enumerate(movement_buttons):
        if i == 3:  # Start new row for backward movements
            move_frame = ttk.Frame(control_frame)
            move_frame.pack(fill=tk.X, pady=5)
        ttk.Button(move_frame, text=text, 
                 command=lambda c=cmd: send_movement_command(tag_id_entry.get(), c),
                 width=12).pack(side=tk.LEFT, padx=2)
    
    # Auto Navigation section
    auto_frame = ttk.LabelFrame(nav_window, text="Auto Navigation")
    auto_frame.pack(fill=tk.X, padx=10, pady=5)
    
    # Target coordinates
    target_frame = ttk.Frame(auto_frame)
    target_frame.pack(fill=tk.X, pady=5)
    
    ttk.Label(target_frame, text="Target X(cm):").pack(side=tk.LEFT, padx=2)
    target_x_entry = ttk.Entry(target_frame, width=8)
    target_x_entry.pack(side=tk.LEFT, padx=2)
    
    ttk.Label(target_frame, text="Target Y(cm):").pack(side=tk.LEFT, padx=2)
    target_y_entry = ttk.Entry(target_frame, width=8)
    target_y_entry.pack(side=tk.LEFT, padx=2)
    
    ttk.Button(target_frame, text="Start Movement", 
              command=lambda: start_movement_to_target(
                  tag_id_entry.get(),
                  target_x_entry.get(),
                  target_y_entry.get()
              )).pack(side=tk.LEFT, padx=2)
    
    ttk.Button(auto_frame, text="Set Target with Mouse", 
              command=set_target_with_mouse).pack(side=tk.LEFT, padx=2, pady=5)
    
    # Status bar
    status_frame = ttk.Frame(nav_window)
    status_frame.pack(fill=tk.X, padx=10, pady=5)
    
    status_var = tk.StringVar(value="System Running...")
    ttk.Label(status_frame, textvariable=status_var).pack(side=tk.LEFT)
    
    # Add sample data to table (replace with real data)
    sample_tags = [
        ("001", "120.5", "80.3", "0.0", "Active", "45.2", "52000"),
        ("002", "200.0", "150.0", "0.0", "Inactive", "90.0", "48000")
    ]
    
    for tag in sample_tags:
        tree.insert("", tk.END, values=tag)
    
    def set_target():
        try:
            x = float(target_x_entry.get())
            y = float(target_y_entry.get())
            tag_manager.set_target(x, y)
        except ValueError:
            safe_log_insert(log_area, "Error", "Invalid coordinates")
    
    ttk.Button(nav_window, text="Set Target", command=set_target).grid(row=0, column=4)
    
    # Click-to-set functionality
    def on_canvas_click(event):
        if event.inaxes:
            target_x_entry.delete(0, tk.END)
            target_x_entry.insert(0, f"{event.xdata:.2f}")
            target_y_entry.delete(0, tk.END)
            target_y_entry.insert(0, f"{event.ydata:.2f}")
            set_target()
    
    # Get the figure canvas and bind the click event
    canvas2d = vis_system.set_up_gui.frame_2d.canvas
    canvas2d.mpl_connect('button_press_event', on_canvas_click)

def change_work_state(tag_id):
    """Toggle between active/inactive states"""
    if not tag_id:
        print("Please enter a tag ID")
        return
    
    current_info = get_tag_info(tag_id)
    if not current_info:
        print(f"Tag {tag_id} not found")
        return
    
    new_state = 'standby' if current_info['status'] == 'Active' else 'active'
    if send_nav_command(tag_id, 'set_state', {'state': new_state}):
        print(f"Changed tag {tag_id} to {new_state} mode")
    else:
        print(f"Failed to change state for tag {tag_id}")

def stop_movement(tag_id):
    """Send emergency stop command"""
    if not tag_id:
        print("Please enter a tag ID")
        return
    
    if send_nav_command(tag_id, 'emergency_stop'):
        print(f"Stopped tag {tag_id}")
    else:
        print(f"Failed to stop tag {tag_id}")

def show_nav_config():
    """Show configuration dialog"""
    config_window = tk.Toplevel()
    config_window.title("Navigation Configuration")
    
    # Add your configuration options here
    ttk.Label(config_window, text="Navigation Parameters").pack(pady=10)
    
    # Example parameter
    ttk.Label(config_window, text="Max Speed (cm/s):").pack()
    speed_entry = ttk.Entry(config_window)
    speed_entry.pack()
    speed_entry.insert(0, "50")  # Default value
    
    def save_config():
        # Save configuration to file or send to device
        print(f"Configuration saved: Max Speed = {speed_entry.get()}")
        config_window.destroy()
    
    ttk.Button(config_window, text="Save", command=save_config).pack(pady=10)

def start_auto_nav():
    """Start autonomous navigation mode"""
    if send_nav_command('all', 'set_mode', {'mode': 'auto'}):
        print("Autonomous navigation started")
    else:
        print("Failed to start autonomous navigation")

def start_manual_control():
    """Start manual control mode"""
    if send_nav_command('all', 'set_mode', {'mode': 'manual'}):
        print("Manual control started")
    else:
        print("Failed to start manual control")

def read_speed(tag_id):
    """Request current speed from tag"""
    if not tag_id:
        print("Please enter a tag ID")
        return

def change_speed(tag_id, speed_level):
    """Set movement speed level"""
    if not tag_id:
        print("Please enter a tag ID")
        return
    
    speed_map = {'1': 10, '2': 25, '3': 50, '4': 75, '5': 100}  # cm/s
    
    if send_nav_command(tag_id, 'set_speed', {'speed': speed_map.get(speed_level, 50)}):
        print(f"Set tag {tag_id} speed to level {speed_level}")
    else:
        print(f"Failed to set speed for tag {tag_id}")

def send_movement_command(tag_id, direction):
    """Send directional movement command"""
    if not tag_id:
        print("Please enter a tag ID")
        return
    
    direction_map = {
        'FL': {'x': -0.7, 'y': 0.7},   # Forward-Left
        'FS': {'x': 0, 'y': 1},        # Forward-Straight
        'FR': {'x': 0.7, 'y': 0.7},    # Forward-Right
        'BL': {'x': -0.7, 'y': -0.7},  # Backward-Left
        'BS': {'x': 0, 'y': -1},       # Backward-Straight
        'BR': {'x': 0.7, 'y': -0.7}    # Backward-Right
    }
    
    if direction in direction_map:
        if send_nav_command(tag_id, 'move', direction_map[direction]):
            print(f"Sent {direction} command to tag {tag_id}")
        else:
            print(f"Failed to send movement command to tag {tag_id}")

def start_movement_to_target(tag_id, target_x, target_y):
    """Navigate to specific coordinates"""
    if not tag_id:
        print("Please enter a tag ID")
        return
    
    try:
        x = float(target_x)
        y = float(target_y)
        
        if send_nav_command(tag_id, 'navigate_to', {'x': x, 'y': y}):
            print(f"Tag {tag_id} navigating to ({x}, {y})")
        else:
            print(f"Failed to start navigation for tag {tag_id}")
    except ValueError:
        print("Invalid coordinates - please enter numbers") 
        
def set_target_with_mouse():
    """Set target by clicking on the 2D view"""
    print("Click on the 2D view to set target")
    
def auto_calibration():
    """Show the auto calibration window"""
    if not anchor_instrument:
        safe_log_insert(log_area, "Anchor connection not established\n")
        return
    
    calibration_window = AnchorCalibrationWindow(root)
    
    try:
        safe_log_insert(log_area, "Starting auto-calibration...\n")
        send_to_major_anchor("calib auto")
        safe_log_insert(log_area, "Auto-calibration initiated. Check anchor responses.\n")
    except Exception as e:
        safe_log_insert(log_area, f"Calibration error: {str(e)}\n")

def show_help(help_type):
    """Show help tooltips"""
    help_texts = {
        "hardware_coords": "When enabled, position calculations are done by the anchor hardware.\n"
                          "When disabled, calculations are done by this software.",
        "auto_position": "When enabled, the anchor will automatically start positioning\n"
                        "after power is applied (saved in NVM)."
    }
    
    help_window = tk.Toplevel(root)
    help_window.title("Help")
    help_window.geometry("400x100")
    ttk.Label(help_window, text=help_texts.get(help_type, "No help available")).pack(padx=10, pady=10)

def select_tag(tag_id):
    """Set the currently active tag for navigation controls"""
    tag_manager.active_tag = tag_id
    if tag_id in tag_manager.tags:
        tag = tag_manager.tags[tag_id]
        # Update UI to show this tag is selected
        print(f"Selected tag {tag_id} at ({tag.x}, {tag.y})")

def validate_anchor_positions(self):
    """Check against installation guidelines"""
    for name, (x, y, z) in self.anchor_positions.items():
        if z < 1.0:
            safe_log_insert(log_area, f"WARNING: Anchor {name} height {z}m < 1m minimum\n")
        if min(x, y) < 0.5:
            safe_log_insert(log_area, f"WARNING: Anchor {name} too close to walls\n")

def read_anchor_config():
    """Read configuration from anchor in a thread-safe way"""
    if not anchor_instrument:
        return
        
    try:
        with anchor_lock:
            # Read basic configuration registers
            config = {
                'MODBUS-ID': anchor_instrument.read_register(0x00, functioncode=3),
                'Kalman-Q': anchor_instrument.read_register(0x10, functioncode=3),
                'Kalman-R': anchor_instrument.read_register(0x11, functioncode=3),
                'Antenna_Delay': anchor_instrument.read_register(0x12, functioncode=3),
                'Device_Type': anchor_instrument.read_register(0x14, functioncode=3),
                'Device_ID': anchor_instrument.read_register(0x15, functioncode=3),
                'Positioning_Mode': anchor_instrument.read_register(0x16, functioncode=3),
                'Ranging_Mode': anchor_instrument.read_register(0x17, functioncode=3),
                'UWB_Channel': anchor_instrument.read_register(0x18, functioncode=3),
                'Data_Rate': anchor_instrument.read_register(0x19, functioncode=3),
                'Network_ID': anchor_instrument.read_register(0x20, functioncode=3)
            }
            
            # Update GUI variables on main thread
            root.after(0, lambda: update_config_gui(config))
            
    except Exception as e:
        safe_log_insert(log_area, f"Config read error: {str(e)}\n")

def update_config_gui(config):
    """Update GUI with received configuration"""
    try:
        config_vars['MODBUS-ID'].set(str(config['MODBUS-ID']))
        config_vars['Kalman-Q'].set(str(config['Kalman-Q']))
        config_vars['Kalman-R'].set(str(config['Kalman-R']))
        config_vars['Antenna Delay'].set(str(config['Antenna_Delay']))
        
        # Map numeric values to string options
        device_type_map = {1: 'Major Anchor', 2: 'Sub Anchor', 3: 'Tag'}
        config_vars['Device Type'].set(device_type_map.get(config['Device_Type'], 'Major Anchor'))
        
        config_vars['Device ID'].set(str(config['Device_ID']))
        
        pos_mode_map = {0: '2D', 1: '3D'}
        config_vars['Positioning Mode'].set(pos_mode_map.get(config['Positioning_Mode'], '3D'))
        
        range_mode_map = {0: 'TWR', 1: 'HDS-TWR', 2: 'TDOA'}
        config_vars['Ranging Mode'].set(range_mode_map.get(config['Ranging_Mode'], 'HDS-TWR'))
        
        config_vars['UWB Channel'].set(str(config['UWB_Channel']))
        
        data_rate_map = {0: '110K', 1: '850K', 2: '6M8'}
        config_vars['Data Rate'].set(data_rate_map.get(config['Data_Rate'], '6M8'))
        
        config_vars['Network ID'].set(str(config['Network_ID']))
        
    except Exception as e:
        safe_log_insert(log_area, f"GUI update error: {str(e)}\n")

def scan_modbus_ids(self):
    """Scan all possible MODBUS IDs (1-255) to find active anchors"""
    self.scan_results = []
    self.current_scan_id = 1
    self.scan_in_progress = True
    
    def scan_next():
        if self.current_scan_id <= 255 and self.scan_in_progress:
            try:
                # Try to read a register from each ID
                instrument = minimalmodbus.Instrument(
                    self.port, 
                    self.current_scan_id,
                    close_port_after_each_call=True
                )
                version = instrument.read_register(0, functioncode=3)
                self.scan_results.append(self.current_scan_id)
                self.safe_log(f"Found device at ID {self.current_scan_id} (Version: {version})\n")
            except:
                pass
            
            self.current_scan_id += 1
            self.root.after(100, scan_next)  # Schedule next scan
        else:
            self.scan_in_progress = False
            self.safe_log("MODBUS ID scan completed\n")
            if self.scan_results:
                self.update_id_dropdown(self.scan_results)
    
    scan_next()  # Start the scan

def write_config(self):
    """Enhanced configuration writing based on C# code"""
    config_bytes = bytearray(122)  # Match C# buffer size
    
    # Serial baudrate
    config_bytes[0:2] = self.config_vars['MODBUS_RATE'].get().to_bytes(2, 'big')
    
    # MODBUS ID
    config_bytes[2:4] = self.config_vars['MODBUS_ID'].get().to_bytes(2, 'big')
    
    # Ranging and positioning modes
    config_bytes[4] = self.config_vars['RANGING_MODE'].get()
    config_bytes[5] = self.config_vars['POSITIONING_MODE'].get()
    
    # Device type and ID
    config_bytes[6] = self.config_vars['DEVICE_TYPE'].get()
    config_bytes[8:10] = self.config_vars['DEVICE_ID'].get().to_bytes(2, 'big')
    
    # UWB channel and data rate
    config_bytes[10] = self.config_vars['UWB_CHANNEL'].get()
    config_bytes[11] = self.config_vars['DATA_RATE'].get()
    
    # Kalman filters
    config_bytes[12:14] = self.config_vars['KALMAN_Q'].get().to_bytes(2, 'big')
    config_bytes[14:16] = self.config_vars['KALMAN_R'].get().to_bytes(2, 'big')
    
    # Antenna delay
    config_bytes[16:18] = self.config_vars['ANTENNA_DELAY'].get().to_bytes(2, 'big')
    
    # Anchor enable flags and positions
    anchor_enable = 0
    for i, anchor in enumerate(self.anchor_manager.anchors.values()):
        if anchor['enabled']:
            anchor_enable |= (1 << i)
        pos = i * 6 + 22
        config_bytes[pos:pos+6] = struct.pack('>hhh', 
            int(anchor['x'] * 100),  # Convert to cm
            int(anchor['y'] * 100),
            int(anchor['z'] * 100))
    
    config_bytes[20:22] = anchor_enable.to_bytes(2, 'big')
    
    # Auto positioning
    config_bytes[118] = 8 if self.auto_position_var.get() else 9
    
    # Send via MODBUS
    try:
        self.anchor_instrument.write_registers(0, config_bytes)
        self.safe_log("Configuration written successfully\n")
    except Exception as e:
        self.safe_log(f"Config write failed: {str(e)}\n")

def create_navigation_controls(self, parent):
    """Create navigation control panel"""
    nav_frame = ttk.LabelFrame(parent, text="Navigation Control")
    
    # Target coordinates
    ttk.Label(nav_frame, text="Target X:").grid(row=0, column=0)
    self.target_x_entry = ttk.Entry(nav_frame, width=8)
    self.target_x_entry.grid(row=0, column=1)
    
    ttk.Label(nav_frame, text="Target Y:").grid(row=0, column=2)
    self.target_y_entry = ttk.Entry(nav_frame, width=8)
    self.target_y_entry.grid(row=0, column=3)
    
    ttk.Button(nav_frame, text="Set Target", 
              command=self.set_navigation_target).grid(row=0, column=4)
    
    # Movement buttons
    directions = ['FL', 'FS', 'FR', 'BL', 'BS', 'BR']
    for i, dir in enumerate(directions):
        btn = ttk.Button(nav_frame, text=dir, 
                        command=lambda d=dir: self.send_movement_command(d))
        btn.grid(row=1 + i//3, column=i%3)
    
    # Speed control
    ttk.Label(nav_frame, text="Speed:").grid(row=3, column=0)
    self.speed_combo = ttk.Combobox(nav_frame, values=["1", "2", "3", "4", "5"])
    self.speed_combo.grid(row=3, column=1)
    self.speed_combo.set("3")
    
    return nav_frame

def connect_serial(self, port, baudrate=115200):
    """Connect to serial port with error handling"""
    try:
        self.serial_conn = serial.Serial(
            port=port,
            baudrate=baudrate,
            timeout=1,
            write_timeout=1
        )
        self.safe_log(f"Connected to {port} at {baudrate} baud\n")
        self.connection_state = ConnectionState.CONNECTED
        return True
    except serial.SerialException as e:
        self.safe_log(f"Serial connection failed: {e}\n")
        self.connection_state = ConnectionState.DISCONNECTED
        return False

def clear_buffers(self):
    """Clear serial buffers"""
    if self.serial_conn and self.serial_conn.is_open:
        self.serial_conn.reset_input_buffer()
        self.serial_conn.reset_output_buffer()
        self.safe_log("Serial buffers cleared\n")

def update_modbus_id_display(value):
    """Update MODBUS ID display with hex representation"""
    hex_str = f"0x{value:02X}"
    modbus_id_var.set(hex_str)
    
def on_modbus_id_changed(event=None):
    """Handle MODBUS ID spinbox changes"""
    try:
        value = int(modbus_id_spinbox.get())
        if 1 <= value <= 255:
            update_modbus_id_display(value)
        else:
            modbus_id_spinbox.set(1)
            update_modbus_id_display(1)
    except ValueError:
        modbus_id_spinbox.set(1)
        update_modbus_id_display(1)

def update_map_origin():
    """Update map origin coordinates"""
    global map_origin_x, map_origin_y
    try:
        map_origin_x = int(origin_x_spinbox.get())
        map_origin_y = int(origin_y_spinbox.get())
        safe_log_insert(log_area, f"Map origin updated to ({map_origin_x}, {map_origin_y})\n")
    except ValueError:
        safe_log_insert(log_area, "Invalid origin coordinates\n")

def update_map_scale():
    """Update map scaling factor"""
    global map_multiple
    try:
        map_multiple = float(scale_spinbox.get())
        safe_log_insert(log_area, f"Map scale updated to {map_multiple}x\n")
    except ValueError:
        safe_log_insert(log_area, "Invalid scale factor\n")

def update_tag_count():
    """Adjust number of tags being tracked"""
    try:
        new_count = int(tag_count_spinbox.get())
        if 1 <= new_count <= max_tags:
            with data_lock:
                # Add or remove tags as needed
                current_count = len(tag_manager.tags)
                if new_count > current_count:
                    for i in range(current_count, new_count):
                        tag_id = str(i)
                        tag_manager.tags[tag_id] = UWBTag(id=tag_id)
                elif new_count < current_count:
                    for tag_id in list(tag_manager.tags.keys())[new_count:]:
                        del tag_manager.tags[tag_id]
                        
            tag_manager.update_tag_table()
            safe_log_insert(log_area, f"Tracking {new_count} tags\n")
        else:
            tag_count_spinbox.set(len(tag_manager.tags))
    except ValueError:
        tag_count_spinbox.set(len(tag_manager.tags))

def create_map_config_frame(parent):
    """Create frame for map configuration controls"""
    frame = ttk.LabelFrame(parent, text="Map Configuration")
    frame.pack(padx=10, pady=5, fill=tk.X)
    
    # Origin controls
    ttk.Label(frame, text="Origin X:").grid(row=0, column=0)
    origin_x_spinbox = ttk.Spinbox(frame, from_=-1000, to=1000, width=8)
    origin_x_spinbox.grid(row=0, column=1)
    origin_x_spinbox.set(map_origin_x)
    
    ttk.Label(frame, text="Origin Y:").grid(row=0, column=2)
    origin_y_spinbox = ttk.Spinbox(frame, from_=-1000, to=1000, width=8)
    origin_y_spinbox.grid(row=0, column=3)
    origin_y_spinbox.set(map_origin_y)
    
    ttk.Button(frame, text="Set Origin", 
              command=update_map_origin).grid(row=0, column=4, padx=5)
    
    # Scale control
    ttk.Label(frame, text="Scale Factor:").grid(row=1, column=0)
    scale_spinbox = ttk.Spinbox(frame, from_=0.1, to=10.0, increment=0.1, width=8)
    scale_spinbox.grid(row=1, column=1)
    scale_spinbox.set(map_multiple)
    
    ttk.Button(frame, text="Set Scale", 
              command=update_map_scale).grid(row=1, column=2, padx=5)
    
    origin_x_spinbox = origin_x_spinbox
    scale_spinbox = scale_spinbox
    
    return frame

def create_tag_management_frame(parent):
    """Create frame for tag management controls"""
    frame = ttk.LabelFrame(parent, text="Tag Management")
    frame.pack(padx=10, pady=5, fill=tk.X)
    
    # Tag count control
    ttk.Label(frame, text="Tag Count:").grid(row=0, column=0)
    tag_count_spinbox = ttk.Spinbox(frame, from_=1, to=max_tags, width=8)
    tag_count_spinbox.grid(row=0, column=1)
    tag_count_spinbox.set(len(tag_manager.tags))
    tag_count_spinbox.bind("<FocusOut>", lambda e: update_tag_count())
    
    # MODBUS ID control
    ttk.Label(frame, text="MODBUS ID:").grid(row=0, column=2)
    modbus_id_spinbox = ttk.Spinbox(frame, from_=1, to=255, width=8)
    modbus_id_spinbox.grid(row=0, column=3)
    modbus_id_spinbox.set(1)
    modbus_id_spinbox.bind("<FocusOut>", on_modbus_id_changed)
    
    ttk.Label(frame, text="Hex:").grid(row=0, column=4)
    modbus_id_var = tk.StringVar(value="0x01")
    ttk.Label(frame, textvariable=modbus_id_var).grid(row=0, column=5)
    
    # Store references to the spinboxes and variables
    modbus_id_spinbox = modbus_id_spinbox
    modbus_id_var = modbus_id_var
    tag_count_spinbox = tag_count_spinbox
    
    # Tag table (existing code)
    columns = ("ID", "X", "Y", "Z", "Angle", "Status", "Channel")
    tag_table = ttk.Treeview(frame, columns=columns, show="headings", height=5)
    for col in columns:
        tag_table.heading(col, text=col)
        tag_table.column(col, width=80, anchor=tk.CENTER)
    tag_table.grid(row=1, column=0, columnspan=6, pady=5)
    
    # Store references
    tag_manager.tag_table = tag_table
    return frame

def load_image(path, scale=0.1):
    img = imread(path)
    return OffsetImage(img, zoom=scale)

def export_to_excel(data_table, filename):
    """Export data to Excel file similar to C# implementation"""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add headers
        headers = [
            "Time", "X", "Y", "Z", 
            "A", "B", "C", "D", "E", "F", "G", "H",
            "I", "J", "K", "L", "M", "N", "O", "P",
            "Flag", "Velocity"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            
        # Add data
        for row_idx, row_data in enumerate(data_table, 2):
            for col_idx, value in enumerate(row_data, 1):
                if col_idx == 1:  # Time column
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
                else:
                    try:
                        num_value = float(value)
                        ws.cell(row=row_idx, column=col_idx, value=num_value)
                    except ValueError:
                        ws.cell(row=row_idx, column=col_idx, value=value)
                        
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Export failed: {e}")
        return False

def create_data_export_frame(parent):
    """Create frame for data export controls"""
    frame = ttk.LabelFrame(parent, text="Data Export")
    frame.pack(padx=10, pady=5, fill=tk.X)
    
    ttk.Button(frame, text="Export Channel 1 Data", 
              command=lambda: tag_manager.export_channel_data(1)).pack(side=tk.LEFT, padx=2)
    ttk.Button(frame, text="Export Channel 2 Data", 
              command=lambda: tag_manager.export_channel_data(2)).pack(side=tk.LEFT, padx=2)
    return frame

def create_trajectory_controls(parent):
    """Create frame for trajectory display controls"""
    frame = ttk.Frame(parent)
    frame.pack(padx=10, pady=5, fill=tk.X)
    
    ttk.Button(frame, textvariable=tag_manager.trajectory_button_text,
              command=tag_manager.toggle_trajectory).pack(side=tk.LEFT)
    return frame

def create_map_config_controls(parent):
    """Create map configuration controls"""
    frame = ttk.LabelFrame(parent, text="Map Configuration")
    frame.pack(padx=10, pady=5, fill=tk.X)
    
    # Map width
    ttk.Label(frame, text="Width:").grid(row=0, column=0)
    width_spin = ttk.Spinbox(frame, from_=100, to=10000, width=8)
    width_spin.grid(row=0, column=1)
    width_spin.set(500)
    
    # Map height
    ttk.Label(frame, text="Height:").grid(row=0, column=2)
    height_spin = ttk.Spinbox(frame, from_=100, to=10000, width=8)
    height_spin.grid(row=0, column=3)
    height_spin.set(500)
    
    # Tag size
    ttk.Label(frame, text="Tag Size:").grid(row=0, column=4)
    size_spin = ttk.Spinbox(frame, from_=1, to=50, width=8)
    size_spin.grid(row=0, column=5)
    size_spin.set(10)
    
    return frame

def create_data_transmission_frame(parent):
    """Create frame for data transmission controls"""
    frame = ttk.LabelFrame(parent, text="Data Transmission")
    frame.pack(padx=10, pady=5, fill=tk.X)
    
    # Data entry
    ttk.Label(frame, text="Data:").grid(row=0, column=0)
    data_entry = ttk.Entry(frame, width=20)
    data_entry.grid(row=0, column=1)
    
    # Hex mode checkbox
    hex_var = tk.BooleanVar()
    ttk.Checkbutton(frame, text="Hex", variable=hex_var).grid(row=0, column=2)
    
    # Tag ID
    ttk.Label(frame, text="Tag ID:").grid(row=0, column=3)
    tag_entry = ttk.Entry(frame, width=5)
    tag_entry.grid(row=0, column=4)
    
    # Interval for continuous transmission
    ttk.Label(frame, text="Interval (ms):").grid(row=0, column=5)
    interval_entry = ttk.Entry(frame, width=6)
    interval_entry.grid(row=0, column=6)
    
    # Buttons
    ttk.Button(frame, text="Send Once", 
              command=lambda: tag_manager.start_data_transmission(
                  tag_entry.get(), data_entry.get(), hex_var.get())).grid(row=1, column=0)
    ttk.Button(frame, text="Start Continuous", 
              command=lambda: tag_manager.start_data_transmission(
                  tag_entry.get(), data_entry.get(), hex_var.get(), 
                  int(interval_entry.get()))).grid(row=1, column=1)
    ttk.Button(frame, text="Stop", 
              command=tag_manager.stop_data_transmission).grid(row=1, column=2)
    
    return frame        


   
target_img = load_image(r"g:\Indoor Positioning\Hexbot\PyScript\Positioning\uwb_positioning_project\assets\target.png", scale=0.05)
anchor_img = load_image(r"g:\Indoor Positioning\Hexbot\PyScript\Positioning\uwb_positioning_project\assets\anchor.png", scale=0.05)
car_img = load_image(r"g:\Indoor Positioning\Hexbot\PyScript\Positioning\uwb_positioning_project\assets\Car_Red.png", scale=0.08)


# === GUI Setup ===
root = tk.Tk()
root.title("Hexapod Control Panel")

tag_manager = TagManager()
anchor_manager = AnchorManager() 
data_logger = DataLogger()

vis_system = VisualizationSystem(root, tag_manager, anchor_manager)

modbus_id_var = tk.StringVar()          # For MODBUS ID display
modbus_id_spinbox = None                # Will be initialized in GUI setup
origin_x_spinbox = None                 # Will be initialized in GUI setup
origin_y_spinbox = None                 # Will be initialized in GUI setup
scale_spinbox = None                    # Will be initialized in GUI setup
tag_count_spinbox = None   

com_port_var = tk.StringVar(value="COM7")  # Default COM port
serial_connection = None  # Holds active serial connection
serial_lock = Lock()

# --- Main Layout ---
notebook = ttk.Notebook(root)
notebook.pack(fill=tk.BOTH, expand=True)

# === Anchor Config Tab ===
anchor_config_tab = ttk.Frame(notebook)
notebook.add(anchor_config_tab, text="Anchor Setup")

# === Anchor Controls ===
control_frame = ttk.LabelFrame(anchor_config_tab, text="Anchor Controls")
control_frame.pack(padx=10, pady=5, fill=tk.X)

# === Frame for anchor setup ===
config_frame = ttk.LabelFrame(anchor_config_tab, text="Anchor Configuration")
config_frame.pack(padx=10, pady=10, fill=tk.X)

# Create two sub-frames for columns
left_frame = ttk.Frame(config_frame)
left_frame.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.BOTH, expand=True)

right_frame = ttk.Frame(config_frame)
right_frame.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.BOTH, expand=True)

separator = ttk.Separator(config_frame, orient=tk.VERTICAL)
separator.pack(side=tk.LEFT, fill=tk.Y, padx=5)

config_vars = {
    # Left column variables
    'Network ID': tk.StringVar(value='1234'),
    'COM Port': tk.StringVar(value='COM5'),
    'MODBUS-ID': tk.StringVar(value='1'),
    'Device Type': tk.StringVar(value='Major Anchor'),
    'Device ID': tk.StringVar(value='0'),
    'Kalman-Q': tk.StringVar(value='3'),
    'Kalman-R': tk.StringVar(value='10'),
    
    # Right column variables
    'Coverage Area': tk.StringVar(value='10'),
    'Antenna Delay': tk.StringVar(value='33040'),
    'UWB Channel': tk.StringVar(value='2'),
    'Data Rate': tk.StringVar(value='6M8'),
    'Positioning Mode': tk.StringVar(value='3D'),
    'Ranging Mode': tk.StringVar(value='HDS-TWR')
}

# === Left Column Layout ===
row = 0
left_fields = ['Network ID', 'COM Port', 'MODBUS-ID', 'Device Type', 'Device ID', 'Kalman-Q', 'Kalman-R']
for label in left_fields:
    ttk.Label(left_frame, text=label).grid(row=row, column=0, padx=5, pady=3, sticky=tk.E)
    if label == "Device Type":
        ttk.Combobox(left_frame, textvariable=config_vars[label], 
                    values=["Major Anchor", "Sub Anchor", "Tag"], width=15).grid(row=row, column=1)
    else:
        ttk.Entry(left_frame, textvariable=config_vars[label], width=20).grid(row=row, column=1)
    row += 1

# === Right Column Layout ===
row = 0
right_fields = ['Coverage Area', 'Antenna Delay', 'UWB Channel', 'Data Rate', 'Positioning Mode', 'Ranging Mode']
for label in right_fields:
    ttk.Label(right_frame, text=label).grid(row=row, column=0, padx=5, pady=3, sticky=tk.E)
    if label == "Data Rate":
        ttk.Combobox(right_frame, textvariable=config_vars[label], 
                    values=["110K", "850K", "6M8"], width=15).grid(row=row, column=1)
    elif label == "Positioning Mode":
        ttk.Combobox(right_frame, textvariable=config_vars[label], 
                    values=["2D", "3D"], width=15).grid(row=row, column=1)
    elif label == "Ranging Mode":
        ttk.Combobox(right_frame, textvariable=config_vars[label], 
                    values=["HDS-TWR", "TWR", "TDOA"], width=15).grid(row=row, column=1)
    else:
        ttk.Entry(right_frame, textvariable=config_vars[label], width=20).grid(row=row, column=1)
    row += 1

ttk.Button(config_frame, text="Save Config", 
          command=lambda: tag_manager.save_config()).pack(side=tk.LEFT, padx=5)


# First row of buttons
button_row1 = ttk.Frame(control_frame)
button_row1.pack(fill=tk.X, pady=5)

ttk.Button(button_row1, text="Load Config", 
          command=lambda: load_config(config_vars, log_area)).pack(side=tk.LEFT, padx=2)
ttk.Button(button_row1, text="Import Config", 
           command=lambda: import_config()).pack(side=tk.LEFT, padx=2)
ttk.Button(button_row1, text="Navigation Control", 
           command=lambda: show_navigation_controls()).pack(side=tk.LEFT, padx=2)

# Options row
options_frame = ttk.Frame(control_frame)
options_frame.pack(fill=tk.X, pady=5)

# Second row of buttons
button_row2 = ttk.Frame(control_frame)
button_row2.pack(fill=tk.X, pady=5)

ttk.Button(button_row2, text="Start Positioning", 
          command=lambda: start_positioning()).pack(side=tk.LEFT, padx=2)
ttk.Button(button_row2, text="Stop Positioning", 
          command=lambda: stop_positioning()).pack(side=tk.LEFT, padx=2)
ttk.Button(button_row2, text="Auto Calibration", 
          command=lambda: auto_calibration()).pack(side=tk.LEFT, padx=2)

# Hardware coordinates option
hardware_coords_var = tk.BooleanVar(value=True)
ttk.Checkbutton(options_frame, text="Coordinates by hardware", 
               variable=hardware_coords_var).pack(side=tk.LEFT, padx=2)

# Help button for hardware coordinates
ttk.Button(options_frame, text="?", 
          command=lambda: show_help("hardware_coords"), 
          width=2).pack(side=tk.LEFT, padx=2)

# Auto-positioning option
auto_position_var = tk.BooleanVar(value=True)
ttk.Checkbutton(options_frame, text="Auto-position after power-on", 
               variable=auto_position_var).pack(side=tk.LEFT, padx=2)

# Help button for auto-positioning
ttk.Button(options_frame, text="?", 
          command=lambda: show_help("auto_position"), 
          width=2).pack(side=tk.LEFT, padx=2)

# === Log output area ===
log_area = tk.Text(anchor_config_tab, height=10, width=70)
log_area.pack(padx=10, pady=5, fill=tk.X)

# === Save Button ===
tag_manager.log_area = log_area
tag_manager.config_vars = config_vars
tag_manager.save_config(config_vars, log_area)

# === Console Frame ===
console_frame = ttk.Frame(notebook)
notebook.add(console_frame, text="Console")

data_export_frame = create_data_export_frame(console_frame)
trajectory_frame = create_trajectory_controls(console_frame)

log_area = scrolledtext.ScrolledText(console_frame, width=80, height=15, wrap=tk.WORD)
log_area.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

waveform_controls = data_logger.setup_waveform_controls(console_frame)
waveform_controls.pack(fill=tk.X, padx=10, pady=5)

entry_frame = ttk.Frame(console_frame)
entry_frame.pack(padx=10, pady=5, fill=tk.X)

entry_field = ttk.Entry(entry_frame, width=40)
entry_field.pack(side=tk.LEFT, fill=tk.X, expand=True)

stats_frame = ttk.LabelFrame(console_frame, text="Real-Time Stats")
stats_frame.pack(padx=10, pady=5, fill=tk.X)

speed_var = tk.StringVar(value="Speed: 0.00")
avg_speed_var = tk.StringVar(value="Avg Speed: 0.00")
distance_var = tk.StringVar(value="Distance: 0.00")
heading_var = tk.StringVar(value="Heading: 0°")

ttk.Label(stats_frame, textvariable=speed_var).pack(side=tk.LEFT, padx=5)
ttk.Label(stats_frame, textvariable=avg_speed_var).pack(side=tk.LEFT, padx=5)
ttk.Label(stats_frame, textvariable=distance_var).pack(side=tk.LEFT, padx=5)
ttk.Label(stats_frame, textvariable=heading_var).pack(side=tk.LEFT, padx=5)

reset_btn = ttk.Button(stats_frame, text="Reset Distance", command=lambda: reset_stats())
reset_btn.pack(side=tk.RIGHT, padx=5)

position_frame = ttk.LabelFrame(console_frame, text="Current Position (XYZ)")
position_frame.pack(padx=10, pady=5, fill=tk.X)

x_var = tk.StringVar(value="X: 0.00")
y_var = tk.StringVar(value="Y: 0.00")
z_var = tk.StringVar(value="Z: 0.00")

ttk.Label(position_frame, textvariable=x_var).pack(side=tk.LEFT, padx=5)
ttk.Label(position_frame, textvariable=y_var).pack(side=tk.LEFT, padx=5)
ttk.Label(position_frame, textvariable=z_var).pack(side=tk.LEFT, padx=5)

def reset_stats():
    global total_distance, total_time, last_time, last_position
    total_distance = 0.0
    total_time = 0.0
    last_time = None
    last_position = None
    distance_var.set("Distance: 0")
    avg_speed_var.set("Avg Speed: 0")
    speed_var.set("Speed: 0")
    heading_var.set("Heading: 0°")
    x_var.set("X: 0")
    y_var.set("Y: 0")
    z_var.set("Z: 0")


send_button = ttk.Button(entry_frame, text="Send Command", command=lambda: send_command(entry_field))
send_button.pack(side=tk.LEFT, padx=5)

# === Add COM Port Frame ===
com_frame = ttk.LabelFrame(console_frame, text="COM Port Configuration")
com_frame.pack(padx=10, pady=5, fill=tk.X)

# COM Port Dropdown
ttk.Label(com_frame, text="COM Port:").pack(side=tk.LEFT)
com_dropdown = ttk.Combobox(console_frame, textvariable=com_port_var, width=10)
com_dropdown.pack(side=tk.LEFT, padx=5)

# Scan Button
scan_button = ttk.Button(console_frame, text="Scan Ports", command=scan_com_ports)
scan_button.pack(side=tk.LEFT, padx=5)

# Connect/Disconnect Button
connect_button = ttk.Button(console_frame, text="Connect", command=toggle_serial_connection)
connect_button.pack(side=tk.LEFT, padx=5)

# Refresh Button (NEW)
refresh_button = ttk.Button(com_frame, text="Refresh", command=refresh_serial)
refresh_button.pack(side=tk.LEFT, padx=5)

clear_buffer_button = ttk.Button(com_frame, text="Clear Buffer", command=clear_serial_buffer)
clear_buffer_button.pack(side=tk.LEFT, padx=5)


# === ADD the "Live plot" tab setup here ===
live_plot_frame = ttk.Frame(notebook)
notebook.add(live_plot_frame, text="Live Plot")

# Create the figure and axes for THIS tab specifically
fig_live, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(8, 6), dpi=100, constrained_layout=True)

canvas_live = FigureCanvasTkAgg(fig_live, master=live_plot_frame)
canvas_live.get_tk_widget().pack(fill=tk.BOTH, expand=True)
toolbar_live = NavigationToolbar2Tk(canvas_live, live_plot_frame)
toolbar_live.update()

def init_live_plot(): # Renamed to be specific
    for ax in [ax1, ax2, ax3]:
        ax.grid(True)
    ax1.set_title("X Coordinate")
    ax2.set_title("Y Coordinate")
    ax3.set_title("Angle (°)")
    return []

def animate_live_plot(i): # Renamed to be specific
    with data_lock:
        if not time_stamps:
            return []
        for ax in [ax1, ax2, ax3]:
            ax.clear() # This is still inefficient, but we'll fix one thing at a time
        ax1.plot(list(time_stamps), list(x_data), 'b-')
        ax2.plot(list(time_stamps), list(y_data), 'g-')
        ax3.plot(list(time_stamps), list(angle_data), 'r-')
        for ax in [ax1, ax2, ax3]:
            ax.tick_params(axis='x', rotation=45, labelsize=8)
            ax.grid(True)
    return ax1.lines + ax2.lines + ax3.lines

# Create the animation object specifically for this figure
ani_live = FuncAnimation(fig_live, animate_live_plot, init_func=init_live_plot, interval=1000, blit=False)


# Configuration frame
config_frame = ttk.LabelFrame(anchor_config_tab, text="Anchor Configuration")
config_frame.pack(padx=10, pady=5, fill=tk.X)

calc_frame = tag_manager.setup_calculation_mode(config_frame)
calc_frame.grid(row=10, column=0, columnspan=2, pady=5)

ranging_combo = ttk.Combobox(config_frame, textvariable=config_vars['Ranging Mode'],
                            values=["TWR", "HDS-TWR", "TDOA"])
ranging_combo.bind("<<ComboboxSelected>>", 
                  lambda e: tag_manager.update_ranging_mode(config_vars['Ranging Mode'].get()))

# Network ID configuration
ttk.Label(config_frame, text="Network ID:").grid(row=row, column=0, sticky=tk.E, padx=5, pady=3)
network_id_var = tk.StringVar()
network_id_entry = ttk.Entry(config_frame, textvariable=network_id_var, width=8)
network_id_entry.grid(row=row, column=1, sticky=tk.W, padx=5, pady=3)
ttk.Button(config_frame, text="Set", 
          command=lambda: set_network_id()).grid(row=row, column=2, padx=5, pady=3)
row += 1  # Increment row counter for next widget

coverage_frame = ttk.Frame(config_frame)
coverage_frame.grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=2)

# Label on the left
ttk.Label(config_frame, text="Coverage Area (m²):").grid(row=row, column=0, sticky=tk.E, padx=5, pady=3)
coverage_var = tk.StringVar(value="10")
ttk.Entry(config_frame, textvariable=coverage_var, width=8).grid(row=row, column=1, sticky=tk.W, padx=5, pady=3)
row += 1

# Entry on the right
ttk.Entry(coverage_frame, textvariable=config_vars['Coverage Area'], width=5).pack(side=tk.LEFT)

# If you need the entry to expand:
# ttk.Entry(coverage_frame, textvariable=config_vars['Coverage Area']).pack(side=tk.LEFT, fill=tk.X, expand=True)

# Update rate configuration
ttk.Label(config_frame, text="Update Rate (ms):").grid(row=row, column=0, sticky=tk.E, padx=5, pady=3)
rate_var = tk.StringVar(value="100")
rate_entry = ttk.Entry(config_frame, textvariable=rate_var, width=8)
rate_entry.grid(row=row, column=1, sticky=tk.W, padx=5, pady=3)
row += 1

# Configuration functions
def set_network_id():
    try:
        network_id = int(network_id_var.get())
        if 1 <= network_id <= 65535:
            if send_to_major_anchor('set_network', {'id': network_id}):
                update_anchor_state('network_id', network_id)
        else:
            safe_log_insert(log_area, "Network ID must be 1-65535\n")
    except ValueError:
        safe_log_insert(log_area, "Invalid network ID format\n")

def set_update_rate():
    try:
        rate = int(rate_var.get())
        if 10 <= rate <= 60000:  # Reasonable bounds
            if send_to_major_anchor('set_rate', {'rate': rate}):
                update_anchor_state('update_rate', rate)
        else:
            safe_log_insert(log_area, "Rate must be 10-60000 ms\n")
    except ValueError:
        safe_log_insert(log_area, "Invalid rate format\n")

# 2D/3D View Frame
frame_2d = ttk.Frame(notebook)
notebook.add(frame_2d, text="2D View")

frame_3d = ttk.Frame(notebook)
notebook.add(frame_3d, text="3D View")

# Map config frame
map_config_frame = create_map_config_frame(frame_2d)
map_config_frame.pack(padx=10, pady=5, fill=tk.X)

map_config_frame = create_map_config_controls(frame_2d)

vis_system.setup_gui_frames(frame_2d, frame_3d)

# === Diagnostics Tab ===
diag_frame = ttk.Frame(notebook)
notebook.add(diag_frame, text="Diagnostics")

# --- Global variables for the diagnostic plot ---
diag_fig, diag_axes = None, None
diag_lines = {}
diag_data = {
    'time': deque(maxlen=100),
    'std_noise': deque(maxlen=100),
    'fp_power': deque(maxlen=100),
    'rx_power': deque(maxlen=100)
}

def create_diagnostic_plots(parent_frame):
    """Creates and configures the diagnostic plots inside the parent frame."""
    global diag_fig, diag_axes, diag_lines, ani_diag

    diag_fig, diag_axes = plt.subplots(3, 1, figsize=(8, 6), sharex=True)
    diag_fig.tight_layout(pad=3.0)

    # --- Setup each subplot ---
    # Standard Noise Plot
    diag_axes[0].set_title("Standard Noise Level")
    diag_axes[0].set_ylabel("ADC Counts")
    diag_axes[0].grid(True)
    diag_lines['std_noise'], = diag_axes[0].plot([], [], 'r-', label='Std Noise')
    diag_axes[0].legend(loc='upper left')

    # First Path Power Plot
    diag_axes[1].set_title("First Path Power (FP Power)")
    diag_axes[1].set_ylabel("dBm")
    diag_axes[1].grid(True)
    diag_lines['fp_power'], = diag_axes[1].plot([], [], 'g-', label='FP Power')
    diag_axes[1].legend(loc='upper left')

    # Receive Power Plot
    diag_axes[2].set_title("Receive Power (RX Power)")
    diag_axes[2].set_ylabel("dBm")
    diag_axes[2].set_xlabel("Time")
    diag_axes[2].grid(True)
    diag_lines['rx_power'], = diag_axes[2].plot([], [], 'b-', label='RX Power')
    diag_axes[2].legend(loc='upper left')
    
    canvas = FigureCanvasTkAgg(diag_fig, master=parent_frame)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    toolbar = NavigationToolbar2Tk(canvas, parent_frame)
    toolbar.update()

    # Start the animation for this specific plot
    ani_diag = FuncAnimation(diag_fig, update_diagnostic_plots, interval=500, blit=False)

def update_diagnostic_plots(i):
    """Animation callback to update diagnostic plot data."""
    # Append new data from the TagManager's latest reading
    diag_data['time'].append(datetime.now())
    diag_data['std_noise'].append(tag_manager.rx_diag.get('Std_noise', 0))
    diag_data['fp_power'].append(tag_manager.rx_diag.get('Fp_power', -150))
    diag_data['rx_power'].append(tag_manager.rx_diag.get('Rx_power', -150))
    
    # Update the lines
    diag_lines['std_noise'].set_data(diag_data['time'], diag_data['std_noise'])
    diag_lines['fp_power'].set_data(diag_data['time'], diag_data['fp_power'])
    diag_lines['rx_power'].set_data(diag_data['time'], diag_data['rx_power'])

    # Rescale axes
    for ax in diag_axes:
        ax.relim()
        ax.autoscale_view()
        
    # Format the x-axis to show time nicely
    diag_axes[2].xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter('%H:%M:%S'))
    diag_fig.autofmt_xdate()

# Call the creation function
create_diagnostic_plots(diag_frame)


def create_channel_selector(self):
    channel_frame = ttk.LabelFrame(self.root, text="UWB Channel")
    channel_frame.pack(padx=10, pady=5, fill=tk.X)
    
    self.channel_var = tk.IntVar(value=2)  # Default to Channel 2
    
    ttk.Label(channel_frame, text="Channel:").pack(side=tk.LEFT)
    channel_dropdown = ttk.Combobox(
        channel_frame,
        textvariable=self.channel_var,
        values=[1, 2, 3, 4, 5, 7],
        width=4
    )
    channel_dropdown.pack(side=tk.LEFT, padx=5)
    
    # Button to send channel command to transmitter
    ttk.Button(
        channel_frame,
        text="Set Channel",
        command=self.send_channel_command
    ).pack(side=tk.LEFT)

channel_var = tk.StringVar(value="Channel: 2")
ttk.Label(stats_frame, textvariable=channel_var).pack(side=tk.LEFT, padx=5)

# Status Label
com_status_var = tk.StringVar(value="Not connected")
ttk.Label(console_frame, textvariable=com_status_var).pack(side=tk.LEFT, padx=5)

# Major Anchor Control Frame 
anchor_frame = ttk.LabelFrame(console_frame, text="Anchor Control")
anchor_frame.pack(padx=10, pady=5, fill=tk.X)

# Status display
anchor_status_var = tk.StringVar(value="Not connected")
ttk.Label(anchor_frame, textvariable=anchor_status_var).pack(side=tk.TOP, fill=tk.X)

# Control buttons
button_frame = ttk.Frame(anchor_frame)
button_frame.pack(side=tk.TOP, fill=tk.X)

# Buttons for common commands
ttk.Button(button_frame, text="Start Positioning", 
          command=lambda: [send_to_major_anchor('start_pos'), 
                         update_anchor_state('positioning', True)]).pack(side=tk.LEFT, padx=2)
ttk.Button(button_frame, text="Stop Positioning", 
          command=lambda: [send_to_major_anchor('stop_pos'), 
                         update_anchor_state('positioning', False)]).pack(side=tk.LEFT, padx=2)
ttk.Button(button_frame, text="Reset Anchor", 
          command=lambda: send_to_major_anchor('reset')).pack(side=tk.LEFT, padx=2)
ttk.Button(button_frame, text="Refresh Info", 
          command=query_anchor_info).pack(side=tk.LEFT, padx=2)

def create_anchor_controls(self):
    self.anchor_frame = ttk.LabelFrame(self.root, text="Anchor Configuration")
    self.anchor_frame.pack(padx=10, pady=5, fill=tk.X)
    
    self.anchor_vars = {}
    for i, anchor_id in enumerate(['A', 'B', 'C', 'D']):
        var = tk.BooleanVar(value=True)
        cb = ttk.Checkbutton(self.anchor_frame, text=anchor_id, variable=var)
        cb.pack(side=tk.LEFT, padx=5)
        self.anchor_vars[anchor_id] = var
        
        # Position entries
        ttk.Label(self.anchor_frame, text=f"{anchor_id} X:").pack(side=tk.LEFT)
        x_entry = ttk.Entry(self.anchor_frame, width=6)
        x_entry.pack(side=tk.LEFT)
        ttk.Label(self.anchor_frame, text=f"{anchor_id} Y:").pack(side=tk.LEFT)
        y_entry = ttk.Entry(self.anchor_frame, width=6)
        y_entry.pack(side=tk.LEFT)
        ttk.Label(self.anchor_frame, text=f"{anchor_id} Z:").pack(side=tk.LEFT)
        z_entry = ttk.Entry(self.anchor_frame, width=6)
        z_entry.pack(side=tk.LEFT)


# ==== Add tooltips ====
# With this (keep the original buttons and add tooltips to them):
add_tooltip(anchor_frame.winfo_children()[0], "Start positioning cycle (sends 'pc')")  # First button
add_tooltip(anchor_frame.winfo_children()[1], "Stop positioning (sends space + newline)")  # Second button

def add_tooltip(widget, text):
    tooltip = tk.Toplevel(widget)
    tooltip.withdraw()
    tooltip.overrideredirect(True)
    label = tk.Label(tooltip, text=text, background="#ffffe0", relief='solid', borderwidth=1, font=("tahoma", 8))
    label.pack()

    def show(event):
        tooltip.geometry(f"+{event.x_root+10}+{event.y_root+10}")
        tooltip.deiconify()

    def hide(event):
        tooltip.withdraw()

    widget.bind("<Enter>", show)
    widget.bind("<Leave>", hide)

# Help dialog functions
def show_command_reference():
    help_window = tk.Toplevel(root)
    help_window.title("Command Reference")
    
    text = scrolledtext.ScrolledText(help_window, width=80, height=20)
    text.pack(padx=10, pady=10)
    
    reference_text = "DWM Anchor Command Reference:\n\n"
    for cmd, info in ANCHOR_COMMANDS.items():
        reference_text += f"{cmd}: {info['desc']}\nCommand: '{info['cmd']}'\n\n"
    
    reference_text += "\nHotkeys:\n"
    reference_text += "Ctrl+S - Start positioning\n"
    reference_text += "Ctrl+X - Stop positioning\n"
    text.insert(tk.END, reference_text)
    text.config(state=tk.DISABLED)

def show_about():
    about_window = tk.Toplevel(root)
    about_window.title("About DWM Receiver")
    msg = ("DWM Positioning Receiver v1.0\n\n"
          "A GUI for visualizing and controlling\n"
          "Decawave DWM1000/PGPlus anchors.\n\n"
          "Supports 2D/3D visualization and\n"
          "real-time position tracking.")
    ttk.Label(about_window, text=msg).pack(padx=20, pady=20)

# Add hotkeys
root.bind('<Control-s>', lambda e: send_to_major_anchor('start_pos'))
root.bind('<Control-x>', lambda e: send_to_major_anchor('stop_pos'))

# Add help menu
menubar = tk.Menu(root)

# --- File Menu (Good Practice) ---
filemenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="File", menu=filemenu)
filemenu.add_command(label="Import Config", command=import_config) # Example
filemenu.add_command(label="Export Channel 1 Data", command=lambda: tag_manager.export_channel_data(1)) # Example
filemenu.add_separator()
filemenu.add_command(label="Exit", command=on_close)

# --- Help Menu (Enhanced) ---
helpmenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="Help", menu=helpmenu)

# Dynamically add help topics from TagManager
for topic, help_text in tag_manager.help_texts.items():
    # Use a lambda to capture the 'help_text' for each iteration
    helpmenu.add_command(
        label=f"{topic} Help", 
        command=lambda text=help_text: tag_manager.show_help_popup(text)
    )

helpmenu.add_separator()
helpmenu.add_command(label="Command Reference", command=show_command_reference)
helpmenu.add_separator()
helpmenu.add_command(label="About", command=show_about)
root.config(menu=menubar)


def launch_background_tasks():
    """Starts all background threads after the GUI is running."""
    # Start the network server thread
    threading.Thread(target=start_server, args=(log_area,), daemon=True).start()
    
    # Scan for COM ports
    scan_com_ports()
    
    vis_system.start_animation()

# Start server thread
threading.Thread(target=start_server, args=(log_area,), daemon=True).start()
root.protocol("WM_DELETE_WINDOW", on_close)

root.after(100, launch_background_tasks)
root.mainloop()

