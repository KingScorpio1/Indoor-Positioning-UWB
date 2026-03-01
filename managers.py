# managers.py

import csv
import time
import math
import queue
import struct
import threading
import pandas as pd
import tkinter as tk
import numpy as np
from tkinter import filedialog, ttk, messagebox
import minimalmodbus
from threading import Lock
from datetime import datetime
from collections import deque
from dataclasses import dataclass, field

# Import default anchor positions from the config file
from config import DEFAULT_ANCHOR_POSITIONS
from collections import deque

# === Globals ===
HOST, PORT = '0.0.0.0', 52740
client_socket = None
shutdown_flag = False
data_lock = Lock()
acc_fsr = 16.0  # ±16g full scale range
gyro_fsr = 2000.0  # ±2000dps full scale range

ser_uwb = None
uwb_lock = Lock()
anchor_instrument = None # MODBUS instrument for anchor communication
anchor_lock = Lock()

ACC_FSR = 16.0      # ±16g full scale range for accelerometer
GYRO_FSR = 2000.0   # ±2000dps full scale range for gyroscope
ANCHOR_MAX_COUNT = 16

# Bitmasks for IMU data types
IMU_DATA_ACC_EN = 0x0001
IMU_DATA_GYRO_EN = 0x0002
IMU_DATA_EULER_EN = 0x0004
IMU_DATA_TEMP_EN = 0x0008
IMU_DATA_Q_EN = 0x0010
IMU_DATA_MAGN_EN = 0x0020

log_queue = queue.Queue()
data_queue = queue.Queue()

map_multiple = 1.0  # Map scaling factor
map_origin_x = 0    # Map origin X coordinate
map_origin_y = 0    # Map origin Y coordinate
max_tags = 100      # Maximum number of tags to track

MODBUS_ADDR_POSITIONING_CONTROL = 59 # Register address for starting/stopping
MODBUS_VAL_START_POSITIONING = 4      # Value to start (matches 0x04 in your bytes)
MODBUS_VAL_STOP_POSITIONING = 0       # Value to stop (matches 0x00 in your bytes)

ANC_PROTOCAL_RTLS = 0      # Base station positioning data flag
ANC_PROTOCAL_DIST = 1      # Base station ranging data flag
ANC_PROTOCAL_RXDIAG = 2    # Base station reception info flag
ANC_PROTOCAL_TIMESTAMP = 3 # Base station timestamp info flag

TAG_OUTPUT_DIST = 0
TAG_OUTPUT_RTLS = 1

@dataclass
class UWBTag:
    def __init__(self, id, max_history=5000):
        self.id = str(id)
        self.x = 0.0
        self.y = 0.0
        self.z = 0.0
        self.angle = 0.0
        self.velocity = 0.0
        self.distance_traveled = 0.0
        self.last_update = 0.0
        self.battery = 0.0
        self.last_seen = time.time()
        self.status = "Inactive"
        self.channel = 2  # Default channel
        self.movement_history = deque(maxlen=200)  # Stores (x, y, z, angle, timestamp)
        self.lock = threading.Lock()

        # 1. History for the AI-Corrected Path (Green)
        self.history_ai = {'x': deque(maxlen=max_history), 'y': deque(maxlen=max_history)}
        
        # 2. History for the Raw Hardware Path (Red)
        self.history_raw = {'x': deque(maxlen=max_history), 'y': deque(maxlen=max_history)}
        
        # IMU data attributes
        self.acc_x, self.acc_y, self.acc_z = 0.0, 0.0, 0.0
        self.gyro_x, self.gyro_y, self.gyro_z = 0.0, 0.0, 0.0
        self.roll, self.pitch, self.yaw = 0.0, 0.0, 0.0 # Euler angles
        self.temperature = 0.0
        self.q0, self.q1, self.q2, self.q3 = 0.0, 0.0, 0.0, 0.0 # Quaternions
        self.magn_x, self.magn_y, self.magn_z = 0.0, 0.0, 0.0
        self.is_get_newdata = False
        
    def update(self, data):
        """
        Updates tag state.
        Expects data to contain: {'x', 'y', 'z', 'raw_x', 'raw_y', ...}
        """
        self.last_seen = time.time()
        self.last_update = time.time()
        
        # Update current 'official' position (usually the AI one)
        self.x = data.get('x', self.x)
        self.y = data.get('y', self.y)
        self.z = data.get('z', self.z)
        self.battery = data.get('battery', self.battery)
        
        # --- MODIFIED SECTION: Append to Histories ---
        self.history_ai['x'].append(self.x)
        self.history_ai['y'].append(self.y)
        
        self.history_raw['x'].append(data.get('raw_x', self.x))
        self.history_raw['y'].append(data.get('raw_y', self.y))
        
        # 2. Save Raw Position
        # If 'raw_x' isn't provided, fallback to the current x
        # raw_x = data.get('raw_x', self.x)
        # raw_y = data.get('raw_y', self.y)
        
        # with self.lock:
        #     # 1. Append to AI History
        #     self.history_ai['x'].append(self.x)
        #     self.history_ai['y'].append(self.y)
            
        #     # 2. Append to Raw History
        #     self.history_raw['x'].append(raw_x)
        #     self.history_raw['y'].append(raw_y)
            
        #     # (Optional) Keep general movement history for velocity calcs
        #     self.movement_history.append((
        #     self.x, 
        #     self.y, 
        #     self.z, 
        #     self.angle, 
        #     self.last_update, # <--- Index 4 used by plotter
        #     self.velocity
        # ))
    
    def update_imu_data(self, imu_data):
        """Update IMU data fields"""
        with self.lock:
            if 'acc_x' in imu_data: self.acc_x = imu_data['acc_x']
            if 'acc_y' in imu_data: self.acc_y = imu_data['acc_y']
            if 'acc_z' in imu_data: self.acc_z = imu_data['acc_z']
            if 'gyro_x' in imu_data: self.gyro_x = imu_data['gyro_x']
            if 'gyro_y' in imu_data: self.gyro_y = imu_data['gyro_y']
            if 'gyro_z' in imu_data: self.gyro_z = imu_data['gyro_z']
            if 'roll' in imu_data: self.roll = imu_data['roll']
            if 'pitch' in imu_data: self.pitch = imu_data['pitch']
            if 'yaw' in imu_data: self.yaw = imu_data['yaw']
            if 'temperature' in imu_data: self.temperature = imu_data['temperature']
            if 'q0' in imu_data: self.q0 = imu_data['q0']
            if 'q1' in imu_data: self.q1 = imu_data['q1']
            if 'q2' in imu_data: self.q2 = imu_data['q2']
            if 'q3' in imu_data: self.q3 = imu_data['q3']
            if 'magn_x' in imu_data: self.magn_x = imu_data['magn_x']
            if 'magn_y' in imu_data: self.magn_y = imu_data['magn_y']
            if 'magn_z' in imu_data: self.magn_z = imu_data['magn_z']
            if 'is_get_newdata' in imu_data: self.is_get_newdata = imu_data['is_get_newdata']
            
    def update_position(self, x: float, y: float, z: float):
        """Thread-safe position update that also calculates velocity."""
        with self.lock:
            now = time.time()
            if self.last_update_time > 0:
                dt = now - self.last_update_time
                if dt > 0.001:
                    dist = ((x - self.x)**2 + (y - self.y)**2 + (z - self.z)**2)**0.5
                    self.velocity = dist / dt
                    self.distance_traveled += dist
                    self.status = "Moving" if self.velocity > 0.1 else "Stationary"
                
            self.x, self.y, self.z = x, y, z
            self.last_update_time = now
            # Store velocity in history as well
            self.movement_history.append((self.x, self.y, self.z, self.angle, now, self.velocity))
    def add_raw_reading(self, distances: dict, rx_diag: dict):
        """Adds a new set of raw readings to the history buffers."""
        with self.lock:
            for aid, dist in distances.items():
                if aid in self.dist_history:
                    self.dist_history[aid].append(dist)
            if rx_diag:
                self.diag_history.append(rx_diag)
                
    def get_position(self):
        """Thread-safe position access"""
        with self.lock:
            return (self.x, self.y, self.z, self.angle)
            
    def calculate_velocity(self, new_x, new_y, new_z):
        """Calculate velocity from new position"""
        with self.lock:
            if len(self.movement_history) > 0:
                prev_x, prev_y, prev_z, _, prev_time = self.movement_history[-1]
                dt = time.time() - prev_time
                
                if dt > 0:
                    dx = new_x - prev_x
                    dy = new_y - prev_y
                    dz = new_z - prev_z
                    distance = (dx**2 + dy**2 + dz**2) ** 0.5
                    
                    self.velocity = distance / dt
                    self.distance_traveled += distance
                    # Update status based on movement
                    self.status = "Moving" if self.velocity > 0.1 else "Stationary"
                  
                    
class TagManager:
    def __init__(self):
        self.tags = {}
        self.active_tag = None
        self.active_tag_id = "0"
                
        self.tag_colors = {
            'default': (255, 255, 255),  # White
            '0': (255, 0, 0),     # Red
            '1': (0, 255, 0),     # Green
            '2': (0, 0, 255),     # Blue
            '3': (255, 255, 0),   # Yellow
            '4': (255, 0, 255),   # Magenta
            '5': (0, 255, 255),   # Cyan
            '6': (255, 128, 0),   # Orange
            '7': (128, 0, 128)    # Purple
        }
        
        # PG-specific data structures
        self.last_cal_data = {
            'Tag_ID': 0,
            'x': 0,
            'y': 0,
            'z': 0,
            'Cal_Flag': 0,
            'Dist': [0] * 16,
            'Time_ts': [0] * 6
        }
        
        self.rx_diag = {
            'Max_noise': 0,
            'Std_noise': 0,
            'Fp_amp1': 0,
            'Fp_amp2': 0,
            'Fp_amp3': 0,
            'Max_growthCIR': 0,
            'Rx_preambleCount': 0,
            'Fp': 0,
            'Fp_power': 0.0,
            'Rx_power': 0.0
        }
        
        self.help_texts = {
            'Ranging Mode': (
                "Ranging Mode Description\n"
                "<Usage> If using HDS-TWR, the rate is fixed at 6.8Mbps and all anchors/tags must use the same ranging mode!\n"
                "<Default Setting> HDS-TWR\n"
                "<Activation> Takes effect immediately after configuration is loaded"
                ),
            'UWB Channel': (
                "UWB Channel\n"
                "<Usage> The wireless channel for DW1000 data transmission\n"
                "Different channels in the same area will cause interference\n"
                "All devices must use the same channel. Channel 2 is most commonly used with the longest range\n"
                "<Default Setting> 2\n"
                "<Activation> Takes effect immediately after configuration is loaded"
                ),
            '3D View': (
                "Dynamic 3D View Configuration Guide\n"
                "<Usage>\n"
                "<X-axis Range> Sets the real-world range represented by the x-axis (blue)\n"
                "<Y-axis Range> Sets the real-world range represented by the y-axis (green)\n"
                "<Z-axis Range> Sets the real-world range represented by the z-axis (red)\n"
                "<X-axis Step> Gridline intervals for x-axis (must be smaller than x-axis range)\n"
                "<Y-axis Step> Gridline intervals for y-axis (must be smaller than y-axis range)\n"
                "<Z-axis Step> Gridline intervals for z-axis (must be smaller than z-axis range)\n"
                "<Trail Length> Maximum length of tag movement trails\n"
                "<Apply Changes> Click 'Update Configuration' to apply settings\n"
                "<View Navigation>\n"
                "Left-click+drag to rotate view | Mouse wheel to zoom\n"
                "W/S/A/D keys move camera forward/back/left/right\n"
                "Z/C keys move camera up/down\n"
                "View adjustments only work during 3D positioning operation"
                ),
            'Hardware Calculation': (
                "Calculation Mode Selection\n"
                "<Usage> Choose between hardware/software calculation\n"
                "<Control>\n"
                "Checked = Hardware calculation | Unchecked = Software calculation\n"
                "Requires writing configuration to take effect\n"
                "In software mode, hardware will report raw ranging data without calculation\n"
                "For 3D positioning with >10 anchors, hardware calculation may fail - use software mode"
                ),
            'Waveform': (
                "Waveform Display\n"
                "<Usage> Waveform 1/2 corresponds to Channel Data Table 1/2\n"
                "<Controls>\n"
                "Edit tag ID to display specific tag data\n"
                "Left-click lines to view detailed values\n"
                "Right-click to pause scrolling and pan view\n"
                "Mouse wheel to zoom in/out\n"
                "After manual interaction, click 'Reset View' to resume auto-scrolling\n"
                "<Default> Channel 1 Tag ID: 0 | Channel 2 Tag ID: 1\n"
                "<Activation> Press Enter after changing ID"
                )
        }
        
        self.config_vars = None
        self.log_area = None  # Will be set later
        self.last_position = None
        self.total_distance = 0.0  # Add these tracking variables
        self.total_time = 0.0
        self.tag_table = None  # Will be set to reference the GUI table
        self.max_tags = max_tags
        self.target_position = None
        self.is_show_trajectory = True
        self.trajectory_button_text_value = "Clear Trajectory Display"
        self.tag_lock = threading.Lock()
        self.data_transmission_enabled = False
        self.transmission_data = bytearray()
        self.transmission_interval = 1000  # ms
        self.transmission_thread = None
        self.transmission_running = False
        self.signal_strength_tag = "0"  # Default to tag 0
        self.signal_filter_alpha = 0.35  # Default filter coefficient
        
    def update_or_create_tag(self, tag_data):
        tag_id = str(tag_data.get('id') or tag_data.get('tag_id'))
        if not tag_id:
            return None

        with self.tag_lock:
            if tag_id not in self.tags:
                self.tags[tag_id] = UWBTag(id=tag_id, max_history=5000) # Increased for analysis

            tag = self.tags[tag_id]
            
            # Record current state for velocity stats
            prev_pos = (tag.x, tag.y, tag.z)
            prev_time = tag.last_update if tag.last_update > 0 else time.time()

            # --- CRITICAL FIX: Call the tag's internal update method ---
            # This triggers history_ai and history_raw appending
            tag.update(tag_data) 
            # -----------------------------------------------------------

            # Calculate velocity and other stats
            self._update_tag_movement_stats(tag, prev_pos, prev_time)
            return tag
        
    def _update_tag_movement_stats(self, tag, prev_pos, prev_time):
        dt = time.time() - prev_time
        if dt > 0.001:
            dx = tag.x - prev_pos[0]
            dy = tag.y - prev_pos[1]
            dz = tag.z - prev_pos[2]
            distance_step = (dx**2 + dy**2 + dz**2)**0.5
            
            tag.velocity = distance_step / dt
            tag.distance_traveled += distance_step
            tag.status = "Moving" if tag.velocity > 0.1 else "Stationary"
        
        if time.time() - tag.last_update > 5:
            tag.status = "Inactive"
            
    def find_tag(self, tag_id):
        """Thread-safe tag lookup"""
        with self.tag_lock:
            return self.tags.get(tag_id, None)
        
    def get_all_tags(self) -> list:
        """Get a thread-safe copy of all tags, updating status."""
        with self.tag_lock:
            now = time.time()
            for tag in self.tags.values():
                if now - tag.last_update > 5.0: # 5 second timeout
                    if tag.status != "Inactive":
                        tag.status = "Inactive"
                        tag.velocity = 0.0
            return list(self.tags.values())
        
    def reset_all_stats(self):
        """Resets distance and velocity for all tags."""
        with self.tag_lock:
            for tag in self.tags.values():
                tag.distance_traveled = 0.0
                tag.velocity = 0.0
                
    def toggle_trajectory(self):
        """Toggle trajectory display on/off"""
        self.is_show_trajectory = not self.is_show_trajectory
        if self.is_show_trajectory:
            self.trajectory_button_text.set("Clear Trajectory Display")
        else:
            self.trajectory_button_text.set("Show Trajectory")
        self.safe_log(f"Trajectory display {'enabled' if self.is_show_trajectory else 'disabled'}\n")
        
    def clear_all_trajectories(self):
        """Clears the movement history for all tags."""
        with self.tag_lock:
            for tag in self.tags.values():
                tag.movement_history.clear()
                
    def get_tag_color(self, tag_id):
        """Get color for a tag, assigning one if it doesn't exist"""
        # Use the tag's ID to get a consistent color
        with self.tag_lock:
            if tag_id not in self.tag_colors:
                # Generate a new color if we run out of predefined ones
                color_idx = len(self.tag_colors) - len(self.tag_colors) % 8  # Cycle through colors
                r = (color_idx * 37) % 256
                g = (color_idx * 73) % 256
                b = (color_idx * 109) % 256
                self.tag_colors[tag_id] = (r, g, b)
            
            return self.tag_colors.get(tag_id, self.tag_colors['default'])

    def get_anchor_color(self, anchor_id):
        """Get color for an anchor"""
        return self.anchor_positions.get(anchor_id, {}).get('color', (255, 255, 255))     
        
    def check_bit_is_true(self, data, b):
        """Check if a specific bit is set"""
        return ((data >> b) & 0x01) == 0x01
    
    def set_log_area(self, log_area):
        """Set the log area after GUI initialization"""
        self.log_area = log_area    
    
    def safe_log(self, message):
        """Thread-safe logging with fallback"""
        if self.log_area:
            try:
                self.log_area.after(0, lambda: self.log_area.insert(tk.END, message))
            except:
                print(f"GUI log failed: {message}")
        else:
            print(message)  # Console fallback
            
    def process_modbus_data(self, data):
        """Process incoming MODBUS data from PG anchors"""
        try:
            if len(data) < 5:
                return False
                
            if data[0] == 0x01 and data[1] == 0x03:  # MODBUS response
                length = data[2]
                if len(data) < length + 5:
                    return False  # Incomplete message
                
                # Verify CRC
                crc = self.crc16(data[:-2])
                if crc != (data[-2] << 8 | data[-1]):
                    self.safe_log("CRC check failed")
                    return False
                
                # Check for calibration data (0xDA 0xDA)
                if data[3] == 0xDA and data[4] == 0xDA:
                    return self.handle_calibration_data(data)
                
                if data[3] == 0xCA and data[4] == 0xDA:
                    self.process_anchor_data(data[5:-2])
                elif data[3] == 0xAC and data[4] == 0xDA:
                    self.process_tag_data(data[5:-2])
                
                return True
            
        except Exception as e:
            self.safe_log(f"Error processing MODBUS data: {str(e)}")
            return False
    
    def process_anchor_data(self, data):
        """Process data from anchor station"""
        ptr = 0
        output_protocal = (data[ptr] << 8) | data[ptr+1]
        ptr += 2
        
        self.last_cal_data['Tag_ID'] = (data[ptr] << 8) | data[ptr+1]
        ptr += 2
        
        # Status flags
        cal_flag = (data[ptr] << 24) | (data[ptr+1] << 16) | (data[ptr+2] << 8) | data[ptr+3]
        ptr += 4
        self.last_cal_data['Cal_Flag'] = cal_flag
        
        # Positioning data
        if self.check_bit_is_true(output_protocal, ANC_PROTOCAL_RTLS):
            if not self.check_bit_is_true(cal_flag, 16):  # Positioning failed
                ptr += 6
                self.safe_log("Positioning calculation failed")
            else:
                self.last_cal_data['x'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.last_cal_data['y'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.last_cal_data['z'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.safe_log(f"Positioning success - X: {self.last_cal_data['x']} Y: {self.last_cal_data['y']} Z: {self.last_cal_data['z']}")
        
        # Distance data
        if self.check_bit_is_true(output_protocal, ANC_PROTOCAL_DIST):
            for i in range(16):
                if self.check_bit_is_true(cal_flag, i):
                    self.last_cal_data['Dist'][i] = (data[ptr] << 8) | data[ptr+1]
                    ptr += 2
                else:
                    self.last_cal_data['Dist'][i] = 0
                    ptr += 2
        
        # Reception diagnostics
        if self.check_bit_is_true(output_protocal, ANC_PROTOCAL_RXDIAG):
            if self.last_cal_data['Tag_ID'] == self.active_tag:
                self.rx_diag['Max_noise'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Std_noise'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Fp_amp1'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Fp_amp2'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Fp_amp3'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Max_growthCIR'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Rx_preambleCount'] = (data[ptr] << 8) | data[ptr+1]
                ptr += 2
                self.rx_diag['Fp'] = ((data[ptr] << 8) | data[ptr+1]) / 64
                ptr += 2
                
                self.rx_diag['Fp_power'] = self.calculate_fp_power(self.rx_diag)
                self.rx_diag['Rx_power'] = self.calculate_rx_power(self.rx_diag)
            else:
                ptr += 16
        
        # Timestamps
        if self.check_bit_is_true(output_protocal, ANC_PROTOCAL_TIMESTAMP):
            if self.last_cal_data['Tag_ID'] == self.active_tag:
                for j in range(6):
                    self.last_cal_data['Time_ts'][j] = (data[ptr] << 24) | (data[ptr+1] << 16) | (data[ptr+2] << 8) | data[ptr+3]
                    ptr += 4
    
    def process_tag_data(self, data):
        """Process data from tag"""
        ptr = 0
        output_protocal = (data[ptr] << 8) | data[ptr+1]
        ptr += 2
        
        # Distance data
        if self.check_bit_is_true(output_protocal, TAG_OUTPUT_DIST):
            dist_flag = (data[ptr] << 8) | data[ptr+1]
            ptr += 2
            
            for i in range(16):
                if self.check_bit_is_true(dist_flag, i):
                    self.last_cal_data['Dist'][i] = (data[ptr] << 8) | data[ptr+1]
                    ptr += 2
                else:
                    ptr += 2
        
        # Positioning data
        if self.check_bit_is_true(output_protocal, TAG_OUTPUT_RTLS):
            if (data[ptr] << 8) | data[ptr+1] == 1:  # Positioning enabled
                ptr += 2
                self.last_cal_data['x'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.last_cal_data['y'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
                self.last_cal_data['z'] = self.decode_int16(data[ptr:ptr+2])
                ptr += 2
            else:
                ptr += 2
    
    def process_rtls_packet(self, data: bytes):
        """
        NEW: This is the primary function to parse the RTLS packet.
        It's based on the C# Rtls_DataRecv function.
        It expects the PAYLOAD part of the Modbus frame (after the 0xCA 0xDA).
        """
        try:
            # --- Unpack header fields ---
            # '>' for big-endian, 'H' for unsigned short, 'I' for unsigned int
            unpacker_header = struct.Struct('>HHI')
            output_protocol, tag_id, cal_flag = unpacker_header.unpack_from(data, 0)
            cursor = unpacker_header.size

            tag = self.get_or_create_tag(str(tag_id))

            # --- Extract Position (if available) ---
            if (output_protocol & (1 << ANC_PROTOCAL_RTLS)):
                position_valid = bool((cal_flag >> 16) & 0x01)
                if position_valid:
                    # '>' for big-endian, 'h' for signed short
                    pos_unpacker = struct.Struct('>hhh')
                    raw_x, raw_y, raw_z = pos_unpacker.unpack_from(data, cursor)
                    # The module sends coordinates in cm, convert to meters for consistency
                    tag.update_position(raw_x / 100.0, raw_y / 100.0, raw_z / 100.0)
                cursor += 6  # Always advance cursor by 6 bytes for position data block

            # --- Extract Distances (if available) ---
            if (output_protocol & (1 << ANC_PROTOCAL_DIST)):
                dist_unpacker = struct.Struct('>H') # unsigned short
                for i in range(ANCHOR_MAX_COUNT):
                    # The distance field is always present, but we check the flag to see if it's valid
                    distance_cm = dist_unpacker.unpack_from(data, cursor)[0]
                    is_valid = bool((cal_flag >> i) & 0x01)
                    if is_valid:
                        # You can store this distance data in your anchor/tag objects if needed
                        # For now, we just log it for debugging
                        if i == 0: # Log distance to anchor A for a quick check
                            # print(f"Tag {tag_id} Dist to A: {distance_cm} cm")
                            pass
                    cursor += 2

            # After parsing, we can update stats and GUI
            # root.after(0, self.update_gui_for_tag, tag)
            return True

        except (struct.error, IndexError) as e:
            self.safe_log_insert(self.log_area, f"ERROR: Failed to parse RTLS packet: {e}\n")
            return False
        
    def decode_int16(self, data):
        """Decode a 16-bit signed integer from bytes"""
        return struct.unpack('>h', bytes(data))[0]
    
    def crc16(self, data):
        """Calculate MODBUS CRC-16 checksum"""
        crc = 0xFFFF
        for byte in data:
            crc ^= byte
            for _ in range(8):
                if crc & 0x0001:
                    crc >>= 1
                    crc ^= 0xA001
                else:
                    crc >>= 1
        return crc

    def filter_multipath(self, x, y, z):
        """Filter positions based on installation guidelines"""
        # Height check (minimum 1m per docs)
        if z < 1.0:
            return False
            
        # Wall proximity check (30cm clearance)
        if x < 0.3 or y < 0.3:
            return False
            
        # Velocity sanity check (5m/s threshold)
        if self.last_position:
            dx = x - self.last_position['x']
            dy = y - self.last_position['y']
            dt = time.time() - self.last_position['time']
            if dt > 0 and (dx**2 + dy**2)**0.5/dt > 5.0:
                return False
                
        self.last_position = {'x': x, 'y': y, 'time': time.time()}
        return True
    
    def export_to_excel(self,  data_table, filename):
        """Export data to Excel file similar to C# implementation"""
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Add headers
            headers = [
                "Time", "X", "Y", "Z", 
                "A", "B", "C", "D", "E", "F", "G", "H",
                "I", "J", "K", "L", "M", "N", "O", "P",
                "Flag", "Velocity"
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            # Add data
            for row_idx, row_data in enumerate(data_table, 2):
                for col_idx, value in enumerate(row_data, 1):
                    if col_idx == 1:  # Time column
                        ws.cell(row=row_idx, column=col_idx, value=str(value))
                    else:
                        try:
                            num_value = float(value)
                            ws.cell(row=row_idx, column=col_idx, value=num_value)
                        except ValueError:
                            ws.cell(row=row_idx, column=col_idx, value=value)
                            
            wb.save(filename)
            return True
        except Exception as e:
            print(f"Export failed: {e}")
            return False
        
    def safe_log_insert(self, log_area, text):
        if not shutdown_flag:
            log_area.after(0, lambda: log_area.insert(tk.END, text))
            
    def send_to_major_anchor(self, command, params=None, retries=3, timeout=0.5):
        """Improved anchor command sending with retries"""
        if not anchor_instrument or not anchor_instrument.serial.is_open:
            self.safe_log_insert(self.log_area, "Serial connection not established\n")
            return False
            
        try:
            for attempt in range(retries):
                try:
                    with anchor_lock:
                        if command == 'start_pos':
                            anchor_instrument.write_register(
                                MODBUS_ADDR_POSITIONING_CONTROL,
                                MODBUS_VAL_START_POSITIONING
                            )
                        elif command == 'stop_pos':
                            anchor_instrument.write_register(
                                MODBUS_ADDR_POSITIONING_CONTROL,
                                MODBUS_VAL_STOP_POSITIONING
                            )
                        elif command == 'set_network':
                            anchor_instrument.write_register(
                                0x20,  # Network ID register
                                params['id']
                            )
                        elif command == 'set_rate':
                            anchor_instrument.write_register(
                                0x82,  # Update rate register
                                params['rate']
                            )
                        # Add other commands as needed
                        
                        return True
                        
                except minimalmodbus.ModbusException as e:
                    if attempt == retries - 1:
                        self.safe_log_insert(self.log_area, f"Failed: {str(e)}\n")
                        raise
                    time.sleep(0.1)
                    
        except Exception as e:
            self.safe_log_insert(self.log_area, f"Command '{command}' failed: {str(e)}\n")
            return False
        
    def get_tag_info(tag_id):
        tag_manager = TagManager()
        """Get current info for a specific tag"""
        with data_lock:
            if tag_id in tag_manager.tags:
                tag = tag_manager.tags[tag_id]
                return {
                    'x': tag.x,
                    'y': tag.y,
                    'z': tag.z,
                    'angle': tag.angle,
                    'status': 'Active' if time.time() - tag.last_update < 5 else 'Inactive'
                }
        return None

    def save_config(self, config_vars=None, log_area=None):
        """Unified configuration save handler"""
        config_vars = config_vars or self.config_vars
        if not config_vars:
            raise ValueError("No configuration variables set")
        log_area = log_area or self.log_area
        
        try:
            device_type = config_vars['Device Type'].get().lower()
            type_map = {
                'major anchor': '1',
                'sub anchor': '2', 
                'tag': '3'
            }
            device_type_value = type_map.get(device_type, '1')

            commands = [
                f"nvm 0x80 {config_vars['MODBUS-ID'].get()}",
                f"kalman_q {config_vars['Kalman-Q'].get()}",
                f"kalman_r {config_vars['Kalman-R'].get()}",
                f"antdly {config_vars['Antenna Delay'].get()}",
                f"mode {config_vars['Positioning Mode'].get().lower()}",
                f"chan {config_vars['UWB Channel'].get()}",
                f"rate {config_vars['Data Rate'].get()}",
                f"nvm 0x84 {device_type_value}",
                f"nvm 0x85 {config_vars['Device ID'].get()}",
                f"ranging {config_vars['Ranging Mode'].get().lower()}",
                f"nvm 0x86 {config_vars['Network ID'].get()}"
            ]

            for cmd in commands:
                if not self.send_to_major_anchor(cmd):
                    raise Exception(f"Command failed: {cmd}")
                    
            self.safe_log_insert(log_area, "Configuration saved successfully\n")
            return True
            
        except Exception as e:
            self.safe_log_insert(log_area, f"Configuration save failed: {str(e)}\n")
            return False

    def start_positioning(self):
        """Start HDS-TWR positioning cycle"""
        try:
            # Set network ID first
            if not self.send_to_major_anchor('set_network', 
                                      {'id': self.config_vars['Network ID'].get()}):
                raise Exception("Network ID config failed")

            # Set data rate based on coverage area
            coverage = float(self.config_vars['Coverage Area'].get())
            rate = 110 if coverage > 15 else 6800  # 110kbps or 6.8Mbps
            self.send_to_major_anchor('set_rate', {'rate': rate})

            # Execute anchor sequence
            seq = [
                ('start_pos', None),    # Anchor A
                ('trigger_B', 0.003),   # 3ms delay
                ('trigger_C', 0.003),   # 3ms delay
                ('trigger_D', 0.003)    # 3ms delay
            ]
            
            for cmd, delay in seq:
                if not self.send_to_major_anchor(cmd):
                    raise Exception(f"{cmd} failed")
                if delay:
                    time.sleep(delay)
                    
            self.safe_log_insert(self.log_area, "HDS-TWR positioning started\n")
            return True
            
        except Exception as e:
            self.safe_log_insert(self.log_area, f"Start failed: {str(e)}\n")
            return False

    def get_active_tag_info(self):
        """Get info for the currently active tag"""
        if not self.active_tag:
            return None
            
        tag = self.find_tag(self.active_tag)
        if not tag:
            return None
            
        return {
            'id': tag.id,
            'x': tag.x,
            'y': tag.y,
            'z': tag.z,
            'angle': tag.angle,
            'velocity': tag.velocity,
            'distance': tag.distance_traveled,
            'status': tag.status,
            'channel': tag.channel,
            'last_update': tag.last_update
        }

    def reset_tag_stats(self, tag_id):
        """Reset distance and velocity for a tag"""
        with self.tag_lock:
            if tag_id in self.tags:
                tag = self.tags[tag_id]
                tag.distance_traveled = 0.0
                tag.velocity = 0.0
                tag.movement_history.clear()
                
    def validate_position(self, x, y, z):
        """Enhanced position validation from C# code"""
        # Ground/obstruction check
        if z < 1.0:  # Minimum height per docs
            self.safe_log(f"WARNING: Invalid height {z}m - must be ≥1m")
            return False
        
        # Origin rejection (common error state)
        if abs(x) < 0.1 and abs(y) < 0.1:
            self.safe_log("WARNING: Position at origin - likely error")
            return False
        
        # Velocity sanity check
        if self.last_position and self.last_time:
            dt = time.time() - self.last_time
            if dt > 0:
                dx = x - self.last_position['x']
                dy = y - self.last_position['y']
                velocity = (dx**2 + dy**2)**0.5 / dt
                if velocity > 5.0:  # 5m/s = 18km/h
                    self.safe_log(f"WARNING: Implausible velocity: {velocity:.2f}m/s")
                    return False
        
        self.last_position = {'x': x, 'y': y, 'time': time.time()}
        return True
    
    def update_anchor_positions(self, anchor_positions):
        """Validate anchor positions against installation guidelines"""
        for anchor_id, pos in anchor_positions.items():
            if pos['z'] < 1.0:
                self.safe_log(f"WARNING: Anchor {anchor_id} height {pos['z']}m < 1m minimum")
            if min(pos['x'], pos['y']) < 0.5:
                self.safe_log(f"WARNING: Anchor {anchor_id} too close to walls")
                
    def calculate_fp_power(self, rx_diag_data):
        """Calculate First Path Power in dBm. Based on C# code."""
        try:
            if rx_diag_data['Rx_preambleCount'] == 0: return -150.0 # Avoid division by zero
            result = (rx_diag_data['Fp_amp1']**2 + rx_diag_data['Fp_amp2']**2 + rx_diag_data['Fp_amp3']**2)
            n_square = rx_diag_data['Rx_preambleCount']**2
            result /= n_square
            # Use math.log10, checking for result > 0
            result = 10 * math.log10(result) if result > 0 else -150.0
            result -= 121.74 # Constant from DW1000/3000 documentation
            return round(result, 2)
        except Exception:
            return -150.0

    def calculate_rx_power(self, rx_diag_data):
        """Calculate Receive Power in dBm. Based on C# code."""
        try:
            if rx_diag_data['Rx_preambleCount'] == 0: return -150.0 # Avoid division by zero
            pow_value = 2**17  # DW1000 constant (use 2**21 for DW3000 if you add that logic)
            n_square = rx_diag_data['Rx_preambleCount']**2
            result = rx_diag_data['Max_growthCIR'] * pow_value / n_square
            result = 10 * math.log10(result) if result > 0 else -150.0
            result -= 121.74
            return round(result, 2)
        except Exception:
            return -150.0
    
    def process_imu_data(self, data):
        """Process IMU data from device"""
        ptr = 0
        print_en = (data[ptr] << 8) | data[ptr+1]
        ptr += 2
        
        imu_data = {}
        
        if print_en & IMU_DATA_ACC_EN:
            imu_data['acc_x'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * acc_fsr
            ptr += 2
            imu_data['acc_y'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * acc_fsr
            ptr += 2
            imu_data['acc_z'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * acc_fsr
            ptr += 2
            
        if print_en & IMU_DATA_GYRO_EN:
            imu_data['gyro_x'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * gyro_fsr
            ptr += 2
            imu_data['gyro_y'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * gyro_fsr
            ptr += 2
            imu_data['gyro_z'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * gyro_fsr
            ptr += 2
            
        if print_en & IMU_DATA_EULER_EN:
            imu_data['roll'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * 180.0
            ptr += 2
            imu_data['pitch'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * 180.0
            ptr += 2
            imu_data['yaw'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0 * 180.0
            ptr += 2
            
        if print_en & IMU_DATA_TEMP_EN:
            imu_data['temperature'] = self.decode_int16(data[ptr:ptr+2]) / 340.0 + 36.53
            ptr += 2
            
        if print_en & IMU_DATA_Q_EN:
            imu_data['q0'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0
            ptr += 2
            imu_data['q1'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0
            ptr += 2
            imu_data['q2'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0
            ptr += 2
            imu_data['q3'] = self.decode_int16(data[ptr:ptr+2]) / 32768.0
            ptr += 2
    
        if print_en & IMU_DATA_MAGN_EN:
            imu_data['magn_x'] = self.decode_int16(data[ptr:ptr+2]) * 0.15
            ptr += 2
            imu_data['magn_y'] = self.decode_int16(data[ptr:ptr+2]) * 0.15
            ptr += 2
            imu_data['magn_z'] = self.decode_int16(data[ptr:ptr+2]) * 0.15
            ptr += 2
            
        return imu_data 
    
    def update_tag_table(self):
        """Update the GUI table with current tag data"""
        if not self.tag_table:
            return
            
        # Clear existing rows
        for row in self.tag_table.get_children():
            self.tag_table.delete(row)
            
        # Add current tag data
        for tag_id, tag in self.tags.items():
            x, y, z, angle = tag.get_position()
            self.tag_table.insert("", "end", 
                                 values=(tag_id, f"{x:.2f}", f"{y:.2f}", f"{z:.2f}", 
                                        f"{angle:.1f}", tag.status, tag.channel))
    
    def set_target(self, x, y):
        """Set target coordinates"""
        self.target_position = (x, y)
        
    def clear_target(self):
        """Clear target"""
        self.target_position = None
        
    def export_channel_data(self, channel, filename=None):
        """Export channel data to Excel file"""
        if not filename:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title=f"Export Channel {channel} Data"
            )
            if not filename:
                return False
                
        table_name = 'trace1' if channel == 1 else 'trace2'
        
        try:
            # Use the enhanced Excel export function
            return self.export_to_excel(self.tables[table_name], filename)
        except Exception as e:
            self.safe_log(f"Export failed: {str(e)}\n")
            return False

    def update_device_type_ui(self, device_type):
        """Update UI based on device type (Master, Sub, Tag)"""
        if device_type == "Tag":
            self.config_widgets['Positioning Mode'].config(state='disabled')
            self.config_widgets['Device ID'].config(state='normal')
            self.config_widgets['Ranging Mode'].config(state='normal')
            self.config_widgets['UWB Channel'].config(state='normal')
            self.config_widgets['Data Rate'].config(state='normal')
            # Disable Kalman filter settings for tags
            self.config_widgets['Kalman-Q'].config(state='disabled')
            self.config_widgets['Kalman-R'].config(state='disabled')
            
        elif device_type == "Sub Anchor":
            self.config_widgets['Positioning Mode'].config(state='disabled')
            self.config_widgets['Device ID'].config(state='normal')
            # Similar enable/disable logic for other controls
            
        elif device_type == "Master Anchor":
            self.config_widgets['Positioning Mode'].config(state='normal')
            self.config_widgets['Device ID'].config(state='disabled')
            # Enable all configuration options
            self.config_widgets['Kalman-Q'].config(state='normal')
            self.config_widgets['Kalman-R'].config(state='normal')

    def validate_config(self):
        """Validate configuration based on device type and ranging mode"""
        errors = []
        
        # Device ID validation
        device_type = self.config_vars['Device Type'].get()
        device_id = int(self.config_vars['Device ID'].get())
        
        if device_type == "Sub Anchor" and device_id > 14:
            errors.append("Sub Anchor ID must be 0-14 (B=0,C=1,...,P=14)")
        elif device_type == "Tag" and device_id > 99:
            errors.append("Tag ID must be 0-99")
        
        # Network ID validation
        network_id = int(self.config_vars['Network ID'].get())
        if network_id < 1 or network_id > 65535:
            errors.append("Network ID must be 1-65535")
        
        return errors

    def save_config_to_device(self):
        """Save current configuration to device"""
        try:
            # Convert configuration to MODBUS registers
            config_bytes = bytearray(122)  # Match C# buffer size
            
            # MODBUS ID (register 0)
            config_bytes[0:2] = int(self.config_vars['MODBUS-ID'].get()).to_bytes(2, 'big')
            
            # Device type (register 6)
            device_type_map = {"Master Anchor": 2, "Sub Anchor": 1, "Tag": 0}
            config_bytes[6] = device_type_map.get(self.config_vars['Device Type'].get(), 0)
            
            # Device ID (registers 8-9)
            config_bytes[8:10] = int(self.config_vars['Device ID'].get()).to_bytes(2, 'big')
            
            # Kalman filters (registers 12-15)
            config_bytes[12:14] = int(self.config_vars['Kalman-Q'].get()).to_bytes(2, 'big')
            config_bytes[14:16] = int(self.config_vars['Kalman-R'].get()).to_bytes(2, 'big')
            
            # Antenna delay (registers 16-17)
            config_bytes[16:18] = int(self.config_vars['Antenna Delay'].get()).to_bytes(2, 'big')
            
            # Send via MODBUS
            if not self.send_modbus_command(0, config_bytes):
                raise Exception("Failed to write configuration")
                
            self.safe_log("Configuration saved to device\n")
            return True
            
        except Exception as e:
            self.safe_log(f"Failed to save config: {str(e)}\n")
            return False

    def handle_ranging_mode(self, mode):
        """Apply ranging mode specific settings"""
        if mode == "HDS-TWR":
            # Set fixed parameters for HDS-TWR
            self.config_vars['Data Rate'].set('6M8' if self.module_chip == 'DW1000' else '850K')
            self.config_widgets['Data Rate'].config(state='disabled')
            
            # Additional HDS-TWR specific initialization
            if not self.send_to_major_anchor('set_hds_mode'):
                raise Exception("Failed to set HDS-TWR mode")
        else:
            self.config_widgets['Data Rate'].config(state='normal')
            
    def create_help_buttons(self, parent_frame):
        """Create help buttons in the specified frame"""
        for key, text in self.help_texts.items():
            btn = ttk.Button(parent_frame, text=f"{key} Help",
                           command=lambda t=text: self.show_help_popup(t))
            btn.pack(side=tk.LEFT, padx=2)
    
    def show_help_popup(self, text):
        """Display help text in a popup window"""
        popup = tk.Toplevel()
        popup.title("Help")
        text_widget = tk.Text(popup, wrap=tk.WORD, width=60, height=20)
        text_widget.insert(tk.END, text)
        text_widget.config(state=tk.DISABLED)
        text_widget.pack(padx=10, pady=10)
        ttk.Button(popup, text="Close", command=popup.destroy).pack(pady=5)
    
    def setup_calculation_mode(self, parent):
        """Create calculation mode controls"""
        frame = ttk.LabelFrame(parent, text="Calculation Mode")
        
        self.calc_mode = tk.BooleanVar(value=True)  # Default to hardware
        
        # Use grid instead of pack
        ttk.Checkbutton(frame, text="Hardware Calculation", 
                        variable=self.calc_mode,
                        command=self.update_calc_mode).grid(row=0, column=0, padx=2, pady=2)
    
        ttk.Button(frame, text="Help", 
                   command=self.show_calc_help).grid(row=0, column=1, padx=2, pady=2)
        
        return frame
    
    def update_calc_mode(self):
        """Send calculation mode to device"""
        mode = "hardware" if self.calc_mode.get() else "software"
        if not self.send_to_major_anchor(f"calc_mode {mode}"):
            self.safe_log("Failed to update calculation mode\n")
            # Revert on failure
            self.calc_mode.set(not self.calc_mode.get())
         
    def update_ranging_mode(self, mode):
        """Handle ranging mode changes"""
        if mode == "HDS-TWR":
            # Set fixed parameters
            rate = "6M8" if self.module_chip == "DW1000" else "850K"
            self.config_vars['Data Rate'].set(rate)
            self.config_widgets['Data Rate'].config(state='disabled')
            
            # Update channel options
            if self.module_chip == "DW3000":
                self.config_widgets['UWB Channel']['values'] = ["1", "2", "3", "4", "5"]
            
            self.safe_log("All devices must be set to HDS-TWR mode!\n")
        else:
            self.config_widgets['Data Rate'].config(state='normal')
            self.config_widgets['UWB Channel']['values'] = ["1", "2", "3", "4", "5", "7"]
        
        # Send command to device
        if not self.send_to_major_anchor(f"ranging {mode.lower()}"):
            self.safe_log("Failed to update ranging mode\n")
    
    def show_calc_help(self):
        """Show help information about calculation modes"""
        help_text = (
            "Calculation Mode Help\n\n"
            "Hardware Calculation: Performed by the anchor's internal processor\n"
            "Software Calculation: Performed by this application\n\n"
            "For large deployments (>10 anchors), software mode is recommended."
        )
        
        # Create a simple help popup
        help_window = tk.Toplevel()
        help_window.title("Calculation Mode Help")
        help_window.geometry("400x200")
        
        text_widget = tk.Text(help_window, wrap=tk.WORD, padx=10, pady=10)
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
        text_widget.pack(fill=tk.BOTH, expand=True)
        
        ttk.Button(help_window, text="Close", command=help_window.destroy).pack(pady=5)
    
    def start_data_transmission(self, tag_id, data, hex_mode=False, interval=0):
        """Start transmitting data to a tag"""
        try:
            if hex_mode:
                self.transmission_data = bytes.fromhex(data.replace(" ", ""))
            else:
                self.transmission_data = data.encode('ascii')[:10]  # Max 10 bytes
            
            if interval > 0:
                self.transmission_interval = interval
                self.transmission_running = True
                self.transmission_thread = threading.Thread(
                    target=self._transmission_loop, args=(tag_id,), daemon=True)
                self.transmission_thread.start()
            else:
                self._send_transmission_data(tag_id)
                
            return True
        except Exception as e:
            self.safe_log(f"Transmission error: {str(e)}\n")
            return False
            
    def _transmission_loop(self, tag_id):
        """Continuous transmission loop"""
        while self.transmission_running:
            self._send_transmission_data(tag_id)
            time.sleep(self.transmission_interval / 1000)
            
    def _send_transmission_data(self, tag_id):
        """Send data to specific tag"""
        if client_socket:
            try:
                cmd = f"AT+DataSend=\"{self.transmission_data.hex()}\",\"{tag_id}\"\n"
                client_socket.sendall(cmd.encode())
            except Exception as e:
                self.safe_log(f"Transmission send error: {str(e)}\n")
                
    def stop_data_transmission(self):
        """Stop continuous transmission"""
        self.transmission_running = False
        if self.transmission_thread:
            self.transmission_thread.join()
    
    def set_anchor_output_protocol(self, positioning=True, distances=True, 
                                 diagnostics=True, timestamps=True):
        """Configure what data the anchor outputs"""
        if not anchor_instrument:
            self.safe_log("No anchor connection\n")
            return False
            
        try:
            # Convert boolean flags to bitmask
            protocol = 0
            if positioning: protocol |= (1 << ANC_PROTOCAL_RTLS)
            if distances: protocol |= (1 << ANC_PROTOCAL_DIST) 
            if diagnostics: protocol |= (1 << ANC_PROTOCAL_RXDIAG)
            if timestamps: protocol |= (1 << ANC_PROTOCAL_TIMESTAMP)
            
            # Send via MODBUS
            with anchor_lock:
                anchor_instrument.write_register(0x30, protocol)
            return True
        except Exception as e:
            self.safe_log(f"Protocol config failed: {str(e)}\n")
            return False
    
    def set_signal_strength_tag(self, tag_id):
        """Set which tag to monitor for signal strength"""
        if tag_id in self.tags:
            self.signal_strength_tag = tag_id
            return True
        return False
        
    def apply_signal_filter(self, new_value, last_value):
        """Apply low-pass filter to signal strength"""
        return last_value * (1 - self.signal_filter_alpha) + new_value * self.signal_filter_alpha
        
    def update_signal_strength_display(self):
        """Update GUI with filtered signal strength"""
        if self.signal_strength_tag in self.tags:
            tag = self.tags[self.signal_strength_tag]
            # Get raw values from rx_diag
            fp_power = self.calculate_fp_power(self.rx_diag)
            rx_power = self.calculate_rx_power(self.rx_diag)
            
            # Apply filter
            tag.filtered_fp_power = self.apply_signal_filter(
                fp_power, getattr(tag, 'filtered_fp_power', fp_power))
            tag.filtered_rx_power = self.apply_signal_filter(
                rx_power, getattr(tag, 'filtered_rx_power', rx_power))
                
            # Update GUI
            if hasattr(self, 'signal_strength_var'):
                self.signal_strength_var.set(
                    f"Tag {tag.id}: FP={tag.filtered_fp_power:.2f}dBm, "
                    f"RX={tag.filtered_rx_power:.2f}dBm")
            
    def handle_calibration_data(self, data):
        """Process calibration data received from anchor"""
        if len(data) < 5 or data[3] != 0xDA or data[4] != 0xDA:
            return False
            
        try:
            # Parse calibration data (similar to C# code)
            anchor_id = chr(65 + data[5])  # Convert index to A-F
            x = struct.unpack('>h', bytes(data[6:8]))[0]  # Big-endian int16
            y = struct.unpack('>h', bytes(data[8:10]))[0]
            z = struct.unpack('>h', bytes(data[10:12]))[0]
            
            # Update anchor position
            if self.anchor_manager: # Check if the manager is set
                self.anchor_manager.update_anchor_position(anchor_id, x / 100.0, y / 100.0, z / 100.0)
                self.safe_log(f"Calibration update: {anchor_id} X={x}cm, Y={y}cm, Z={z}cm\n")
                return True
                
        except Exception as e:
            self.safe_log(f"Error processing calibration data: {str(e)}\n")
            return False
        
    def get_active_tag_data(self):
        """Returns a copy of the active tag's data for thread-safe UI updates."""
        with self.tag_lock:
            tag = self.tags.get(self.active_tag_id)
            if tag:
                return {
                    "id": tag.id,
                    "x": tag.x,
                    "y": tag.y,
                    "z": tag.z,
                    "velocity": tag.velocity,
                    "distance_traveled": tag.distance_traveled,
                    "status": tag.status,
                    "channel": tag.channel,
                    "last_update": tag.last_update
                }
            return None
        
    
class AnchorManager:
    def __init__(self):
        self.anchors = {}  # Dictionary of anchors keyed by ID
        self.active_anchors = set()
        self.anchor_lock = threading.Lock()
        self.max_anchors = 16  # PGPlus supports up to 16 anchors
        
        # Initialize default anchor positions
        self.anchor_positions = {
            'A': {'x': 0, 'y': 0, 'z': 1.2, 'enabled': True, 'color': (0, 255, 0)},
            'B': {'x': .8, 'y': 6.76, 'z': 1.2, 'enabled': True, 'color': (0, 0, 255)},
            'C': {'x': 6.29, 'y': 6.55, 'z': 1.2, 'enabled': True, 'color': (255, 0, 0)},
            'D': {'x': 6.16, 'y': 0, 'z': 1.2, 'enabled': True, 'color': (255, 255, 0)}
        }
        
        # Anchor colors for visualization
        self.anchor_colors = [
            (255, 0, 0),    # Red - A
            (0, 255, 0),    # Green - B
            (0, 0, 255),    # Blue - C
            (255, 255, 0),  # Yellow - D
            (255, 0, 255),  # Magenta - E
            (0, 255, 255),  # Cyan - F
            (255, 128, 0),  # Orange - G
            (128, 0, 128),  # Purple - H
            # Add more colors as needed
        ]
        
        self.initialize_anchors()

    def initialize_anchors(self):
        """Initialize anchor objects with default positions"""
        with self.anchor_lock:
            for i, (anchor_id, pos) in enumerate(self.anchor_positions.items()):
                self.anchors[anchor_id] = {
                    'x': pos['x'],
                    'y': pos['y'],
                    'z': pos['z'],
                    'enabled': pos['enabled'],
                    'color': self.anchor_colors[i % len(self.anchor_colors)],
                    'distance_history': deque(maxlen=100),
                    'last_update': 0
                }
            # Enable first anchor by default
            if self.anchors:
                first_anchor = next(iter(self.anchors))
                self.active_anchors.add(first_anchor)

    def update_anchor_position(self, anchor_id, x, y, z):
        """Update anchor position with thread safety"""
        with self.anchor_lock:
            if anchor_id in self.anchors:
                self.anchors[anchor_id]['x'] = x
                self.anchors[anchor_id]['y'] = y
                self.anchors[anchor_id]['z'] = z
                self.anchors[anchor_id]['last_update'] = time.time()
                return True
            return False

    def get_anchor_positions(self):
        """Get thread-safe copy of anchor positions"""
        with self.anchor_lock:
            return {k: v.copy() for k, v in self.anchors.items() if v['enabled']}

    def toggle_anchor(self, anchor_id, enabled):
        """Enable/disable an anchor"""
        with self.anchor_lock:
            if anchor_id in self.anchors:
                self.anchors[anchor_id]['enabled'] = enabled
                if enabled:
                    self.active_anchors.add(anchor_id)
                else:
                    self.active_anchors.discard(anchor_id)
                return True
            return False

    def calibrate_anchor(self, anchor_id, x, y, z):
        """Calibrate anchor position"""
        with self.anchor_lock:
            if anchor_id in self.anchors:
                self.anchors[anchor_id]['x'] = x
                self.anchors[anchor_id]['y'] = y
                self.anchors[anchor_id]['z'] = z
                return True
            return False


class DiagnosticsManager:
    """Manages time-series data for diagnostics plots."""
    def __init__(self, max_len=200):
        self.max_len = max_len
        self.data = {
            'time': deque(maxlen=max_len),
            'std_noise': deque(maxlen=max_len),
            'fp_power': deque(maxlen=max_len),
            'rx_power': deque(maxlen=max_len)
        }
        self.lock = threading.Lock()

    def update(self, rx_diag: dict):
        """
        Add a new diagnostics entry from parsed rx_diag.
        Use this method name consistently.
        """
        if not rx_diag:
            return

        with self.lock:
            current_time = time.time()
            self.data['time'].append(current_time)
            self.data['std_noise'].append(rx_diag.get('Std_noise', 0))
            self.data['fp_power'].append(rx_diag.get('Fp_power', -150.0))
            self.data['rx_power'].append(rx_diag.get('Rx_power', -150.0))

    def get_data(self):
        """Thread-safe copy of the data for plotting."""
        with self.lock:
            return {key: list(val) for key, val in self.data.items()}

    def clear(self):
        """Clear all stored data."""
        with self.lock:
            for key in self.data:
                self.data[key].clear()


class DataLogger:
    def __init__(self, max_records=10000):
        self.max_records = max_records
        self.data_lock = threading.Lock()
        self.range_analysis_tag = "0"
        self.recording = False
        
        # Initialize data tables
        self.tables = {
            '1': deque(maxlen=max_records),
            '2': deque(maxlen=max_records)
        }
        self.headers = ["Timestamp", "TagID", "X", "Y", "Z", "Status", "Dist_A", "Dist_B", "Dist_C", "Dist_D"]
        
    def _create_table_template(self):
        """Create template for trace tables"""
        columns = [
            'Time', 'x', 'y', 'z', 'A', 'B', 'C', 'D', 
            'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 
            'M', 'N', 'O', 'P', 'Flag', 'Velocity'
        ]
        return {col: deque(maxlen=self.max_records) for col in columns}
        
    def _create_analysis_table(self):
        """Create analysis-specific table"""
        columns = [
            'Time', 'AnchorID', 'Distance', 'MaxNoise', 'StdNoise',
            'FirstPathAmp', 'CIR', 'PreambleCount', 'RxPower'
        ]
        return {col: deque(maxlen=self.max_records) for col in columns}
        
    def log_trace_data(self, channel, tag, distances):
        """Logs a snapshot of tag data."""
        channel_str = str(channel)
        if channel_str not in self.tables:
            return
            
        with self.data_lock:
            self.tables[channel_str].append({
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                "TagID": tag.id, "X": tag.x, "Y": tag.y, "Z": tag.z, "Status": tag.status,
                "Dist_A": distances.get('A'), "Dist_B": distances.get('B'),
                "Dist_C": distances.get('C'), "Dist_D": distances.get('D')
            })
            
    def log_analysis_data(self, anchor_id, diagnostic_data):
        """Log anchor diagnostic data"""
        with self.data_lock:
            timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
            self.tables['analysis']['Time'].append(timestamp)
            self.tables['analysis']['AnchorID'].append(anchor_id)
            self.tables['analysis']['Distance'].append(str(diagnostic_data['distance']))
            self.tables['analysis']['MaxNoise'].append(str(diagnostic_data['Max_noise']))
            self.tables['analysis']['StdNoise'].append(str(diagnostic_data['Std_noise']))
            self.tables['analysis']['FirstPathAmp'].append(
                f"{diagnostic_data['Fp_amp1']}/{diagnostic_data['Fp_amp2']}/{diagnostic_data['Fp_amp3']}"
            )
            self.tables['analysis']['CIR'].append(str(diagnostic_data['Max_growthCIR']))
            self.tables['analysis']['PreambleCount'].append(str(diagnostic_data['Rx_preambleCount']))
            self.tables['analysis']['RxPower'].append(f"{diagnostic_data['Rx_power']:.2f}")
    
    def log_position_data(self, tag_id, x, y, z, velocity, status):
        """Log position data with timestamp"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
        with self.data_lock:
            self.position_log.append({
                'timestamp': timestamp,
                'tag_id': tag_id,
                'x': x,
                'y': y,
                'z': z,
                'velocity': velocity,
                'status': status
                })
            
    def log_ai_comparison(self, tag):
        """Records the raw vs ai position for error analysis."""
        with self.data_lock:
            # Calculate Euclidean Error: distance between Raw and AI
            raw_x = tag.history_raw['x'][-1] if tag.history_raw['x'] else tag.x
            raw_y = tag.history_raw['y'][-1] if tag.history_raw['y'] else tag.y
            
            error = math.sqrt((tag.x - raw_x)**2 + (tag.y - raw_y)**2)
    
            entry = {
                "Timestamp": datetime.now().strftime("%H:%M:%S.%f"),
                "TagID": tag.id,
                "Raw_X": round(raw_x, 3),
                "Raw_Y": round(raw_y, 3),
                "AI_X": round(tag.x, 3),
                "AI_Y": round(tag.y, 3),
                "Z": round(tag.z, 3),
                "Error_m": round(error, 4)
            }
            # We'll use table '1' for this session data
            self.tables['1'].append(entry)
            
    def get_dataframe(self):
        """Converts the collected data in table '1' to a Pandas DataFrame."""
        with self.data_lock:
            if not self.tables['1']:
                return pd.DataFrame()
            return pd.DataFrame(list(self.tables['1']))
            
    def export_to_csv(self, table_name, filename):
        """Export data table to CSV file"""
        if table_name not in self.tables:
            return False
            
        try:
            with self.data_lock:
                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)
                    # Write header
                    writer.writerow(self.tables[table_name].keys())
                    # Write data
                    for i in range(len(next(iter(self.tables[table_name].values())))):
                        row = [col_data[i] for col_data in self.tables[table_name].values()]
                        writer.writerow(row)
            return True
        except Exception as e:
            print(f"Export failed: {e}")
            return 
        
    def export_to_excel(self, channel):
        """Exports the data for a given channel to an Excel file."""
        channel_str = str(channel)
        if not self.tables[channel_str]:
            messagebox.showinfo("Export Empty", f"No data has been logged for Channel {channel_str} to export.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All files", "*.*")],
            title=f"Export Channel {channel_str} Data"
        )
        if not filepath:
            return

        try:
            with self.data_lock:
                # Convert deque of dicts to a DataFrame
                df = pd.DataFrame(list(self.tables[channel_str]))
            
            # Ensure proper column order
            df = df[self.headers]
            df.to_excel(filepath, index=False, engine='openpyxl')
            messagebox.showinfo("Export Successful", f"Data for Channel {channel_str} successfully exported to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"An error occurred during export: {e}")
        
    def setup_waveform_controls(self, parent):
        """Create waveform control panel"""
        control_frame = ttk.Frame(parent)
        
        # Channel selection
        self.channel_vars = {}
        for i in [1, 2]:
            ch_frame = ttk.LabelFrame(control_frame, text=f"Channel {i}")
            ttk.Label(ch_frame, text="Tag ID:").pack()
            var = tk.StringVar(value=str(i-1))
            entry = ttk.Entry(ch_frame, textvariable=var, width=5)
            entry.bind("<Return>", lambda e, ch=i: self.set_channel_tag(ch))
            entry.pack()
            self.channel_vars[i] = var
        
        # View controls
        btn_frame = ttk.Frame(control_frame)
        ttk.Button(btn_frame, text="Reset View", 
                  command=self.reset_waveform_view).pack(pady=2)
        ttk.Button(btn_frame, text="Pause", 
                  command=self.toggle_pause).pack(pady=2)
        
        return control_frame
    
    def set_channel_tag(self, channel):
        """Change which tag is displayed on a channel"""
        tag_id = self.channel_vars[channel].get()
        if tag_id in self.tag_manager.tags:
            self.active_tags[channel] = tag_id
            self.clear_channel_data(channel)
            
    def reset_waveform_view(self):
        """Reset the waveform view to default"""
        # Add your implementation here
        print("Waveform view reset")
    
    def toggle_pause(self):
        """Toggle pause state of waveform updates"""
        # Add your implementation here
        print("Waveform pause toggled")
    
    def clear_channel_data(self, channel):
        """Clear data for a specific channel"""
        # Add your implementation here
        print(f"Channel {channel} data cleared")
    
    def start_range_recording(self, tag_id):
        """Start recording range analysis data"""
        if tag_id in self.tag_manager.tags:
            self.range_analysis_tag = tag_id
            self.recording = True
            return True
        return False
        
    def stop_range_recording(self):
        """Stop recording"""
        self.recording = False
        
    def export_range_data(self, filename=None):
        """Export recorded range data to Excel"""
        if not filename:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Range Analysis Data")
            if not filename:
                return False
                
        try:
            # Create DataFrame from analysis table
            df = pd.DataFrame({
                'Time': list(self.tables['analysis']['Time']),
                'AnchorID': list(self.tables['analysis']['AnchorID']),
                'Distance': list(self.tables['analysis']['Distance']),
                'FP_Power': list(self.tables['analysis']['FirstPathAmp']),
                'RX_Power': list(self.tables['analysis']['RxPower'])
            })
            
            df.to_excel(filename, index=False)
            return True
        except Exception as e:
            print(f"Export error: {e}")
            return False
      
        
class EnhancedDataLogger(DataLogger):
    def __init__(self):
        super().__init__()
        # Add analysis-specific columns
        self.tables['analysis']['columns'].extend([
            'Cal_Flag', 'Max_noise', 'Std_noise', 'Fp_amp1', 'Fp_amp2', 'Fp_amp3',
            'Max_growthCIR', 'Rx_preambleCount', 'Fp', 'Fp_power', 'Rx_power'
        ])
        
    def save_tag_data(self, tag_id, x, y, z, dists, cal_flag, velocity, mode=1):
        """Save tag data with enhanced fields from C# code"""
        table_name = 'trace1' if mode == 1 else 'trace2'
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        
        with self.data_lock:
            self.tables[table_name]['Time'].append(timestamp)
            self.tables[table_name]['x'].append(x)
            self.tables[table_name]['y'].append(y)
            self.tables[table_name]['z'].append(z)
            
            for i in range(16):  # ANCHOR_MAX_COUNT
                anchor_id = chr(65 + i)  # A-P
                self.tables[table_name][anchor_id].append(dists[i] if i < len(dists) else 0)
                
            self.tables[table_name]['Flag'].append(cal_flag)
            self.tables[table_name]['Velocity'].append(velocity)
       